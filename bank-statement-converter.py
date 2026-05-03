"""
BCA Bank Statement Converter
=============================
Converts BCA PDF bank statements to Excel, XML, JSON, CSV, TSV,
Markdown, YAML, RTF, ODT, or Parquet.

Requirements:
    pip install pdfplumber openpyxl python-dotenv

Optional (for extra export formats):
    pip install pyyaml odfpy pandas pyarrow

Run:
    python bca_converter.py

Compile to .exe:
    pyinstaller --onefile --windowed --name "BankStatementConverter_v1.4.0" bank-statement-converter.py

"""

import re
import sys
import os
import json
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# Load environment variables from .env file
from dotenv import load_dotenv
load_dotenv()

EXCHANGE_RATE_API_KEY = os.getenv("EXCHANGERATE_API_KEY")

# Self-reference that works both as a script and inside a PyInstaller .exe
# (where __name__ == '__main__' and 'bca_converter' is not in sys.modules)
import sys as _sys
_this_module = _sys.modules[__name__]

# Resource path helper for PyInstaller --onefile bundles
def _resource_path(relative):
    """Return absolute path to a resource, works for dev and PyInstaller."""
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative)

__version__ = "1.4.0"

# ── Embedded logo (base64 PNG) ──────────────────────────────────────────────
LOGO_B64 = (
    "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZ"
    "WiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAA"
    "ACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAA"
    "AChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAA"
    "AAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAA"
    "AAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAA"
    "E9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBu"
    "AGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQa"
    "FRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4e"
    "Hh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAIAAgADASIAAhEBAxEB/8QAHQABAQAD"
    "AQADAQAAAAAAAAAAAAgGBwkFAQMEAv/EAE4QAAEDAgMFAwgECgcHBQEAAAABAgMEBQYRIQcSMUFRE2Fx"
    "CBQYIiRmpeMyQoGRFlJlcoKhpLHR4hUjNKKjweEXM2NkwsPwU4OSsrOE/8QAGwEBAAIDAQEAAAAAAAAA"
    "AAAAAAQFAgMGAQf/xAA0EQEAAQIDBgUCBQUBAQEAAAAAAQIDERNhBAUVYqHhEiExQXEUUQZCgcHRIjKR"
    "sfBS8WP/2gAMAwEAAhEDEQA/AIyAAAAAAAAAAAAAAAB/cMUk0rYomOfI9cmtRM1VTN7HhCniY2a5r20q"
    "69k1cmt8V5/uGALU2Kl/pOZucsuaRZ/VbzX7f3eJlZLs2Yw8VSJevTj4aX009LTU7UbBTxRInJjEQ+4A"
    "lIwAA8AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA+FRHJkqIqd58gDz66y2utaqTUcW8"
    "v12Juu+9DDcRYWnt7HVNI51RTpq5FT12J39U7zYQNddqmpsou1UtNAyPG1mbb6ptXTM3aaZclanBjung"
    "v8THCBVTNM4Sn01RVGMAAMWQAAAAAAAAAAAAAAAAAAAAA2/Qwtp6OGBqZJHG1qfYh9wBaqtb2xrZnZMF"
    "4bopJLfBNfJYmyVdXIxHSNeqZqxir9FqcNMs8s1NiH4ZpHPcuvq8kPrIUxM+cpkTEeUPSB5oPPC98T0g"
    "eaB4TxPSB5oHhPE9IHmgeE8T0geaB4TxPSB5oHhPE9IHmgeE8T0geaB4TxPSB5oHhPE9IHmgeE8T0gea"
    "B4TxPSB5oHhPE9IHmgeE8T0geaB4TxPSB5oHhPE9IHmgeE8T0geaB4TxIc2wYgkxNtKvl1dJvxLVOip+"
    "iRRruMy6Zo1F8VUxIJomQJsRhGCFM4zioLYfsrtq2emxLiWkZVz1LUkpaWVM4441+i5zeDlVNURdERU5"
    "8NzeZUXmvmvmlP5vll2XZpuZdMuB+GeZKeJlFS+pFCxGJlpoiZZIfk3nZ728ufXM3Rbxhzt/bMa592st"
    "uOyy2pZ6jEuG6RlJPTNWSqpYm5RyRp9JzWpo1UTVUTRURefGfS2oJkqIn0VV/WRTMWNc9dFTLJSJTCqn"
    "wys9hv5tE6POxJSNrbJVQqmbtxXM/OTVP3GqzcjkRzVavBUyNNkLaY84ldbNPlMAAIqSAAAAAAAAAAAA"
    "AAAAAAAAANygAtVU6IAAiJYAAAAAAAAAAAAAH01tXS0NJJV1tTDTU0Td6SWZ6MYxOquXRENd7WNr9iwS"
    "klvpkbdL3l/ZY35Mhz5yu5dd1NV04IuZLuOccYlxnWrUX24vljR29FSx5sgi4/RZwz1yzXN2XFVM6aJl"
    "hVXEKMxp5QGErM99PZIJ79UtXJXRL2UHHJU7RUVV8WtVF6mpMRbe8f3Nytoqiis8WqbtLTo5yovV0m9r"
    "3t3TVQNsW4hqm5MvZuuLMUXVisueI7vWRque5NWSPang1VyQ8ZURVzXVeoBlgxxkTRc00Xqe1asWYptT"
    "EZbcR3ejYi57kNZI1q+LUXJTxQMDFtXDu3zH1scja6oorxFom7VU6NciJ0dHu697t423gzygcI3h7Ke9"
    "wVFhqHLkjpF7aBVzyRN9qIqeLmoidSTwYzREsouTDoVRVVLXUkdXRVMNTTyt3o5YXo9j06oqaKh9xCeB"
    "scYlwZW+cWK4vijc5HS00nrwS8PpM4Z6ZZpk7LgqFQ7J9r9ixt2dvqUba73l/ZZH5smVOKxO59d1dU14"
    "oiqaqqJhtpriWywAYMwAAAAAAAAAAAABzvABLRFnquaqq8VPgAluMfKKqKipxQjAs4jE03fZdbo/P+n7"
    "hpo3KaaK/afZ0mze4ACKlAAAAAAAAAAAAAAAAAAAAADcoALVVOiAAIiWAAAAAAAAAAAaF29bZltklRhb"
    "CFSnnqZx1texc+wXgsca/j8ld9Xgnrat9DykNp78O0bsK2Cq3LvUx51U8a+tSROTREXlI5OHNE10VWqS"
    "ybKKMfOWuuvDyh/Uj3ySOkke573KrnOcuaqq8VVT+T5Y1z3oxjVc5y5IiJmqqUBsh2FpLHDe8bxORrk3"
    "4rZmrVy5LKqap+Yn280MNq2u1stHiuT+nvLyxYrvVYUtSYIwHijGM+7ZLZI+BFyfVS+pCzxcvFe5M17j"
    "d+EvJ2tFMjJsTXaevlyzWClTsokXorlzc5O9N03ZTQ0lBSx0tLDFTwRNRscUTEa1qJyRE0RD4dUL9VMv"
    "E5HbfxBdqnCmfDGnr/n/AOL3Z92W6fOqMZY9ZdnWB7QxraLC9sRW8HzQpM9P0n5r+syOnpaanajIKeGJ"
    "qcEYxGon3H0rI9eLlP5VVXiqlDc3jNc4zjPzKxp2eKYwjyfpqKamqGq2op4pWrxR7Ecn6zG7zs6wNd2O"
    "bW4Xtiq7i+GFIXr+kzJf1nt5qnNT+kkenByi3vGaJxjGPiSrZ4qjCfNpjFvk7Wipa+bDN2noJeKQVX9b"
    "Evcjkyc3xXeNIY2wHijB027e7Y+OBXbrKqP14X+Dk4L3LkvcW02oX6yZ+B81MNJX0klLVQxVEErVbJFK"
    "xHNci8UVF0VC+2L8QXaZwqnxRr6/5/8Aqu2jdlurziMJc/j+o3vjkbJG5zHtVHNc1clRU4Kim/tr+wzs"
    "Y5r3gmJzmNRXzWzNVVE5rEvFfzF+zkhoBzVa5WuRUci5Kipqh12y7Xa2qjxW57KK9Yrs1YVKV2C7Zluc"
    "lPhbF9Snnrso6K4PXLt14JHIv4/JHfW4L62rt9HO8qbyb9p78R0aYVv9Vv3emjzpZ5F9ariRNUVecjU4"
    "81TXVUcpnXRh5w9orx8pbqABrbAAAAAAAAAAAc7wAS0RZwAJbjAjEs4jE03fZdbo/P8Ap+4aaNymmiv2"
    "n2dJs3uAAipQAAAAAAAAAAAAAAAAAAAAA3KAC1VSiPSa9yfivyR6TXuT8V+STuDDLpZ5lSiPSa9yfivy"
    "R6TXuT8V+STuBl0mZUoj0mvcn4r8ky7ZntyseLrzHZrhb32SuqHbtNvzpLFK7kzf3W5OXkipkvDPNURZ"
    "JPuofOvPYPMe1867RvYdlnv7+fq7uXPPLI8m3S9i5Vi6FAA0N4Y1tNxbS4KwdWX2oRr5WJ2dLCq5dtM7"
    "6LfDiq5a7qKvIyUlHyqMWrecasw9Szb1FZ27r0auj6hyIr10XXdTdbrqio/qZU04zgxqqwjFqe7XCsut"
    "zqblcJ3VFXUyulmkdxc5VzVe7wTRD8oNs+TdgNmJ8SuvdyhR9qtb0XccmbZp+LW96J9Jf0U4KZbRfp2e"
    "1Nyr0hrtWqrtcUx7s/8AJ62VR2qmgxZiOmR1xlaklFTSN/szV4Pci/XXl+Knfw3TLNl6rPvPieXP1Wrp"
    "zU+k+abx3lc2m5NWP/faHW7NstNmiIgVVVc11APx3q6W+zWye5XSqjpaSBu9JI9dE/iq8kTVSqppmqcI"
    "85lKmYiMZfsNC+Urj262+5RYUs9VLRsWBJqyWJyte/ezyYipqiZJmuXHNE65/ixn5QFdLNJT4Ut0UEKK"
    "qJVVabz3d6MRcm/bn4IaexJfbriO7SXW81S1VZIiNdJuNboiZImTURDstybgvW70XtppjCPSJ9cfhRbw"
    "3lRVRNu1Pn93xZb7eLLcW3C13Kppalrt7fZIvrfnJwcncuaFibMsSLizBNvvckbY55mK2djeCSNVWuy7"
    "lVM07lIqN1bDtq9ow1Z4cNXullhp0mc9lbGu+jd5c/XbxRE6pn4Fl+It3TtNiK7VGNcT7euH7ou69qi1"
    "cmmurCJ/2pAIqouaH1008NTTx1FPKyWGVqPjkY7NrmqmaKi80PsPnXnEuoffFNn6r/vNK+UNsrjudNPi"
    "3DtOja+JqyV1NG3+0NTVZGon105p9ZO/juM+6CXL1XLpyUtt3byubNciqJ/77SibTstN6iYlz/P1Wm4V"
    "lpudNc7fO6nq6WVssMjeLXIuaePguimzfKPwG3C+JW3m2wJHabm5XI1qerDNxczuRfpJ+kiaIaoPpez3"
    "6NotRcp9Jcldt1Wq5pn2XXszxZS41wdRX2nRrJJE7OphRc+xmb9NvhzTPXdVF5mSko+Svi1bNjV+Hqqb"
    "dorw3djRy6MqGoqtXVdN5N5umqrudCrjGqnCWymrGMQAGLJqTaZtzseErzJZrfb33uup3btTuTpFFE7m"
    "zf3XZuTmiJknDPNFRMS9Jr3J+K/JJ9rfOvPZ/Pe1867R3bdrnv7+frb2fPPPM+k3xbpwaJuVYqI9Jr3J"
    "+K/JHpNe5PxX5JO4Pcul5mVKI9Jr3J+K/JHpNe5PxX5JO4GXSZlQADNgs4AEtxgRiWcRiabvsut0fn/T"
    "9w00blNNFftPs6TZvcABFSgAAAAAAAAAAAAAAAAAAAABuUAFqqlEejL77fCvnD0Zffb4V84ogEbMqScu"
    "lO/oy++3wr5w9GX32+FfOKIAzKjLpTv6Mvvt8K+cZdsz2G2PCN5jvNwuD73XQO3qbfgSKKJ3J+5vOzcn"
    "JVXJOOWaIqbaAmuqSKIgABiyefia6w2LDtxvM7d+OhpZKhzc8lduNVd1O9csvtIGuFXUV9fUV9ZKstTU"
    "yummkXi97lVXL9qqpWvlTXR1v2Tz0zW5rcayGlzz+iiKsqr/AIWX2khm61Hli03Z88H9RsfJI2ONqve5"
    "URrUTNVVeRbuznDkWD8D26xsRvbsj36lyfWldq9fv0TuRCW9gllbe9qloikYjoaV61cufSNN5v8Ae3U+"
    "0sGV2+9V+45P8U7ZNMU2afn/AL/vdc7nsRONyfh/IAOIX4Sz5RuMp77i6Ww00ypbbU9Y9xF0knTR7l8F"
    "zangvUqYg66TS1Fzqp5lVZZJnvfnx3lcqqdV+FNmouX67tX5YjD9f/im3zdqpt00R7/s/MAVlgHZhs/g"
    "w9R1UNBS3tZ4mvWsn/rElVU4o1fVandlmnPNTq95b0tbvoiq5Ezj9lNsmx17VVMUzhgk0FZYt2MYLvVM"
    "5KKi/oery9Sal+jn3sVd1U8Ml7zCLV5O723BrrpiRj6Nrs1bT06tkenTNVyb+shWfxLsNyiaqpmmftMf"
    "xikV7p2imrCIxZt5N09ZNsro/OlcrI55WQK7nGjv3I7eT7DZJ+Sz26itFrprZboGwUlNGkcUbeCIn717"
    "+Z+s+f7Zei/fru0xhEzMulsW5t26aJ9oAAR214e0fDkWMcDXGySInbuj36Zy/Vlbqxfv0XuVSInscx7m"
    "ParXNXJyKmSovQvuJ249F+8j3b1ZW2PaneIY49yCpkSriy4Kkibzsv0t5PsO4/C22TXFVmr5/wC/72c/"
    "vixhhcj4YZb6uot9wp6+jlWKpppWzQvTi17VRWr9iohfOGrrDfMPW68wN3Iq6ljqGtzzVu+1Hbq96Z5f"
    "Yc/yvPJaui3DZNBTOauduq5qXNV+kiqkqfqly+w6u7Hlip7U+eDagANLc1LtM2G2PF15kvNvuD7JXVDt"
    "6p3IElildzfubzcnLzVFyXjlmqquI+jL77fCvnFEAyiuqGM0Uynf0Zffb4V84ejL77fCvnFEAZlRl0p3"
    "9GX32+FfOHoy++3wr5xRAGZUZdLneACSjLOABLcYEYlnEYmm77LrdH5/0/cNNG5TTRX7T7Ok2b3AARUo"
    "AAAAAAAAAAAAAAAAAAAAAblABaqpRHpNe5PxX5I9Jr3J+K/JJ3Bhl0s8ypRHpNe5PxX5Jl2zPblY8XXm"
    "OzXC3vsldUO3abfnSWKV3Jm/utycvJFTJeGeaoiySfdQ+deeweY9r512jew7LPf38/V3cueeWR5Nul7F"
    "yrF0KABobwAAT95ZVVIy34ZokcvZyy1Mrk6qxI0T/wC6k4FD+Wdxwn//AGf9gngkW/7Ue5/c3d5I9Gjs"
    "SXy4qmsFEyFF6do/P/tlGE/eSK9EqcSR83Mpl+5Zf4lAnzj8SVzVt9UfaI/1Dqd1Rhs1M/P+wAFCsQjf"
    "bLhubDO0C40qxq2mqZFqqV2WixvVVyTwXNv2FkGK7ScEWrHFlShrlWGoiVXUtUxuboXL3c2rpmnPuVEU"
    "udybyjYNoxr/ALavKf5QN4bJO028KfWPRF57WGsVYiw3Ir7JeKqiRVzcxj843L1Vi5tX7UPTxzs8xPhC"
    "d/8ASNA+WjRfUrIEV8Lk6qv1V7nZGJH0imuztVvGMKqZ/WHLTFyzVhOMS3NhzygcQUqsjvlro7jGmiyQ"
    "qsMnjzav3IbgwLtMwpi9zaehrVpq5U/slUiMkX83XJ32Kq9yEcn9RvfHI2SNzmPaqK1zVyVFTgqKU22f"
    "hzY78TNEeCdPT/H8YJ1jet+3P9U+KNf5XwDT/k+7SZ8SQOw5fJu0ulNHvwTuXWojTjvdXp15pryVTcBw"
    "O2bJc2O9Nq56x11dJYv036Irp9AAEVuCc/K4o0biSx3FE9aeifCv/tvz/wC4UYT75XUiLU4bj5tZUqv2"
    "rF/Avvw3XNO30x94n/Uq7esY7NVPx/tocpDyNauR9uxLQqq9nDNTytTor0kRf/zQm8ofyMfpYs8KP/vn"
    "0e5/a5a3/cokAEdIAABqTaZtzseErzJZrfb33uup3btTuTpFFE7mzf3XZuTmiJknDPNFRMS9Jr3J+K/J"
    "J9rfOvPZ/Pe1867R3bdrnv7+frb2fPPPM+k3xbpwaJuVYqI9Jr3J+K/JHpNe5PxX5JO4Pcul5mVAAM2C"
    "zgY9+E3/ACX+L/oPwm/5L/F/0JPjhx3gqZCRiVb+E3/Jf4v+hKRquTE4YLndNMx48dP3DTRuU00QNp9n"
    "R7N7gAIqUAAAAAAAAAAAAAAAAAAAAANygAtVUoj0Zffb4V84ejL77fCvnFEAjZlSTl0p39GX32+FfOMu"
    "2Z7DbHhG8x3m4XB97roHb1NvwJFFE7k/c3nZuTkqrknHLNEVNtATXVJFEQAAxZAAAn/yyqWR9uw1XI1e"
    "zilqInLyRXpGqf8A0Um8rzyprW64bJ56lq626rhqskT6SKqxL/8Arn9hIZvtz5NFyPNuDyWa9tPi+vpH"
    "Ll29M13/AMXZf9aFNEZbIrolq2gW2Vz1bHO9ad/6aZN/vbq/YWLb50qaRkueqpk7x5nz/wDFFiaNs8f/"
    "AKiP4dJui54rHh+0v0AA5tagBh21rGsWB8LrcWwtqK2d/Y0kLl9VX5Ku87LXdREzX7E0zzNtizXfuRbt"
    "xjMsLlym3TNVXpDMHIjmq1yIqKmSovMw/EmzHBF+3nVdhp4Zna9tSp2L8+q7uSKviik5P2v7QnV/naX9"
    "zfWz7JII+zROm7u8P195tDCnlA2mWiZHiW2VVPVtTJ0tI1HxP78lVFb4a+Jf17i3lsWFyzOM8szj+ytp"
    "3jst/wDprj/MPB2gbB6i20FRcsMXCStjhasjqSoana7qaruuTRy92SfauhpEo7Fe3+xstk0WHrfW1Fa9"
    "itjkqGNjjYqp9LRVVcumSeJOJ1W5K9uqtT9ZHxj6/qp94U7PFcZE/P2ezgi7y2HF1qu8TlatNVMc7Lmz"
    "PJyfa1VT7S4SErDQS3S+UFtharpKqojhaidXORP8y7Sh/F0U5lqY9cJ/x5YfusdyTPhrj28gAHHrwJl8"
    "qavbU4voKRrs+wpnO8N52X/QUlcJ0pqR8ueqJk3x5EdbXbol2x/cpWP3o4HpTsX8xMl/vbx0n4XsTXtm"
    "Z/5if4VW97nhseH7yxIpDyNaSRluxLXK1ezmmp4mu6qxJFVP8RPvJvK88lq1ut2yanqXOXO41c1Vuqn0"
    "URUiT70iz+0+gXJ8nN24821AAaG8AAGpdpmw2x4uvMl5t9wfZK6odvVO5AksUrub9zebk5eaouS8cs1V"
    "VxH0Zffb4V84ogGUV1QxmimU7+jL77fCvnD0Zffb4V84ogDMqMulzvABJRmxv9p/5D/a/wCQf7T/AMh/"
    "tf8AIa5B7jKL9FY/89ZbG/2n/kP9r/kNcgHmLbasUWsfBGGIaaNxzSNiifK9cmsarnL0RDThE2n2WGze"
    "4ACKlAAAAAAAAAAAAAAAAAAAAADcoALVVOiAAIiWAAAAAAAAAAAaF29bZltklRhb"
    "CFSnnqZx1texc+wXgsca/j8ld9Xgnrat9DykNp78O0bsK2Cq3LvUx51U8a+tSROTREXlI5OHNE10VWqS"
    "ybKKMfOWuuvDyh/Uj3ySOkke573KrnOcuaqq8VVT+T5Y1z3oxjVc5y5IiJmqqUBsh2FpLHDe8bxORrk3"
    "4rZmrVy5LKqap+Yn280MNq2u1stHiuT+nvLyxYrvVYUtSYIwHijGM+7ZLZI+BFyfVS+pCzxcvFe5M17j"
    "d+EvJ2tFMjJsTXaevlyzWClTsokXorlzc5O9N03ZTQ0lBSx0tLDFTwRNRscUTEa1qJyRE0RD4dUL9VMv"
    "E5HbfxBdqnCmfDGnr/n/AOL3Z92W6fOqMZY9ZdnWB7QxraLC9sRW8HzQpM9P0n5r+syOnpaanajIKeGJ"
    "qcEYxGon3H0rI9eLlP5VVXiqlDc3jNc4zjPzKxp2eKYwjyfpqKamqGq2op4pWrxR7Ecn6zG7zs6wNd2O"
    "bW4Xtiq7i+GFIXr+kzJf1nt5qnNT+kkenByi3vGaJxjGPiSrZ4qjCfNpjFvk7Wipa+bDN2noJeKQVX9b"
    "Evcjkyc3xXeNIY2wHijB027e7Y+OBXbrKqP14X+Dk4L3LkvcW02oX6yZ+B81MNJX0klLVQxVEErVbJFK"
    "xHNci8UVF0VC+2L8QXaZwqnxRr6/5/8Aqu2jdlurziMJc/j+o3vjkbJG5zHtVHNc1clRU4Kim/tr+wzs"
    "Y5r3gmJzmNRXzWzNVVE5rEvFfzF+zkhoBzVa5WuRUci5Kipqh12y7Xa2qjxW57KK9Yrs1YVKV2C7Zluc"
    "lPhbF9Snnrso6K4PXLt14JHIv4/JHfW4L62rt9HO8qbyb9p78R0aYVv9Vv3emjzpZ5F9ariRNUVecjU4"
    "81TXVUcpnXRh5w9orx8pbqABrbAAAAAAAAAAAc7wAS0RZwAJbjAjEs4jE03fZdbo/P8Ap+4aaNymmiv2"
    "n2dJs3uAAipQAAAAAAAAAAAAAAAAAAAAA3KAC1VSiPSa9yfivyR6TXuT8V+STuDDLpZ5lSiPSa9yfivy"
    "R6TXuT8V+STuBl0mZUoj0mvcn4r8ky7ZntyseLrzHZrhb32SuqHbtNvzpLFK7kzf3W5OXkipkvDPNURZ"
    "JPuofOvPYPMe1867RvYdlnv7+fq7uXPPLI8m3S9i5Vi6FAA0N4Y1tNxbS4KwdWX2oRr5WJ2dLCq5dtM7"
    "6LfDiq5a7qKvIyUlHyqMWrecasw9Szb1FZ27r0auj6hyIr10XXdTdbrqio/qZU04zgxqqwjFqe7XCsut"
    "zqblcJ3VFXUyulmkdxc5VzVe7wTRD8oNs+TdgNmJ8SuvdyhR9qtb0XccmbZp+LW96J9Jf0U4KZbRfp2e"
    "1Nyr0hrtWqrtcUx7s/8AJ62VR2qmgxZiOmR1xlaklFTSN/szV4Pci/XXl+Knfw3TLNl6rPvPieXP1Wrp"
    "zU+k+abx3lc2m5NWP/faHW7NstNmiIgVVVc11APx3q6W+zWye5XSqjpaSBu9JI9dE/iq8kTVSqppmqcI"
    "85lKmYiMZfsNC+Urj262+5RYUs9VLRsWBJqyWJyte/ezyYipqiZJmuXHNE65/ixn5QFdLNJT4Ut0UEKK"
    "qJVVabz3d6MRcm/bn4IaexJfbriO7SXW81S1VZIiNdJuNboiZImTURDstybgvW70XtppjCPSJ9cfhRbw"
    "3lRVRNu1Pn93xZb7eLLcW3C13Kppalrt7fZIvrfnJwcncuaFibMsSLizBNvvckbY55mK2djeCSNVWuy7"
    "lVM07lIqN1bDtq9ow1Z4cNXullhp0mc9lbGu+jd5c/XbxRE6pn4Fl+It3TtNiK7VGNcT7euH7ou69qi1"
    "cmmurCJ/2pAIqouaH1008NTTx1FPKyWGVqPjkY7NrmqmaKi80PsPnXnEuoffFNn6r/vNK+UNsrjudNPi"
    "3DtOja+JqyV1NG3+0NTVZGon105p9ZO/juM+6CXL1XLpyUtt3byubNciqJ/77SibTstN6iYlz/P1Wm4V"
    "lpudNc7fO6nq6WVssMjeLXIuaePguimzfKPwG3C+JW3m2wJHabm5XI1qerDNxczuRfpJ+kiaIaoPpez3"
    "6NotRcp9Jcldt1Wq5pn2XXszxZS41wdRX2nRrJJE7OphRc+xmb9NvhzTPXdVF5mSko+Svi1bNjV+Hqqb"
    "dorw3djRy6MqGoqtXVdN5N5umqrudCrjGqnCWymrGMQAGLJqTaZtzseErzJZrfb33uup3btTuTpFFE7m"
    "zf3XZuTmiJknDPNFRMS9Jr3J+K/JJ9rfOvPZ/Pe1867R3bdrnv7+frb2fPPPM+k3xbpwaJuVYqI9Jr3J"
    "+K/JHpNe5PxX5JO4Pcul5mVAAM2CzgY9+E3/ACX+L/oPwm/5L/F/0JPjhx3gqZCRiVb+E3/Jf4v+hKR"
    "quTE4YLndNMx48dP3DTRuU00QNp9nR7N7gAIqUAAAAAAAAAAAAAAAAAAAAANygAtVUoj0Zffb4V84ejL"
    "77fCvnFEAjZlSTl0p39GX32+FfOMu2Z7DbHhG8x3m4XB97roHb1NvwJFFE7k/c3nZuTkqrknHLNEVNtA"
    "TXVJFEQAAxZAAAn/yyqWR9uw1XI1ezilqInLyRXpGqf8A0Um8rzyprW64bJ56lq626rhqskT6SKqxL/8A"
    "Arn9hIZvtz5NFyPNuDyWa9tPi+vpHLl29M13/AMXZf9aFNEZbIrolq2gW2Vz1bHO9ad/6aZN/vbq/YWL"
    "b50qaRkueqpk7x5nz/wDFFiaNs8f/qI/h0m6LniseH7S/QADm1qAGHbWsaxYHwutxbC2orZ39jSQuX1V"
    "fkq7zstd1ETNfsTTPM22LNd+5Fu3GMywuXKbdM1VekMwciOarXIioqZKiczD8SbMcEX7edV2Gnhmdr21K"
    "nZvz6ru5Iq+KKTk/a/tCdX+dpf3N9bPskgj7NE6bu7w/X3m0MKeUDaZaJkeJbZVU9W1MnS0jUfE/vyVU"
    "Vvhr4l/XuLeWxYXLMYzyzOP7K2neOy3/APprj/MPB2gbB6i20FRcsMXCStjhasjqSoana7qaruuTRy92S"
    "fauhpEo7Fe3+xstk0WHrfW1Fa9itjkqGNjjYqp9LRVVcumSeJOJ1W5K9uqtT9ZHxj6/qp94U7PFcZE/P"
    "2ezgi7y2HF1qu8TlatNVMc7LmzPJyfa1VT7S4SErDQS3S+UFtharpKqojhaidXORP8y7Sh/F0U5lqY9c"
    "J/x5YfusdyTPhrj28gAHHrwJl8qavbU4voKRrs+wpnO8N52X/QUlcJ0pqR8ueqJk3x5EdbXbol2x/cpW"
    "P3o4HpTsX8xMl/vbx0n4XsTXtmZ/5if4VW97nhseH7yxIpDyNaSRluxLXK1ezmmp4mu6qxJFVP8AET7y"
    "byvPJatbrdsmp6lzlzuNXNVbqp9FEVIk+9Is/tPoFyfJzduPNtQAGhvAABqXaZsNseLrzJebfcH2SuqH"
    "b1TuQJLFK7m/c3m5OXmqLkvHLNVVcR9GX32+FfOKIBlFdUMZoplO/oy++3wr5w9GX32+FfOKIAzKjLpc"
    "7wASUZsb/af+Q/2v+Qf7T/yH+1/yGuQe4yi/RWP/AD1lsb/af+Q/2v8AkNcgHmLbasUWsfBGGIaaNxzS"
    "NiifK9cmsarnL0RDThE2n2WGze4ACKlAAAAAAAAAAAAAAAAAAAAADcoALVVKI9Jr3J+K/JHpNe5PxX5J"
    "O4MMulnmVKI9Jr3J+K/JHpNe5PxX5JO4GXSZlSiPSa9yfivyTLtme3Kx4uvMdmuFvfZK6odu02/OksUr"
    "uTN/dbk5eSKmS8M81RFkk+6h8689g8x7XzrtG9h2We/v5+ru5c88sjybdL2LlWLoUADQ3vPxLaob5h64"
    "2aoduxV1NJTuciZq3faqbyd6Z5p4EDXCkqKCvqKCriWKpppXQzRrxY9qqjk+xUU6Eko+VRhJ1mxqzEVL"
    "FlRXhu89WpkjKhqIjk0TTeTddrqq7/QNlufPBruRjGLTzHOY9r2OVrmrmiouSovUrrZTiiO94doq9z0z"
    "nYjJ0/Embo7wTP8AUqEiGwtiWLW2K+La66Xdt9e5E3lXSKXg13ci8F+xeRU7/wBg+q2fxUx/VT5/p7pe"
    "7NpybuE+kq4B5NluCSNbTTu9dNGOX63d4nrHzWqmaZwl1UTiGovKhw/XXXCFFc6KJ8yWyZz52MTNUjci"
    "Ir8u5Wpn3Kq8jbp8KiKmSpmhJ2Laqtkv03qYxwar9mL1ubc+6BgVRjrYlhm/zSVlse+yVj1zd2LEdC5e"
    "qx6ZfoqidxLdRGsNRJCq5qx6tz65LkfTN3b0sbwpmbfrHrE+zk9q2O5s04V+kvrAKdwBsQw3a0p7leZn"
    "3mo3WyNjkYjIGqqZ6szXe+1cu493hvOxsFEVXfWfSI93my7Jc2mrCj2Yt5NeAKl9ezGd2gdHTxNVLex6"
    "ZLI5UyWXL8VEzROqrny1oY+GNa1qMY1GtamSIiZIiHyfNt47fc26/N2v9I+0Or2XZqdntxRSAHk3q4Ix"
    "rqeB3rro9ycu7xIVNM1ThCRM4MW2rYojshHq2va9ucDFZAi/Xmdo3x1/Uikivc573Pe5XOcuaqq6qpsL"
    "bbi1t9vjbXQyo+30DlTeauksvBzu9E4J9vU14fStwbB9Ls3iqj+qrz/T2cpvPac67hHpD77fSVNwr6eg"
    "o4llqamVsMMacXvcqI1PtVUL5w3aobHh+3WanduRUMNNHTtcqZK5GNRN5e9cs18SYvJWwk68Y1fiKpi"
    "zornbniS5NH1DkVGpqmu6m87qi7nUq0trk4zgiW4wjEABrbGpNpm3Ox4SvMlmt9vfe66ndu1O5OkUUTu"
    "bN/ddm5OaImScM80VExL0mvcn4r8kn2t8689n897XzrtHdt2ue/v5+tvZ8888z6TfFunBom5Vioj0mv"
    "cn4r8kek17k/Ffkk7g9y6XmZUoj0mvcn4r8kek17k/Ffkk7gZdJmVAAM2DDfw5/Jf7R/KPw5/Jf7R/K"
    "YYCvz6/usMij7Mz/Dn8l/tH8o/Dn8l/tH8phgGfX9zIo+z377iisuUC0zI200LvpI12au7lXoeAAYVV"
    "TVOMs6aYpjCAAGLIAAAAAAAAAAAAAAAAAAAAAblABaqpRHoy++3wr5w9GX32+FfOKIBGzKknLpTv6Mvv"
    "t8K+cPRl99vhXziiAMyoy6U7+jL77fCvnGXbM9htjwjeY7zcLg+910Dt6m34Eiiidyfubzs3JyVVyTjl"
    "miKm2gJrqkiiIAAYsgxraZhOlxrg6ssVSrWSPTtKWZU/3MzfoO8OS5aq1VTmZKAOfN2t9ZarnU224U76"
    "erppHRTRu4tci5L4+KaKflKn8pDZg/EdG7FVgpd+70seVVBG31qqJE4onORqcOappqqNQlgk01eKEaqn"
    "wy3lsa2gNr4YcPXmfdrY0RtLO93++RODVX8ZOXXx47ttt2yRIqpVVOCSfxIhaqtcjmqqKi5oqcjcGzja"
    "v2bI7XimRzkTJsVdlmvhJ1/OT7epyG+dwzMzesRjHvH8fwvNg3lGEW7s/r/KmGua9qOa5HNXgqKfJh9s"
    "uSpCyooalksEibzXMcjmPTqmWh7NNe43ZJPErF6t1Q42qzVSvYqiXrkibYMAXfC+Ja2pjo5prRUTOlp6"
    "mNiua1HLnuOVPoqmeWvHLNCsoq6klT1KiPwVcl/Wfe1zXfRci+Ck/de87m7rk1UxjE+sIu2bJTtVMRM4"
    "TCN9mmArzjG908UVJNFbUei1VW5ioxjM9URebl4Iifu1LJa1GtRrUyaiZInQ+HOa36TkTxU+iWupIk9e"
    "oj8EXNf1GW9d6XN5VxM04RHpHqbHsdOy0zETjMv0nw5zWtVznI1E4qqnk1N7jbmkEavXq7RDxrncnLC+"
    "oralkUEabznPcjWMTqvIrqbNVSVNUQ9W5XbNFipVXosn8DSe2XaA2ghmw9Zp866RN2qnYv8AuWrxai/j"
    "Lz6ePDzdo+1dJI5bXhZ7kRc2yV2WS96Rp/1L9nJTT7nK5yucqq5VzVVXVTstzbhmJi9fjCPaP5/hRbfv"
    "KMJt2p/X+HwfqtNvrLtc6a2W+B1RV1UrYoY28XOVck8PFdEPylT+Tfswfhyjbiq/0u5eKqPKlgkb61JE"
    "vNU5PcnHmiaaKrkOvqq8MKOmnxS2JszwnS4LwdRWKnVr5I29pVTImXbTO+m7w5JnqjURORkoBGSQAAal"
    "2mbDbHi68yXm33B9krqh29U7kCSxSu5v3N5uTl5qi5LxyzVVXEfRl99vhXziiAZRXVDGaKZTv6Mvvt8"
    "K+cPRl99vhXziiAMyoy6U7+jL77fCvnD0Zffb4V84ogDMqMulzvABJRmmgAVS1AAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAG5QAWqqdEAAREsAAAAAAAAAAA0Lt72Mrc31GKsIU6efOVZK23sTLt14rJGn4/NW/W4p6"
    "2jt9A9iZifJ5MRMYS54yMfHI6ORjmPYqtc1yZKipxRU6n8ljbWNkFhxuklwp1S2Xvd0qo25smVOCSt5"
    "6abyappxREQl3HOBsTYLrOwvtufFG5ytiqY/Xgl4/Rf10zyXJ2XFEN9NcS0VUTD6MLYtv2Gpd611zmR"
    "Kub4H+tE/wAWrw8UyXvNp4e2z22ZrY75bpqSTgstP/WR+KovrJ/eNHgg7XuvZtq866fP7x5T/wB8t9j"
    "bb1nypny+yp7ZjTClxai0t+oc14NlkSJy/ovyU9uKogmbvRTxyJ1a9FQj0FPX+F6Jn+i5MfMY/vCwp"
    "3zV+ajqsKaogharppo40Tir3In7zxLnjTClua5aq/UObeLY5Ukd9zc1JYAo/C9ET/XcmfiMP3kq3zV+"
    "Wjq3hiHbPbYWujsdumq5OCS1H9WxO/JNV/UasxTi2/Yllzulc98SLm2BnqxN8Gpx8VzXvPCBcbJuvZtl"
    "86KfP7z5z/3wr7+23r3lVPl9g/qNj5JGxxsc971RrWtTNVVeCInUyLA2BsS40rfN7FbnyxtdlLVSepBF"
    "w+k/rrnkmbsuCKVDsn2QWLBCR3CpVt0ve7rVSMyZDnxSJvLpvLquvBFVCdVXFLRTRMsQ2C7GVtklPin"
    "F9MnnrcpKK3vTPsF4pJIn4/NG/V4r62jd9AGiZmfVviIiMIAAePQAAAAAAAAAAc7wAS0RpoAFUtQAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAABuUAFqqnRAAERLAAAAAAAAAAAAAA+mspaatpZKSsp4amnlbuyRSsR"
    "7Hp0VF0VD7gBqDGfk/4RvL31FlmqLDUuXPdiTtYM881Xs3Kip4NciJ0NSYi2B4+tr1WggorxFqu9TVC"
    "MciJ1bJu69zd4roGUVzDGaIlBN1wliq1MdJcsN3ekjauSyTUcjWZ/nKmS/eeKui5LovedEAZ5rDKhzv"
    "TVck1Xoh7Vqwliq6xtktuG7vVxuXJJIaORzM/zkTJC9gM3QyoSLh3YHj65PRa+Cis8Wiq6pqEe5UXo2"
    "Pe17nbptvBnk/4Rsz2VF6mnv1S1c8pU7KDPPNF7Nqqq+DnKi9Db4MJrmWcURD6aKlpqKkjpKKmhpqe"
    "Ju7HFCxGMYnRETREPuAMWQAAAAAAAAAAAAAAADneACWiNNAAqlqAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AANygAtVU6IAAiJYCINqePrvjm/1FTU1MzLY2RfMqLeyjiYme6qt4K9U1Vy65rkmSIiJhxti01zddE"
    "Ac7wMrV5m6OiAOd4GVqZujogDneBlambo6IA53gZWpm6OiAOd4GVqZujogDneBlambo6IA53gZWpm6"
    "OiAOd4GVqZujogDneBlambo6IA53gZWpm6OiAOd4GVqZujogDneBlambo6IA53gZWpm6OiAOd4GVqZ"
    "ujogDneZjssx9d8DX+nqaapmfbHSIlbRbyrHKxct5UbwR6Jqjk1zTJc0VUVNp7F1b4ANTY53gAlojT"
    "QAKpagAAAAAAAAAAAAAAAAAAAAAAAAAAAADcoALVVOiAAIiW53gAlogC8Ae4IP1vL1QeC8AMD63l6oP"
    "BeAGB9by9UHgvADA+t5eqDwXgBgfW8vVB4LwAwPreXqg8F4AYH1vL1QeC8AMD63l6oPBeAGB9by9UHg"
    "vADA+t5eqDwXgBgfW8vVB4LwAwPreXqg8F4AYH1vL1QeC8AMD63l6oPAB4nOiAAIiW53gAlojTQAKpa"
    "gAAAAAAAAAAAAAAAAAAAAAAAAAAAADcoALVVOiAAIiW53gAloizgAS3GAJe2i40ueJr1Ue1Sx21kit"
    "p6djlRm6i6OVObl469TEjVN1b2901VUxNVWE/CzgRiDzN0Z8I5+ndZwIxAzdDhHP07rOBGIGbocI5+n"
    "dZwIxAzdDhHP07rOBGIGbocI5+ndZwIxAzdDhHP07rOBGIGbocI5+ndZwIxMt2c40ueGb1T+1SyW18"
    "iNqKd7lVm6q5K5qcnJx06ZHsXWFzdNVNMzTVjPwqEAG1UAJe2i40ueJr1Ue1Sx21kitp6djlRm6i6O"
    "VObl469TEjVN1b2901VUxNVWE/CzgRiZbs5xpc8M3qn9qlktr5EbUU73KrN1VyVzU5OTjp0yEXS5um"
    "qmmZpqxn4VCADaqEYgAiOzdEAAREtzvABLRGmgAVS1AAAAAAAAAAAAAAAAAAAAAAAAAAAAAG5QAWqq"
    "dEAAREtzvABLRFnAAluMS9tFwXc8M3qo9llktr5FdT1DGqrN1V0a5eTk4a9DEizgaptLe3vaqmmIqp"
    "xn5RiCzgeZWrPi/J17IxBZwGVqcX5OvZGILOAytTi/J17IxBZwGVqcX5OvZGILOAytTi/J17IxBZw"
    "GVqcX5OvZGILOAytTi/J17IxMt2c4LueJr1T+yyx21kiOqKh7VRu6i6tavNy8NOpUIPYtMLm9qqqZ"
    "imnCfkABtVCXtouC7nhm9VHsssltfIrqeoY1VZuqujXLycnDXoYkWcDVNpb297VU0xFVOM/KMTLdn"
    "OC7nia9U/sssdtZIjqioe1UbuourWrzcvDTqVCBFoub2qqpmKacJ+QAG1UIxABEdm6IAAiJbneAC"
    "WiNNAAqlqAAAAAAAAAAAAAAAAAAAAAAAAAAAAANygAtVU6IAAiJbneACWiLOABLcYAl7aLjS54mv"
    "VR7VLHbWSK2np2OVGbqLo5U5uXjr1MSNU3Vvb3TVVTE1VYT8LOBGIPM3Rnwjn6d1nAjEDN0OEc/Tu"
    "s4EYgZuhwjn6d1nAjEDN0OEc/Tus4EYgZuhwjn6d1nAjEDN0OEc/Tus4EYgZuhwjn6d1nAjEy3Zzj"
    "S54ZvVP7VLJbXyI2op3uVWbqrkrmpycnHTpkexdYXN01U0zNNWM/CoQAbVQAl7aLjS54mvVR7VLH"
    "bWSK2np2OVGbqLo5U5uXjr1MSNU3Vvb3TVVTE1VYT8LOBGJluznGlzwzeqf2qWS2vkRtRTvcqs3"
    "VXJXNTk5OOnTIRdLm6aqaZmmrGfhUIANqoRiACI7N0QABES3O8AEtEaaABVLUAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAblABaqp0QABES3O8AEtEWcACW4xL20XBdzwzeqj2WWS2vkV1PUMaqs3VXRrl5O"
    "Thr0MSLOBqm0t7e9qqaYiqnGflGILOB5las+L8nXsjEFnAZWpxfk69kYgs4DK1OL8nXsjEFnAZ"
    "Wpxfk69kYgs4DK1OL8nXsjEFnAZWpxfk69kYgs4DK1OL8nXsjEy3Zzgu54mvVP7LLHbWSI6oqHt"
    "VG7qLq1q83Lw06lQg9i0wub2qqpmKacJ+QAG1UJe2i4LueGb1UeyyyW18iup6hjVVm6q6NcvJ"
    "ycNehiRZwNU2lvb3tVTTEVU4z8oxMt2c4LueJr1T+yyx21kiOqKh7VRu6i6tavNy8NOpUIEWi"
    "5vaqqmYppwn5AAbVQjEAER2bogACIlud4AJaI00ACqWoAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "3KAC1VTogACIlud4AJaIs4AEtxgCXtouNLnia9VHtUsdtZIraenY5UZuoujlTm5eOvUxI1TdW"
    "9vdNVVMTVVhPws4EYg8zdGfCOfp3WcCMQM3Q4Rz9O6zgRiBm6HCOfp3WcCMQM3Q4Rz9O6zgRi"
    "Bm6HCOfp3WcCMQM3Q4Rz9O6zgRiBm6HCOfp3WcCMTLdnONLnhm9U/tUsltfIjaine5VZuquSu"
    "anJycdOmR7F1hc3TVTTM01Yz8KhABtVACXtouNLnia9VHtUsdtZIraenY5UZuoujlTm5eOvUx"
    "I1TdW9vdNVVMTVVhPws4EYmW7OcaXPDN6p/apZLa+RG1FO9yqzdVclc1OTk46dMhF0ubpqppm"
    "aasZ+FQgA2qhGIAIjs3RAAERLc7wAS0RpoAFUtQAAAAAAAAAAAAAAAAAAAAAAAAAAAABuUAFq"
    "qnRAAERLc7wAS0RZwAJbjEvbRcF3PDN6qPZZZLa+RXU9QxqqzdVdGuXk5OGvQxIs4GqbS3t7"
    "2qppiKqcZ+UYgs4HmVqz4vydeyMQWcBlanF+Tr2RiCzgMrU4vydeyMQWcBlanF+Tr2RiCzgM"
    "rU4vydeyMQWcBlanF+Tr2RiCzgMrU4vydeyMTLdnOC7nia9U/sssdtZIjqioe1UbuourWrzc"
    "vDTqVCD2LTC5vaqqmYppwn5AAbVQl7aLgu54ZvVR7LLJbXyK6nqGNVWbqro1y8nJw16GJFnA"
    "1TaW9ve1VNMRVTjPyjEy3Zzgu54mvVP7LLHbWSI6oqHtVG7qLq1q83Lw06lQgRaLm9qqqZim"
    "nCfkABtVCMQARHZuiAAIiW53gAlojTQAKpagAAAAAAAAAAAAAAAAAAAAAAAAAAAADcoALVVO"
    "iAAIiW53gAloizgAS3GAJe2i40ueJr1Ue1Sx21kitp6djlRm6i6OVObl469TEjVN1b2901VU"
    "xNVWE/CzgRiDzN0Z8I5+ndZwIxAzdDhHP07rOBGIGbocI5+ndZwIxAzdDhHP07rOBGIGbocI"
    "5+ndZwIxAzdDhHP07rOBGIGbocI5+ndZwIxMt2c40ueGb1T+1SyW18iNqKd7lVm6q5K5qcnJ"
    "x06ZHsXWFzdNVNMzTVjPwqEAG1UAJe2i40ueJr1Ue1Sx21kitp6djlRm6i6OVObl469TEjVN"
    "1b2901VUxNVWE/CzgRiZbs5xpc8M3qn9qlktr5EbUU73KrN1VyVzU5OTjp0yEXS5umqmmZpq"
    "xn4VCADaqEYgAiOzdEAAREtzvABLRGmgAVS1AAAAAAAAAAAAAAAAAAAAAPfseF6y5QJUvkb"
    "TQu+irm5q7vROh4BuOKNsUTImJk1jUa1OiIb7FuK5nFovXJoiMGH/gN+VP2f+YfgN+VP2f8Am"
    "MyBJyKPsjZ9f3Yb+A35U/Z/5h+A35U/Z/5jMgMij7GfX9wAG1qdEAAREtzvABLRFnAAluMS"
    "9tFwXc8M3qo9llktr5FdT1DGqrN1V0a5eTk4a9DEizgaptLe3vaqmmIqpxn5RiCzgeZWrPi"
    "/J17IxBZwGVqcX5OvZGILOAytTi/J17IxBZwGVqcX5OvZGILOAytTi/J17IxBZwGVqcX5Ov"
    "ZGILOAytTi/J17IxMt2c4LueJr1T+yyx21kiOqKh7VRu6i6tavNy8NOpUIPYtMLm9qqqZim"
    "nCfkABtVCXtouC7nhm9VHsssltfIrqeoY1VZuqujXLycnDXoYkWjFG+WRscbFe9y5IiIe/b"
    "8PJkj6165/8ApsX96ke5FNPrK/2Ta796PDFGOHvjhH+kHmW7OcF3PE16p/ZZY7ayRHVFQ9qo"
    "3dRdWtXm5eGnUtqCho4ERIqaNuXPdzX71P0oiJoiIhozYifRazZuVU4Y4T/n+GvQbCBs+q"
    "0VXA//ANOndzvB0QBpzdF1lagANTa53gAlojTQAKpagAAAAAAAAAAAAAAAAAAAAAblNNG5"
    "SVs3ui7T7Bvj0d/fD4b800OW/wD0n/wP7/8AoTaaZn0U+2bRNnw+eGLTHo7++Hw35o9Hf3w"
    "+G/NNz/0n/wAD+/8A6D+k/wDgf3/9D3LlB4hP/rp2RAADBduiAAIiW53gAloiiwToD3FVcM"
    "5undRYJ0AxOGc3TuosE6AYnDObp3UWCdAMThnN07qLBOgGJwzm6d1FgnQDE4ZzdO6iwToBi"
    "cM5undRYJ0AxOGc3TuosE6AYnDObp3UWCdAMThnN07qLPtpKeaqqY6anjWSWRyNY1OaqTeU"
    "L5JGD45HVmNKyLN0blpaHPkuX9Y/7lRqeLjyqvCMWVO68Zw8XTu3dhDDlNYaNNGyVkif1su"
    "X91vRP3nugxTadjq0YCsC3K5Ks1RKqspKRjsn1D05J0amaZu5ZpxVURYszNUruiim1T4afK"
    "IZWCJcc7UsZYtnkSsustJRPzRtFRuWKJG6aOyXN/D6yr3ZcDCTOLU+7ybsezogDneD3K1eZ"
    "ujogDneBlambo6IA53gZWpm6AANzS00ACqWoAAAAAAAAAAAAAAAAAAAAAG5TTRuUlbN7ou0"
    "+wWcRiWcWFr3c3vf8n6/sAA3KVGIAIjs3RAAERLRBtTwDd8DX+opqmlmfbHSL5lWo1VjlYqr"
    "uoruCPRNFauuaZpmioq4cdEAbYutc2nO8HRADN0eZWrneDogBm6GVq53g6IAZuhlaud4OiAG"
    "boZWrneDogBm6GVq53g6IAZuhlaud4OiAGboZWrneDogBm6GVq53g6IAZuhlaud4OiAGboZW"
    "rneXNshtjLRsxw7RMa5q+YRzPR3FHyJ2j0/+TlMqBjVX4mVNHhCJdtuK5cW7Q7jWJNv0VNI"
    "tLRIi5tSJiqm8n5y5u/Sy5IWo53ntqPPF5dnywAC47HaqCyWuC222mjp6aBiNa1jcs8ua9V"
    "Xiq8zeg3r2Vh5IcBeAPcGj63l6oPBeB5eKr/bMM2Wa7XaoSGniTRE1dI7k1qc3L/wA0GBG2"
    "TM4RSiMGU7SMbXPGt6WsrFWGkiVUpYVrs2xN/wA3LzX/AAskMWPE2mZmPMPLxLdYbXb3vVy"
    "du9qtiZzVevgh/d+u9PaaTtZV3pXaRxourl/h3mtLjW1Fwq31NS/ee77kToncaL13w+UeqR"
    "ZteKcZ9H5gAQU4AAAAAAAAAAAAAAAAAAAAADcppo3KStm90XafYLOIxLOLC17ub3v+T9f2A"
    "AblKjEAER2bogACIlgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABzvOiBzvNtr3arvsFoSPd"
    "I9XvVVVSLyziXa91Bvf8n6/sAHm4mvluw7aJbnc5kjhj0RE1dI7k1qc1X/zQ3KammapwgxN"
    "fLdh20S3O5zdnDHoiJq6R3JrU5qv/AJoTLjzFtxxbd1q6tVjp480pqZq5tib/AJuXmvPwyQ"
    "Y8xbccW3daurVY6ePNKamaubYm/wCbl5rz8MkMdI9deLoti2KLEeKr+7/QedfrvT2mk7WVd"
    "6V2kcaLq5f4d4v13p7TSdrKu9K7SONF1cv8O81pcq2ouFW+pqX7z3fc1OidxFu3fB5R6re1"
    "a8fnPoXKtqLhVvqal+8933NToncfmAIMzinRGAAAAAAAAAAAAAAAAAAAAAAAAAblNNG5SVs"
    "3ui7T7BZxGJZrHI5qOTVFTNCwte7m97/k/X9nyAfD3I1iuXRETNTcpUZAAiOzdEAAREsAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAOd50QOd5tte7Vd9gs4jErzE18t2HbRLc7nN2cMeiIm"
    "rpHcmtTmq/68CXa91DvamapoiNf2MTXy3YdtEtzucyRwx6Iiaukdya1Oar/AOaEy48xbccW"
    "3daurVY6ePNKamaubYm/5uXmvPwyQY8xbccW3daurVY6ePNKamaubYm/5uXmvPwyQx0xrrx"
    "SNi2KLEeKr+7/AEHnX6709ppO1lXeldpHGi6uX+HeL9d6e00nayrvSu0jjRdXL/DvNaXKtq"
    "LhVvqal+8933NToncRbt3weUeq3tWvH5z6Fyrai4Vb6mpfvPd9zU6J3H5gCDM4p0RgAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAbkaqOaipwVMzTZtXDlW2tslLMi5u3Ea/85NF/cStmnzmEXaY8"
    "ol6BVuGr3BLRxRTyI1N1FjkVdHNXhqSkbK2fYxpW0MVqu0yQviTchmf9FzeTVXkqdeGX65t"
    "NWEqPeNibtETHsoLtI9zf327vXPQ8DEt7gio5ooJEcm6qySIujWpx1MY85p+x7ft4uyyz39"
    "9N37zA9oOMaV1DLabVMkz5U3ZpmLm1rebUXmq/dl+rObnkqLGy1XK4iGtQvAGQ7OLO++Y3t"
    "VvRm/G6obJN07NnrO/Uip4qhpmcIxdTEYzgu8GDgiYpeDOAYOBiYM4Bg4GJgzgGDgYmDOAY"
    "OBiYM4Bg4GJgzgGDgYmDOAYOBiYM4Bg4GJgzgGDgYmDOAYOBiYM4Bg4GJgzgGDgYmDOAYOB"
    "iYM4Bg4GJgzgGDgYmDODneWwRObrXu03fYMix5i244tu61lYvZ08eaU1M1c2xN/zcvNefhk"
    "hjoN2KPNFM1RVMecB51+u9PaaTtZV3pXaRxourl/h3i/XentNJ2svrSu0jjRdXL/DvNaXGt"
    "qLhVvqal+8933InRO40Xbvg8o9Um1a8fnPoXKtqLhVvqal+8933NToncfmAIMzinRGAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAGR4JvLbfVOpKl+7TTLo5eDHdfBf4GOAypqmmcYY1UxVGEty"
    "g17h3FM9vY2mq2uqKZNGqi+uxO7qncZjQ3u11rUWGsi3l+o9d133KT6LtNSBXaqpeiD4aqO"
    "TNqoqdx8mxrAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADzr9d6e00nayrvSu0jjR"
    "dXL/DvP4vt7o7XTuc97ZJ8vUiauqr39ENb3GtqLhVvqal+8933NToncaLt2KfKPVvtWpq85"
    "9C41tRcKt9TUv3nu+5qdE7j8wBBmcU6IwAAAAAAAAAAAAAAAAAAAAAAAAf/9k="
)



# ════════════════════════════════════════════════════════════════════
#  CORE — PDF PARSING
# ════════════════════════════════════════════════════════════════════

MONTH_MAP = {
    "JANUARI": "01", "FEBRUARI": "02", "MARET": "03", "APRIL": "04",
    "MEI": "05", "JUNI": "06", "JULI": "07", "AGUSTUS": "08",
    "SEPTEMBER": "09", "OKTOBER": "10", "NOVEMBER": "11", "DESEMBER": "12",
}

AMOUNT_RE = re.compile(
    r'([\d,]+\.\d+)'
    r'\s*(DB)?'
    r'\s*([\d,]+\.\d+)?'
    r'\s*$'
)

SKIP_PATTERNS = [
    r'^REKENING TAHAPAN$', r'^REKENING TAPRES$', r'^KCP ', r'^TANGGAL\s+KETERANGAN',
    r'^CATATAN', r'^Apabila', r'^dengan akhir', r'^tercantum',
    r'^BCA berhak', r'^Rekening\.', r'^Bersambung',
    r'NO\.\s*REKENING', r'NOMOR REKENING', r'^HALAMAN', r'^PERIODE', r'^MATA UANG',
    r'^KETERANGAN\s*:', r'^FASILITAS\s*:', r'^DEPOK', r'^INDONESIA',
    r'^BEJI', r'^PERUM', r'TRANSAKSI TIDAK TERSEDIA',
    r'SALDO AWAL\s*:', r'SALDO AKHIR\s*:',
    r'MUTASI CR\s*:', r'MUTASI DB\s*:', r'^\s*$', r'^•',
]
SKIP_RE = re.compile('|'.join(SKIP_PATTERNS), re.IGNORECASE)


def parse_amount(s):
    if s is None:
        return None
    try:
        return float(s.replace(',', ''))
    except ValueError:
        return None


def extract_page_meta(text):
    meta = {}
    # Regular BCA Tahapan: NO. REKENING
    m = re.search(r'NO\.\s*REKENING\s*:\s*(\d+)', text)
    if m:
        meta['account_no'] = m.group(1)
    # BCA Tapres: NOMOR REKENING (no dot)
    if 'account_no' not in meta:
        m = re.search(r'NOMOR REKENING\s*:\s*(\d+)', text)
        if m:
            meta['account_no'] = m.group(1)

    # Account holder name — before NO./NOMOR REKENING on same line
    m_name = re.search(r'^(.+?)\s+NO\.\s*REKENING\s*:', text, re.MULTILINE)
    if not m_name:
        m_name = re.search(r'^(.+?)\s+NOMOR REKENING\s*:', text, re.MULTILINE)
    if m_name:
        meta['account_name'] = m_name.group(1).strip().title()
    else:
        # Tapres: name is typically the first non-empty line at top
        for line in text.splitlines():
            line = line.strip()
            if line and not re.match(r'^\d|^REKENING|^KCP|^JSEB|^JL\.', line, re.IGNORECASE):
                meta['account_name'] = line.title()
                break
        else:
            meta['account_name'] = ''

    m = re.search(r'MATA UANG\s*:\s*(\w+)', text)
    if m:
        meta['currency'] = m.group(1).upper()

    # Detect Tapres account type
    meta['is_tapres'] = bool(re.search(r'REKENING TAPRES', text, re.IGNORECASE))

    m_fas = re.search(r'FASILITAS\s*:\s*(.+)', text)
    m_ket = re.search(r'KETERANGAN\s*:\s*(.+)', text)
    keterangan_header = m_ket.group(1).strip().upper() if m_ket else ''
    if keterangan_header == '-':
        keterangan_header = ''
    meta['notes'] = keterangan_header.title() if keterangan_header else ''

    if meta.get('is_tapres'):
        meta['fasilitas'] = 'Tapres'
    elif m_fas:
        meta['fasilitas'] = m_fas.group(1).strip().title()
    elif not keterangan_header:
        meta['fasilitas'] = 'Main Account'
    elif keterangan_header == 'POKET VALAS':
        meta['fasilitas'] = 'Poket Valas'
    else:
        meta['fasilitas'] = 'Tabungan Tujuan'

    # Regular BCA: PERIODE  01 JANUARI 2026
    m = re.search(r'PERIODE\s*:\s*(\w+)\s+(\d{4})', text)
    if m:
        meta['month'] = MONTH_MAP.get(m.group(1).upper(), '01')
        meta['year']  = m.group(2)
    # Tapres: PERIODE : 01-11-2025 S/D 30-11-2025
    if 'year' not in meta:
        m = re.search(r'PERIODE\s*:\s*\d{2}-(\d{2})-(\d{4})', text)
        if m:
            meta['month'] = m.group(1)
            meta['year']  = m.group(2)
    return meta


def infer_type(description):
    """Assign a transaction type label. Can be monkey-patched by GUI rules."""
    d = description.upper()
    if 'SALDO AWAL'        in d: return 'OPENING'
    if 'TARIKAN ATM'       in d: return 'ATM'
    if 'BIAYA ADM'         in d: return 'ADMIN FEE'
    if 'FLAZZ BCA'         in d: return 'FLAZZ'
    if 'KARTU KREDIT'      in d: return 'CC PAYMENT'
    if 'KARTU DEBIT'       in d: return 'CARD'
    if 'BERLIAN SISTEM INF'in d: return 'SALARY/REIMBURSEMENT'
    if 'SWITCHING'         in d: return 'SWITCHING'
    if 'KR OTOMATIS'       in d: return 'AUTO CR'
    if 'DB OTOMATIS'       in d: return 'AUTO DB'
    if 'QR'                in d: return 'QRIS'
    if 'TRANSAKSI DEBIT'   in d: return 'QRIS'
    if 'E-BANKING'         in d: return 'TRANSFER'
    if 'DEPOSITO'          in d: return 'DEPOSITO'
    if 'DB DEBIT DOMESTIK' in d: return 'DEBIT'
    if 'BI-FAST'           in d: return 'BI-FAST'
    return 'OTHER'


def parse_transactions(text, account_no, currency, year, month,
                       fasilitas='Main Account', notes='', account_name=''):
    transactions = []
    lines = text.splitlines()

    table_start = 0
    for i, line in enumerate(lines):
        if re.search(r'TANGGAL\s+KETERANGAN', line):
            table_start = i + 1
            break

    current = None

    def flush():
        if current:
            transactions.append(dict(current))

    for line in lines[table_start:]:
        line = line.strip()
        if SKIP_RE.search(line):
            flush(); current = None
            continue

        date_match = re.match(r'^(\d{2}/\d{2})\s+(.*)', line)
        if date_match:
            flush()
            day_month = date_match.group(1)
            rest      = date_match.group(2).strip()
            am        = AMOUNT_RE.search(rest)

            if not am:
                current = {
                    'date': f"{day_month}/{year}", 'account_no': account_no,
                    'bank_name': 'BCA', 'account_name': account_name,
                    'currency': currency, 'fasilitas': fasilitas, 'notes': notes,
                    'keterangan': rest, '_pending_amount': True,
                }
                continue

            amount_str  = am.group(1)
            is_debit    = am.group(2) == 'DB'
            balance_str = am.group(3)
            description = rest[:am.start()].strip()
            amount      = parse_amount(amount_str)
            balance     = parse_amount(balance_str)

            current = {
                'date': f"{day_month}/{year}", 'account_no': account_no,
                'bank_name': 'BCA', 'account_name': account_name,
                'currency': currency, 'fasilitas': fasilitas, 'notes': notes,
                'keterangan': description,
                'debit':   amount if is_debit   else None,
                'credit':  amount if not is_debit else None,
                'balance': balance,
                '_pending_amount': False,
            }

        elif current is not None:
            if current.get('_pending_amount'):
                am = AMOUNT_RE.search(line)
                if am:
                    amount_str  = am.group(1)
                    is_debit    = am.group(2) == 'DB'
                    balance_str = am.group(3)
                    extra_desc  = line[:am.start()].strip()
                    if extra_desc:
                        current['keterangan'] = (current['keterangan'] + ' ' + extra_desc).strip()
                    amount  = parse_amount(amount_str)
                    balance = parse_amount(balance_str)
                    current['debit']           = amount if is_debit   else None
                    current['credit']          = amount if not is_debit else None
                    current['balance']         = balance
                    current['_pending_amount'] = False
                else:
                    current['keterangan'] = (current['keterangan'] + ' ' + line).strip()
            else:
                am = AMOUNT_RE.search(line)
                if am and current.get('balance') is None:
                    balance_str = am.group(3) or am.group(1)
                    current['balance'] = parse_amount(balance_str)
                    extra = line[:am.start()].strip()
                    if extra:
                        current['keterangan'] = (current['keterangan'] + ' ' + extra).strip()
                else:
                    current['keterangan'] = (current['keterangan'] + ' ' + line).strip()

    flush()

    # Post-process
    result = []
    for t in transactions:
        t['keterangan'] = ' '.join(t.get('keterangan', '').split())
        if not t.get('keterangan'):
            continue
        t['type'] = infer_type(t['keterangan'])

        # OPENING balance: move credit value → balance field
        if t['type'] == 'OPENING' and t.get('credit') is not None:
            t['balance'] = t['credit']
            t['credit']  = None

        if t.get('debit') is None and t.get('credit') is None and t.get('balance') is None:
            continue
        result.append(t)

    return result


TAPRES_SKIP_RE = re.compile(
    r'SALDO AWAL\s*:|SALDO AKHIR\s*:|MUTASI CR\s*:|MUTASI DB\s*:'
    r'|REKENING TAPRES|^KCP\b|^JSEB\b|^JL\.'
    r'|NOMOR REKENING|^HALAMAN|^PERIODE|^MATA UANG'
    r'|^CATATAN|Rekening Koran|Bank setiap|Bilamana|sanggahan|setujui'
    r'|^\d+\s*/\s*\d+\s*$'   # page markers like "1 /2"
    r'|DEPOK\s+\d|INDONESIA|^BEJI\b|^PERUM\b'
    r'|^\s*$',
    re.IGNORECASE | re.MULTILINE
)

# Matches a raw unformatted amount: digits only, no commas e.g. "4725.00"
TAPRES_RAW_AMT_RE = re.compile(r'^\d+\.\d{2}$')


def parse_tapres_transactions(text, account_no, currency, year, month, account_name=''):
    """
    Parser for BCA Rekening Tapres (investment savings) statements.

    Actual pdfplumber layout (per raw text inspection):
      10/11 TRSF E-BANKING CR 1011/FTSCY/WS95051 4,725.00
      4725.00            ← raw unformatted amount  (skip as data)
      botd? P-483022     ← description continuation
      Dividen ST011T2
      STOCKBIT SEKURITAS
      10/11 ...          ← next transaction

    Key: the formatted amount (with commas, optional "DB", optional balance) is
    always ON THE SAME LINE as the date. Subsequent raw-amount lines are skipped.
    """
    transactions = []
    lines        = text.splitlines()
    fasilitas    = 'Tapres'
    notes        = ''

    # Find table start after "TANGGAL KETERANGAN" header
    table_start = 0
    for i, line in enumerate(lines):
        if re.search(r'TANGGAL\s+KETERANGAN', line, re.IGNORECASE):
            table_start = i + 1
            break

    current = None

    def flush():
        nonlocal current
        if current:
            transactions.append(dict(current))
        current = None

    for raw_line in lines[table_start:]:
        line = raw_line.strip()
        if not line:
            continue
        if TAPRES_SKIP_RE.search(line):
            flush()
            continue

        date_m = re.match(r'^(\d{2}/\d{2})\s+(.*)', line)
        if date_m:
            flush()
            day_month = date_m.group(1)
            rest      = date_m.group(2).strip()
            am        = AMOUNT_RE.search(rest)

            if not am:
                # No amount on line yet (unusual) — set pending
                is_debit = bool(re.search(r'\bDB\b', rest))
                current = {
                    'date': f"{day_month}/{year}",
                    'account_no': account_no, 'bank_name': 'BCA',
                    'account_name': account_name, 'currency': currency,
                    'fasilitas': fasilitas, 'notes': notes,
                    'keterangan': rest, '_is_debit': is_debit,
                    '_pending_amount': True, '_raw_skipped': False,
                }
                continue

            # Amount is on the date line — parse immediately
            amount_str  = am.group(1)
            has_db      = am.group(2) == 'DB'
            balance_str = am.group(3)
            description = rest[:am.start()].strip()

            # Detect debit: DB marker OR keterangan contains " DB "
            is_debit = has_db or bool(re.search(r'\bDB\b', description))
            amount   = parse_amount(amount_str)
            balance  = parse_amount(balance_str)

            # Special: CR line showing "amount balance" (e.g. "211,539.58 211,539.71")
            # group(2) is None (no DB), group(3) is the balance
            if not has_db and balance_str and not is_debit:
                pass   # balance already set correctly

            current = {
                'date': f"{day_month}/{year}",
                'account_no': account_no, 'bank_name': 'BCA',
                'account_name': account_name, 'currency': currency,
                'fasilitas': fasilitas, 'notes': notes,
                'keterangan': description,
                'debit':   amount if is_debit  else None,
                'credit':  amount if not is_debit else None,
                'balance': balance,
                '_pending_amount': False,
                '_raw_skipped': False,   # next raw-amount line should be skipped
            }
            continue

        if current is None:
            continue

        # Skip the raw unformatted amount line (first continuation line)
        if TAPRES_RAW_AMT_RE.match(line) and not current.get('_raw_skipped'):
            current['_raw_skipped'] = True
            continue

        # For pending-amount rows, try to find the amount
        if current.get('_pending_amount'):
            am = AMOUNT_RE.search(line)
            if am:
                has_db   = am.group(2) == 'DB'
                is_debit = has_db or current.get('_is_debit', False)
                amount   = parse_amount(am.group(1))
                balance  = parse_amount(am.group(3)) if am.group(3) else None
                current['debit']           = amount if is_debit  else None
                current['credit']          = amount if not is_debit else None
                current['balance']         = balance
                current['_pending_amount'] = False
                current['_raw_skipped']    = False
            else:
                if not TAPRES_RAW_AMT_RE.match(line):
                    current['keterangan'] = (current['keterangan'] + ' ' + line).strip()
            continue

        # Description continuation (already have amount)
        if not TAPRES_RAW_AMT_RE.match(line):
            # Skip page markers like "1 /2"
            if not re.match(r'^\d+\s*/\s*\d+$', line):
                current['keterangan'] = (current['keterangan'] + ' ' + line).strip()

    flush()

    # Post-process — same as regular parser
    result = []
    for t in transactions:
        t['keterangan'] = ' '.join(t.get('keterangan', '').split())
        if not t.get('keterangan'):
            continue
        t['type'] = infer_type(t['keterangan'])
        if t['type'] == 'OPENING' and t.get('credit') is not None:
            t['balance'] = t['credit']
            t['credit']  = None
        if t.get('debit') is None and t.get('credit') is None and t.get('balance') is None:
            continue
        result.append(t)
    return result


def parse_pdf(pdf_path, pdf_password=""):
    try:
        import pdfplumber
    except ImportError:
        raise ImportError("pdfplumber not installed. Run: pip install pdfplumber")

    import io as _io
    all_transactions = []
    pdf_source  = pdf_path
    pdf_open_pw = pdf_password or ""

    if pdf_password:
        decrypted = False

        # Strategy 0: bytes password direct to pdfplumber (pdfminer needs bytes not str)
        for _enc in ('latin-1', 'utf-8'):
            try:
                _pwd_bytes = pdf_password.encode(_enc)
                with pdfplumber.open(pdf_path, password=_pwd_bytes) as _t:
                    _ = _t.pages[0]
                pdf_source  = pdf_path
                pdf_open_pw = _pwd_bytes
                decrypted   = True
                break
            except Exception:
                continue

        # Strategy 1: pypdf — decrypt and strip encryption
        if not decrypted:
            try:
                from pypdf import PdfReader, PdfWriter
                _reader = PdfReader(pdf_path)
                if _reader.is_encrypted:
                    _res = _reader.decrypt(pdf_password)
                    _res_int = int(_res) if hasattr(_res, '__int__') else _res
                    if _res_int == 0:
                        raise ValueError(
                            f"Wrong password for: {os.path.basename(pdf_path)}\n"
                            f"BCA: password is Date of Birth (DDMMYYYY).")
                _writer = PdfWriter()
                _writer.append(_reader)
                try: _writer._encrypt = None
                except Exception: pass
                _buf = _io.BytesIO()
                _writer.write(_buf)
                _buf.seek(0)
                pdf_source  = _buf
                pdf_open_pw = ""
                decrypted   = True
            except ImportError:
                pass
            except ValueError:
                raise
            except Exception:
                pass

        # Strategy 2: pikepdf
        if not decrypted:
            try:
                import pikepdf as _pikepdf
                _new = _pikepdf.Pdf.new()
                with _pikepdf.open(pdf_path, password=pdf_password) as _locked:
                    _new.pages.extend(_locked.pages)
                _buf = _io.BytesIO()
                _new.save(_buf)
                _buf.seek(0)
                pdf_source  = _buf
                pdf_open_pw = ""
                decrypted   = True
            except ImportError:
                pass
            except Exception as _e:
                if 'password' in str(_e).lower() or 'incorrect' in str(_e).lower():
                    raise ValueError(
                        f"Wrong password for: {os.path.basename(pdf_path)}\n"
                        f"BCA: password is Date of Birth (DDMMYYYY).")

        if not decrypted:
            pdf_open_pw = pdf_password   # final fallback

    try:
        pdf_file = pdfplumber.open(pdf_source, password=pdf_open_pw)
    except Exception as e:
        err = str(e).lower()
        if pdf_password and ("password" in err or "encrypted" in err
                             or "incorrect" in err or not str(e)):
            raise ValueError(
                f"Could not open password-protected PDF: {os.path.basename(pdf_path)}\n\n"
                f"• Check the password is correct (BCA: Date of Birth DDMMYYYY)\n"
                f"• Install decryption support:  pip install pypdf pikepdf")
        raise
    with pdf_file:
        for page in pdf_file.pages:
            text = page.extract_text()
            if not text:
                continue
            meta      = extract_page_meta(text)
            account_no   = meta.get('account_no', '')
            currency     = meta.get('currency', 'IDR')
            year         = meta.get('year', '')
            month        = meta.get('month', '')
            fasilitas    = meta.get('fasilitas', 'Main Account')
            notes        = meta.get('notes', '')
            account_name = meta.get('account_name', '')

            if meta.get('is_tapres'):
                txns = parse_tapres_transactions(
                    text, account_no, currency, year, month, account_name)
            elif re.search(r'MUTASI CR\s*:\s*[\d,\.]+\s+\d+\s*\n.*MUTASI DB', text, re.DOTALL):
                lines_after = text[text.find('TANGGAL KETERANGAN'):]
                if not re.search(r'^\d{2}/\d{2}\s', lines_after, re.MULTILINE):
                    continue
                txns = parse_transactions(text, account_no, currency, year, month, fasilitas, notes, account_name)
            else:
                txns = parse_transactions(text, account_no, currency, year, month, fasilitas, notes, account_name)
            all_transactions.extend(txns)

    return all_transactions


# ════════════════════════════════════════════════════════════════════
#  BANK JAGO — PDF PARSING
# ════════════════════════════════════════════════════════════════════

JAGO_MONTH_MAP = {
    'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
    'Mei': '05', 'Jun': '06', 'Jul': '07', 'Agu': '08',
    'Sep': '09', 'Okt': '10', 'Nov': '11', 'Des': '12',
}

# Lines to skip inside Jago transaction pages
JAGO_SKIP_RE = re.compile(
    r'^PT Bank Jago'
    r'|^www\.jago\.com'
    r'|^Laporan Keuangan Bulanan'
    r'|^Akun Aktif'
    r'|^Total Pemasukan'
    r'|^Total Pengeluaran'
    r'|^Saldo Akhir'
    r'|^INFO PENTING'
    r'|^\*\s*Dokumen ini'
    r'|^NILAI TUKAR MATA UANG'
    r'|^Nilai tukar terhadap'
    r'|^Mata Uang\s+Nilai Kurs'
    r'|^\d+\s+[A-Z]{3}\s+[\d\.]+$'   # currency rate rows
    r'|^RINGKASAN SALDO'
    r'|^SOROTAN'
    r'|^KANTONG PERSONAL'
    r'|^KANTONG BERSAMA'
    r'|^Nama Kantong'
    r'|^Total Saldo'
    r'|^merupakan peserta',
    re.IGNORECASE | re.MULTILINE
)


def _parse_jago_amount(s):
    """Convert Jago Indonesian-format number string to float.
    e.g. '10.067.671' -> 10067671.0  |  '59.900,90' -> 59900.9
    """
    if not s:
        return None
    s = s.strip().lstrip('+').lstrip('-')
    # Replace dots (thousands separators) then comma (decimal separator)
    s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def _parse_jago_date(s):
    """Convert 'DD Mon YYYY' to 'DD/MM/YYYY'. Returns None if unparseable."""
    parts = s.strip().split()
    if len(parts) != 3:
        return None
    day, mon, year = parts
    mm = JAGO_MONTH_MAP.get(mon)
    if not mm:
        return None
    try:
        return f"{int(day):02d}/{mm}/{year}"
    except ValueError:
        return None


# Matches the start of a transaction line:
# e.g. "24 Apr 2026  BAKMI ONO  Pembayaran QRIS  -50.000  10.067.671"
JAGO_TXN_RE = re.compile(
    r'^(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})'
    r'\s+(.+?)'
    r'\s+([+-][\d\.]+(?:,[\d]+)?)'
    r'\s+([\d\.]+(?:,[\d]+)?)\s*$'
)

# Also handle transactions where balance is bare 0
JAGO_TXN_ZERO_BAL_RE = re.compile(
    r'^(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})'
    r'\s+(.+?)'
    r'\s+([+-]\d[\d\.]*(?:,[\d]+)?)'
    r'\s+(0)\s*$'
)

# Matches the time line that immediately follows a transaction line
JAGO_TIME_RE = re.compile(r'^(\d{2}\.\d{2})\s*(.*)')

# Matches a pocket header: a line that is just a pocket name followed by "Saldo Sebelumnya"
# The name occupies the full line before the summary block
JAGO_POCKET_HEADER_RE = re.compile(
    r'^(.+?)\s+Saldo Sebelumnya'
)

# Matches an ID Kantong line to capture pocket ID
JAGO_ID_KANTONG_RE = re.compile(r'ID Kantong\s+(\S+)')

# Matches currency line
JAGO_CURRENCY_RE = re.compile(r'Mata Uang Dalam\s+(\w+)')


def _parse_jago_page(text, account_name):
    """
    Parse a single page of a Jago PDF (from page 3 onward).
    Returns a list of transaction dicts.
    One page may contain multiple Kantong sections.
    """
    transactions = []
    lines = text.splitlines()

    current_pocket = None
    current_id = None
    current_currency = 'IDR'
    current_txn = None
    in_txn_section = False

    def flush_txn():
        nonlocal current_txn
        if current_txn:
            transactions.append(dict(current_txn))
        current_txn = None

    i = 0
    while i < len(lines):
        raw = lines[i]
        line = raw.strip()
        i += 1

        if not line:
            continue

        # Skip boilerplate lines
        if JAGO_SKIP_RE.search(line):
            continue

        # Detect start of a new Kantong section from header line
        # e.g. "Kantong Utama   Saldo Sebelumnya 0,39"
        ph = JAGO_POCKET_HEADER_RE.match(line)
        if ph:
            flush_txn()
            current_pocket = ph.group(1).strip()
            in_txn_section = False
            # Check if ID Kantong is embedded in this same line
            id_inline = JAGO_ID_KANTONG_RE.search(line)
            if id_inline:
                current_id = id_inline.group(1)
            continue

        # Capture ID Kantong (standalone line)
        id_m = JAGO_ID_KANTONG_RE.search(line)
        if id_m:
            current_id = id_m.group(1)
            continue

        # Capture currency
        cur_m = JAGO_CURRENCY_RE.search(line)
        if cur_m:
            current_currency = cur_m.group(1).upper()
            in_txn_section = False
            continue

        # Column header line — next lines are transactions
        if re.match(r'^Tanggal\s*&\s*Waktu', line, re.IGNORECASE):
            in_txn_section = True
            continue

        if not in_txn_section or current_pocket is None:
            continue

        # Skip ID# lines and raw transaction codes
        if re.match(r'^ID#\s+', line):
            continue
        if re.match(r'^[A-Z0-9]{10,}', line) and ' ' not in line:
            continue

        # Try to match a transaction line (handles both decimal and integer amounts)
        txn_m = JAGO_TXN_RE.match(line) or JAGO_TXN_ZERO_BAL_RE.match(line)
        if txn_m:
            flush_txn()
            date_raw   = txn_m.group(1)
            desc_raw   = txn_m.group(2).strip()
            amount_raw = txn_m.group(3)
            bal_raw    = txn_m.group(4)

            date_str = _parse_jago_date(date_raw)
            if not date_str:
                continue

            amount = _parse_jago_amount(amount_raw)
            balance = _parse_jago_amount(bal_raw)
            is_debit = amount_raw.startswith('-')

            current_txn = {
                'date':         date_str,
                'account_no':   current_id or '',
                'bank_name':    'Jago',
                'account_name': account_name,
                'currency':     current_currency,
                'fasilitas':    current_pocket,
                'notes':        '',
                'keterangan':   desc_raw,
                'debit':        amount if is_debit else None,
                'credit':       amount if not is_debit else None,
                'balance':      balance,
            }
            continue

        # Time line (HH.MM + extra info) — second line of a transaction
        time_m = JAGO_TIME_RE.match(line)
        if time_m and current_txn is not None:
            extra = time_m.group(2).strip()
            # extra may contain bank name + account, or just bank name
            # Append to keterangan only if it's not a bare ID# or code
            if extra and not re.match(r'^ID#\s+', extra) and not re.match(r'^[A-Z0-9]{10,}$', extra):
                # Remove trailing ID# portion if present
                extra = re.sub(r'\s*ID#\s+\S+.*$', '', extra).strip()
                if extra:
                    current_txn['keterangan'] = (current_txn['keterangan'] + ' ' + extra).strip()
            continue

        # Any other continuation line — append to keterangan if we have an active txn
        if current_txn is not None:
            # Skip pure ID# lines, transaction codes, and page markers
            if (re.match(r'^ID#\s+', line)
                    or re.match(r'^[A-Z0-9]{15,}$', line)
                    or re.match(r'^\d+$', line)
                    or re.match(r'^\d+\s*/\s*\d+$', line)):
                continue
            # Append meaningful continuation text
            current_txn['keterangan'] = (current_txn['keterangan'] + ' ' + line).strip()

    flush_txn()
    return transactions


def parse_jago_pdf(pdf_path, pdf_password=""):
    """
    Parse a Bank Jago monthly statement PDF.
    - Skips page 1 (summary page) and any page without transaction data.
    - Each Kantong is treated as a separate account (account_no = ID Kantong).
    Returns a list of transaction dicts.
    """
    try:
        import pdfplumber
    except ImportError:
        raise ImportError("pdfplumber not installed. Run: pip install pdfplumber")

    all_transactions = []
    open_kw = {'password': pdf_password.encode('latin-1')} if pdf_password else {}

    with pdfplumber.open(pdf_path, **open_kw) as pdf:
        # Extract account name from page 1
        account_name = ''
        if pdf.pages:
            p1_text = pdf.pages[0].extract_text() or ''
            m = re.search(r'^(.+?)\s*/\s*[\d]+', p1_text, re.MULTILINE)
            if m:
                account_name = m.group(1).strip().title()

        # Skip page 1 (always summary); parse from page 2 onward
        for page in pdf.pages[1:]:
            text = page.extract_text() or ''
            # Skip pages that have no transaction section
            if not re.search(r'Tanggal\s*&\s*Waktu', text, re.IGNORECASE):
                continue
            txns = _parse_jago_page(text, account_name)
            all_transactions.extend(txns)

    # Post-process: clean up keterangan & infer type
    result = []
    for t in all_transactions:
        t['keterangan'] = ' '.join(t.get('keterangan', '').split())
        if not t.get('keterangan'):
            continue
        if t.get('debit') is None and t.get('credit') is None and t.get('balance') is None:
            continue
        result.append(t)
    return result


# ════════════════════════════════════════════════════════════════════
#  BANK ADAPTER PLUGIN SYSTEM
# ════════════════════════════════════════════════════════════════════

class BankAdapter:
    """Base class for bank statement parsers."""
    name     = "Unknown"
    code     = "UNKNOWN"
    ready    = False   # True when parser is fully implemented
    note     = ""      # shown in UI when not ready

    def parse(self, pdf_path, pdf_password=""):
        raise NotImplementedError(f"{self.name} parser not yet implemented.")


class BCAAdapter(BankAdapter):
    name  = "BCA — Bank Central Asia"
    code  = "BCA"
    ready = True

    def parse(self, pdf_path, pdf_password=""):
        return parse_pdf(pdf_path, pdf_password=pdf_password)


class JagoAdapter(BankAdapter):
    name  = "Jago — Bank Jago"
    code  = "Jago"
    ready = True

    def parse(self, pdf_path, pdf_password=""):
        return parse_jago_pdf(pdf_path, pdf_password=pdf_password)


class MandiriAdapter(BankAdapter):
    name  = "Mandiri — Bank Mandiri"
    code  = "Mandiri"
    ready = False
    note  = "Parser coming soon. Contributions welcome."

    def parse(self, pdf_path):
        # TODO: implement Mandiri e-Statement PDF parser
        raise NotImplementedError(
            "Mandiri parser is not yet implemented.\n"
            "Please use BCA for now.")


class BNIAdapter(BankAdapter):
    name  = "BNI — Bank Negara Indonesia"
    code  = "BNI"
    ready = False
    note  = "Parser coming soon."

    def parse(self, pdf_path):
        raise NotImplementedError(
            "BNI parser is not yet implemented.\n"
            "Please use BCA for now.")


class BRIAdapter(BankAdapter):
    name  = "BRI — Bank Rakyat Indonesia"
    code  = "BRI"
    ready = False
    note  = "Parser coming soon."

    def parse(self, pdf_path):
        raise NotImplementedError(
            "BRI parser is not yet implemented.\n"
            "Please use BCA for now.")


# Registry: ordered list of adapters
BANK_ADAPTERS = [BCAAdapter(), JagoAdapter(), MandiriAdapter(), BNIAdapter(), BRIAdapter()]
BANK_ADAPTER_MAP = {a.code: a for a in BANK_ADAPTERS}


# ════════════════════════════════════════════════════════════════════
#  FX RATE LOOKUP  (ExchangeRate-API)
# ════════════════════════════════════════════════════════════════════

# In-process cache: (currency, date_str) → rate (float, IDR per 1 unit of currency)
_fx_cache = {}


def _parse_date(date_str):
    """Parse DD/MM/YYYY → (year, month, day) ints. Returns None on failure."""
    try:
        parts = date_str.strip().split('/')
        if len(parts) == 3:
            return int(parts[2]), int(parts[1]), int(parts[0])
    except Exception:
        pass
    return None


def fetch_fx_rate(currency, date_str, api_key=""):
    """
    Return the IDR exchange rate for 1 unit of *currency* on *date_str* (DD/MM/YYYY).

    Strategy:
      1. Return 1.0 immediately for IDR.
      2. If api_key is provided → try historical endpoint (paid plan):
           https://v6.exchangerate-api.com/v6/{KEY}/history/{CCY}/{Y}/{M}/{D}
      3. Fall back to latest rates (free, no key):
           https://open.er-api.com/v6/latest/{CCY}
    Caches results per (currency, date) to avoid redundant HTTP calls.
    Returns None if all attempts fail.
    """
    import urllib.request
    import urllib.error

    if currency == 'IDR':
        return 1.0, "IDR"

    cache_key = (currency.upper(), date_str)
    if cache_key in _fx_cache:
        return _fx_cache[cache_key]

    rate = None
    source = None

    # ── 1. Historical endpoint (paid API key) ─────────────────────────
    if api_key:
        parsed = _parse_date(date_str)
        if parsed:
            y, m, d = parsed
            url = (f"https://v6.exchangerate-api.com/v6/{api_key}"
                   f"/history/{currency.upper()}/{y}/{m:02d}/{d:02d}")
            try:
                with urllib.request.urlopen(url, timeout=8) as r:
                    data = json.loads(r.read())
                if data.get('result') == 'success':
                    rates = data.get('conversion_rates', {})
                    rate  = rates.get('IDR')
                    if rate:
                        source = "historical"
            except Exception:
                pass

    # ── 2. Fawaz Ahmed Free Historical API ────────────────────────────
    if rate is None:
        parsed = _parse_date(date_str)
        if parsed:
            y, m, d = parsed
            fawaz_date = f"{y}-{m:02d}-{d:02d}"
            url = f"https://cdn.jsdelivr.net/npm/@fawazahmed0/currency-api@{fawaz_date}/v1/currencies/{currency.lower()}.json"
            try:
                with urllib.request.urlopen(url, timeout=8) as r:
                    data = json.loads(r.read())
                rate = data.get(currency.lower(), {}).get('idr')
                if rate:
                    source = "historical"
            except Exception:
                pass

    # ── 3. Latest rates fallback (free, no key needed) ────────────────
    if rate is None:
        url = f"https://open.er-api.com/v6/latest/{currency.upper()}"
        try:
            with urllib.request.urlopen(url, timeout=8) as r:
                data = json.loads(r.read())
            if data.get('result') == 'success':
                rate = data.get('rates', {}).get('IDR')
                if rate:
                    source = "latest fallback"
        except Exception:
            pass

    if rate is not None:
        _fx_cache[cache_key] = (rate, source)
    return (rate, source) if rate else (None, None)


def apply_fx_rates(transactions, api_key="", log_fn=None):
    """
    Enrich transactions in-place with '_fx_rate', '_local_amount', '_local_balance'.
    Skips IDR rows (already local). Logs progress via log_fn(msg, tag).
    """
    # Collect unique (currency, date) pairs that need conversion
    needed = {
        (t['currency'], t.get('date', ''))
        for t in transactions
        if t.get('currency', 'IDR') != 'IDR'
    }
    if not needed:
        return

    fetched, failed = 0, 0
    for currency, date_str in sorted(needed):
        rate_info = fetch_fx_rate(currency, date_str, api_key)
        if isinstance(rate_info, tuple):
            rate, source = rate_info
        else:
            rate, source = rate_info, "unknown"
            
        if rate:
            fetched += 1
            if log_fn:
                log_fn(f"  FX {currency}/{date_str}: 1 {currency} = {rate:,.2f} IDR ({source})", "info")
        else:
            failed += 1
            if log_fn:
                log_fn(f"  ⚠ FX {currency}/{date_str}: rate unavailable", "warn")

    # Apply rates to each transaction
    for t in transactions:
        currency = t.get('currency', 'IDR')
        if currency == 'IDR':
            t['_fx_rate']       = 1.0
            t['_fx_rate_mode']  = 'IDR'
            t['_local_amount']  = None
            t['_local_balance'] = None
        else:
            cached = _fx_cache.get((currency.upper(), t.get('date', '')))
            if cached and isinstance(cached, tuple):
                rate, source = cached
            elif cached:
                rate, source = cached, "unknown"
            else:
                rate, source = None, None
                
            t['_fx_rate'] = rate
            if rate is not None:
                t['_fx_rate_mode'] = 'Exact Date' if source == 'historical' else "Today's Rate"
                debit  = t.get('debit')
                credit = t.get('credit')
                if debit is not None:
                    t['_local_amount'] = round(-debit  * rate, 2)
                elif credit is not None:
                    t['_local_amount'] = round(credit  * rate, 2)
                else:
                    t['_local_amount'] = None
                cb = t.get('_calc_balance')
                t['_local_balance'] = round(cb * rate, 2) if cb is not None else None
            else:
                t['_fx_rate_mode']  = 'Unavailable'
                t['_local_amount']  = None
                t['_local_balance'] = None

    if log_fn:
        log_fn(f"  FX lookup: {fetched} OK, {failed} failed", "ok" if not failed else "warn")


# ════════════════════════════════════════════════════════════════════
#  EXPORT FUNCTIONS
# ════════════════════════════════════════════════════════════════════

def _transaction_to_dict(t):
    debit  = t.get('debit')
    credit = t.get('credit')
    if debit is not None:
        dc, signed = 'DB', round(-debit, 2)
    elif credit is not None:
        dc, signed = 'CR', round(credit, 2)
    else:
        dc, signed = '', None

    currency     = t.get('currency', 'IDR')
    currency_amt = signed

    # Use FX-converted local amount/balance if available, else fall back to
    # same value for IDR, or None for foreign currency without a rate yet.
    if currency == 'IDR':
        local_amt = signed
    elif '_local_amount' in t:
        local_amt = t['_local_amount']
    else:
        local_amt = None

    calc_bal = t.get('_calc_balance')
    currency_bal = round(calc_bal, 2) if calc_bal is not None else None

    if currency == 'IDR':
        local_bal = currency_bal
    elif '_local_balance' in t:
        local_bal = t['_local_balance']
    else:
        local_bal = None

    return {
        'seq_no':           t.get('seq_no', ''),
        'account_no':       t.get('account_no', ''),
        'bank_name':        t.get('bank_name', 'BCA'),
        'account_name':     t.get('account_name', ''),
        'date':             t.get('date', ''),
        'currency':         currency,
        'fx_rate_mode':     t.get('_fx_rate_mode', '—'),
        'exchange_rate':    t.get('_fx_rate'),
        'fasilitas':        t.get('fasilitas', ''),
        'notes':            t.get('notes', ''),
        'keterangan':       t.get('keterangan', ''),
        'type':             t.get('type', ''),
        'dc':               dc,
        'currency_amount':  currency_amt,
        'local_amount':     local_amt,
        'currency_balance': currency_bal,
        'local_balance':    local_bal,
    }

_HEADERS       = ['seq_no','account_no','bank_name','account_name','date','currency','fx_rate_mode',
                  'exchange_rate','fasilitas','notes','keterangan','type','dc',
                  'currency_amount','local_amount','currency_balance','local_balance']
_HEADER_LABELS = ['Seq No.','Account No.','Bank Name','Account Name','Date','Currency','FX Rate Mode',
                  'Exchange Rate','Fasilitas','Notes','Keterangan','Type','DC',
                  'Currency Amount','Local Amount','Currency Balance','Local Balance']

# Runtime column configuration — overridden by App at startup
_active_col_order   = list(_HEADERS)
_active_col_visible = {h: True for h in _HEADERS}

def _active_headers():
    """Returns (keys, labels) filtered and ordered per current column config."""
    keys   = [h for h in _active_col_order if _active_col_visible.get(h, True)]
    lbl_map = dict(zip(_HEADERS, _HEADER_LABELS))
    labels = [lbl_map[h] for h in keys]
    return keys, labels


def _calc_running_balance(transactions):
    """Compute running balance per sub-account in-place. Skips if already done."""
    if not transactions or transactions[0].get('_balance_done'):
        return
    running = {}
    for t in transactions:
        key = (t.get('account_no'), t.get('currency'),
               t.get('fasilitas'),  t.get('notes'))
        if t.get('type') == 'OPENING':
            running[key] = t.get('balance') or 0.0
            t['_calc_balance'] = running[key]
        else:
            prev  = running.get(key, 0.0)
            calc  = round(prev - (t.get('debit') or 0) + (t.get('credit') or 0), 2)
            running[key] = calc
            t['_calc_balance'] = calc
    transactions[0]['_balance_done'] = True   # guard against redundant recalculation


def _assign_seq_numbers(transactions):
    """Assign a sequential Seq No. per group of (bank_name, account_no, date, fasilitas).
    OPENING rows are assigned seq 0; regular transactions are numbered 1, 2, 3…
    Seq resets for each new group combination.
    """
    counters = {}   # group_key → next seq int
    for t in transactions:
        key = (
            t.get('bank_name',  'BCA'),
            t.get('account_no', ''),
            t.get('date',       ''),
            t.get('fasilitas',  ''),
        )
        if t.get('type') == 'OPENING':
            t['seq_no'] = 0
        else:
            counters[key] = counters.get(key, 0) + 1
            t['seq_no'] = counters[key]


def write_excel(transactions, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")

    _calc_running_balance(transactions)

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    HEADER_BG  = "2E75B6"
    ALT_ROW    = "EBF3FB"
    WHITE      = "FFFFFF"
    LIGHT_BLUE = "BDD7EE"
    thin   = Side(style='thin', color='B8CCE4')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    data_font= Font(name="Arial", size=9)
    db_font  = Font(name="Arial", size=9, color="C00000")
    cr_font  = Font(name="Arial", size=9, color="1A6B1A")
    ob_font  = Font(name="Arial", bold=True, size=9)

    def fill(hex_c):
        return PatternFill("solid", fgColor=hex_c)

    active_keys, active_labels = _active_headers()

    col_width_map = dict(zip(_HEADERS,
                             [8, 14, 10, 22, 13, 10, 14, 14, 60, 14, 6, 18, 18, 18, 18]))
    col_widths = [col_width_map.get(h, 14) for h in active_keys]

    for c, (h, w) in enumerate(zip(active_labels, col_widths), 1):
        cell = ws.cell(1, c, h)
        cell.font = hdr_font
        cell.fill = fill(HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 20

    numeric_cols = {"exchange_rate", "currency_amount", "local_amount", "currency_balance", "local_balance"}
    align_map = {
        "seq_no": "center", "account_no": "center", "bank_name": "center",
        "account_name": "left", "date": "center", "currency": "center",
        "fx_rate_mode": "center", "exchange_rate": "right",
        "fasilitas": "center", "notes": "center", "keterangan": "left",
        "type": "center", "dc": "center",
        "currency_amount": "right", "local_amount": "right",
        "currency_balance": "right", "local_balance": "right",
    }

    for idx, t in enumerate(transactions):
        row        = idx + 2
        is_opening = t.get('type') == 'OPENING'
        row_fill   = fill(LIGHT_BLUE) if is_opening else (fill(ALT_ROW) if idx % 2 == 0 else fill(WHITE))

        debit    = t.get('debit')
        credit   = t.get('credit')
        currency = t.get('currency', 'IDR')
        if debit is not None:
            dc, signed = 'DB', -debit
        elif credit is not None:
            dc, signed = 'CR', credit
        else:
            dc, signed = '', None

        currency_amt = signed
        local_amt    = signed if currency == 'IDR' else None
        currency_bal = t.get('_calc_balance')
        local_bal    = t.get('_calc_balance') if currency == 'IDR' else None

        val_map = {
            'seq_no': t.get('seq_no', ''), 'account_no': t.get('account_no'),
            'bank_name': t.get('bank_name', 'BCA'), 'account_name': t.get('account_name', ''),
            'date': t.get('date'), 'currency': currency,
            'fx_rate_mode': t.get('_fx_rate_mode', '—'),
            'exchange_rate': t.get('_fx_rate'),
            'fasilitas': t.get('fasilitas'), 'notes': t.get('notes'),
            'keterangan': t.get('keterangan'), 'type': t.get('type'), 'dc': dc,
            'currency_amount': currency_amt, 'local_amount': local_amt,
            'currency_balance': currency_bal, 'local_balance': local_bal,
        }

        for c, key in enumerate(active_keys, 1):
            v    = val_map.get(key)
            a    = align_map.get(key, "left")
            cell = ws.cell(row, c, v)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(horizontal=a, vertical="center",
                                       wrap_text=(key == "keterangan"))
            if key in numeric_cols and v is not None:
                cell.number_format = '#,##0.00'
                cell.font = db_font if dc == 'DB' else (cr_font if dc == 'CR' else ob_font)
            else:
                cell.font = ob_font if is_opening else data_font

        ws.row_dimensions[row].height = 16

    ws.freeze_panes = "A2"
    wb.save(output_path)


def write_csv(transactions, output_path):
    import csv
    _calc_running_balance(transactions)
    rows = [_transaction_to_dict(t) for t in transactions]
    active_keys, active_labels = _active_headers()
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=active_keys, extrasaction='ignore')
        f.write(",".join(active_labels) + "\n")
        writer.writerows(rows)


def write_tsv(transactions, output_path):
    import csv
    _calc_running_balance(transactions)
    rows = [_transaction_to_dict(t) for t in transactions]
    active_keys, active_labels = _active_headers()
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=active_keys, delimiter="\t", extrasaction='ignore')
        f.write("\t".join(active_labels) + "\n")
        writer.writerows(rows)

def write_json(transactions, output_path):
    import json as _json
    _calc_running_balance(transactions)
    active_keys, _ = _active_headers()
    rows = [{k: r[k] for k in active_keys}
            for r in (_transaction_to_dict(t) for t in transactions)]
    with open(output_path, "w", encoding="utf-8") as f:
        _json.dump(rows, f, indent=2, ensure_ascii=False)


def write_xml(transactions, output_path):
    import xml.etree.ElementTree as ET
    from xml.dom import minidom
    _calc_running_balance(transactions)
    active_keys, active_labels = _active_headers()
    root_el = ET.Element("BankStatement")
    for t in transactions:
        tx = ET.SubElement(root_el, "Transaction")
        d = _transaction_to_dict(t)
        for key, label in zip(active_keys, active_labels):
            el = ET.SubElement(tx, label.replace('.', '').replace(' ', ''))
            el.text = str(d[key]) if d[key] is not None else ""
    raw    = ET.tostring(root_el, encoding="unicode")
    pretty = minidom.parseString(raw).toprettyxml(indent="  ")
    lines  = pretty.split("\n")[1:]
    with open(output_path, "w", encoding="utf-8") as f:
        f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
        f.write("\n".join(lines))


def write_markdown(transactions, output_path):
    _calc_running_balance(transactions)
    rows = [_transaction_to_dict(t) for t in transactions]
    active_keys, active_labels = _active_headers()

    def fmt(v):
        if v is None: return ""
        if isinstance(v, float): return f"{v:,.2f}"
        return str(v).replace("|", "\\|")

    col_widths = [max(len(active_labels[i]),
                      max((len(fmt(r[h])) for r in rows), default=0))
                  for i, h in enumerate(active_keys)]

    def pad(text, w): return text.ljust(w)

    lines = [
        "# BCA Bank Statement\n",
        "| " + " | ".join(pad(h, col_widths[i]) for i, h in enumerate(active_labels)) + " |",
        "| " + " | ".join("-" * w for w in col_widths) + " |",
    ]
    for r in rows:
        lines.append("| " + " | ".join(pad(fmt(r[h]), col_widths[i]) for i, h in enumerate(active_keys)) + " |")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


def write_yaml(transactions, output_path):
    try:
        import yaml
    except ImportError:
        raise ImportError("PyYAML not installed. Run: pip install pyyaml")
    _calc_running_balance(transactions)
    active_keys, _ = _active_headers()
    rows = [{k: r[k] for k in active_keys}
            for r in (_transaction_to_dict(t) for t in transactions)]
    with open(output_path, "w", encoding="utf-8") as f:
        yaml.dump({"bank_statement": rows}, f,
                  allow_unicode=True, default_flow_style=False, sort_keys=False)


def write_rtf(transactions, output_path):
    _calc_running_balance(transactions)
    rows = [_transaction_to_dict(t) for t in transactions]
    active_keys, active_labels = _active_headers()

    def esc(s):
        return str(s).replace("\\","\\\\").replace("{","\\{").replace("}","\\}") if s else ""

    def fmt(v):
        if v is None: return ""
        if isinstance(v, float): return f"{v:,.2f}"
        return esc(str(v))

    n     = len(active_keys)
    col_w = 1600
    lines = [
        r"{\rtf1\ansi\deff0",
        r"{\fonttbl{\f0 Arial;}}",
        r"{\colortbl;\red46\green117\blue182;\red255\green255\blue255;\red235\green243\blue251;}",
        r"\f0\fs18",
    ]

    def row_rtf(cells, bold=False, bg=1):
        r = f"\\trowd\\trgaph108\\trrh260\\rowshdng10000\\trcbpat{bg}"
        for i in range(n):
            r += f"\\cellx{col_w*(i+1)}"
        r += "\n"
        for cell in cells:
            b = r"\b " if bold else ""
            r += f"\\pard\\intbl {b}{cell}\\b0\\cell\n"
        r += "\\row\n"
        return r

    lines.append(row_rtf(active_labels, bold=True, bg=1))
    for i, row in enumerate(rows):
        lines.append(row_rtf([fmt(row[h]) for h in active_keys], bg=3 if i % 2 == 0 else 2))
    lines.append("}")

    with open(output_path, "w", encoding="ascii", errors="replace") as f:
        f.write("\n".join(lines))


def write_odt(transactions, output_path):
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        from odf.style import Style, TableCellProperties, TextProperties
    except ImportError:
        raise ImportError("odfpy not installed. Run: pip install odfpy")

    _calc_running_balance(transactions)
    rows = [_transaction_to_dict(t) for t in transactions]
    active_keys, active_labels = _active_headers()
    doc  = OpenDocumentSpreadsheet()

    hdr_style = Style(name="HeaderCell", family="table-cell")
    hdr_style.addElement(TableCellProperties(backgroundcolor="#2E75B6"))
    hdr_style.addElement(TextProperties(color="#FFFFFF", fontweight="bold"))
    doc.automaticstyles.addElement(hdr_style)

    table   = Table(name="Transactions")
    hdr_row = TableRow()
    for label in active_labels:
        cell = TableCell(valuetype="string", stylename="HeaderCell")
        cell.addElement(P(text=label))
        hdr_row.addElement(cell)
    table.addElement(hdr_row)

    for row in rows:
        tr = TableRow()
        for h in active_keys:
            v = row[h]
            if isinstance(v, float):
                cell = TableCell(valuetype="float", value=str(v))
                cell.addElement(P(text=f"{v:,.2f}"))
            else:
                cell = TableCell(valuetype="string")
                cell.addElement(P(text=str(v) if v is not None else ""))
            tr.addElement(cell)
        table.addElement(tr)

    doc.spreadsheet.addElement(table)
    doc.save(output_path)


def write_parquet(transactions, output_path):
    try:
        import pandas as pd
    except ImportError:
        raise ImportError("pandas not installed. Run: pip install pandas pyarrow")
    try:
        import pyarrow  # noqa
    except ImportError:
        raise ImportError("pyarrow not installed. Run: pip install pyarrow")

    _calc_running_balance(transactions)
    active_keys, _ = _active_headers()
    rows = [{k: r[k] for k in active_keys}
            for r in (_transaction_to_dict(t) for t in transactions)]
    pd.DataFrame(rows, columns=active_keys).to_parquet(output_path, index=False)


# ════════════════════════════════════════════════════════════════════
#  GUI CONFIG & RULES
# ════════════════════════════════════════════════════════════════════

CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".bca_converter_config.json")

# Banks available for type rules
RULE_BANKS = ["All Banks"] + [a.code for a in BANK_ADAPTERS]

DEFAULT_RULES = [
    {"keyword": "SALDO AWAL",         "type": "OPENING",              "bank": "BCA", "locked": True},
    {"keyword": "BERLIAN SISTEM INF",  "type": "SALARY/REIMBURSEMENT", "bank": "BCA", "locked": False},
    {"keyword": "TARIKAN ATM",         "type": "ATM",                  "bank": "BCA", "locked": False},
    {"keyword": "BIAYA ADM",           "type": "ADMIN FEE",            "bank": "BCA", "locked": False},
    {"keyword": "FLAZZ BCA",           "type": "FLAZZ",                "bank": "BCA", "locked": False},
    {"keyword": "KARTU KREDIT",        "type": "CC PAYMENT",           "bank": "BCA", "locked": False},
    {"keyword": "KARTU DEBIT",         "type": "CARD",                 "bank": "BCA", "locked": False},
    {"keyword": "SWITCHING",           "type": "SWITCHING",            "bank": "BCA", "locked": False},
    {"keyword": "KR OTOMATIS",         "type": "AUTO CR",              "bank": "BCA", "locked": False},
    {"keyword": "DB OTOMATIS",         "type": "AUTO DB",              "bank": "BCA", "locked": False},
    {"keyword": "TRANSAKSI DEBIT",     "type": "QRIS",                 "bank": "BCA", "locked": False},
    {"keyword": "QR",                  "type": "QRIS",                 "bank": "BCA", "locked": False},
    {"keyword": "E-BANKING",           "type": "TRANSFER",             "bank": "BCA", "locked": False},
    {"keyword": "DEPOSITO",            "type": "DEPOSITO",             "bank": "BCA", "locked": False},
    {"keyword": "DB DEBIT DOMESTIK",   "type": "DEBIT",                "bank": "BCA", "locked": False},
    {"keyword": "BI-FAST",             "type": "BI-FAST",              "bank": "BCA", "locked": False},
    {"keyword": "BUNGA POKET",         "type": "INTEREST",             "bank": "BCA",  "locked": False},
    {"keyword": "TRF BERKALA",         "type": "AUTO TRANSFER",        "bank": "BCA",  "locked": False},
    {"keyword": "SETORAN AWAL",        "type": "INITIAL DEPOSIT",      "bank": "BCA",  "locked": False},
    # ── Jago rules ────────────────────────────────────────────────────
    {"keyword": "Bunga",               "type": "INTEREST",             "bank": "Jago", "locked": False},
    {"keyword": "Pajak Bunga",         "type": "TAX",                  "bank": "Jago", "locked": False},
    {"keyword": "Transfer Masuk",      "type": "TRANSFER IN",          "bank": "Jago", "locked": False},
    {"keyword": "Transfer Keluar",     "type": "TRANSFER OUT",         "bank": "Jago", "locked": False},
    {"keyword": "Pembayaran QRIS",     "type": "QRIS",                 "bank": "Jago", "locked": False},
    {"keyword": "Transaksi POS",       "type": "CARD",                 "bank": "Jago", "locked": False},
    {"keyword": "Pindah uang antar",   "type": "POCKET TRANSFER",      "bank": "Jago", "locked": False},
    {"keyword": "Pencairan Dana dari", "type": "GOPAY",                "bank": "Jago", "locked": False},
    {"keyword": "Transaksi GoPay",     "type": "GOPAY",                "bank": "Jago", "locked": False},
    {"keyword": "Gopay Savings",       "type": "GOPAY",                "bank": "Jago", "locked": False},
    {"keyword": "Jago Pay",            "type": "JAGO PAY",             "bank": "Jago", "locked": False},
    {"keyword": "Pencairan Reksa Dana","type": "MUTUAL FUND",          "bank": "Jago", "locked": False},
    {"keyword": "Tambah Uang Kantong", "type": "POCKET TRANSFER",      "bank": "Jago", "locked": False},
]

FORMAT_MAP = [
    ("xlsx",    ".xlsx",    write_excel),
    ("xml",     ".xml",     write_xml),
    ("json",    ".json",    write_json),
    ("csv",     ".csv",     write_csv),
    ("tsv",     ".tsv",     write_tsv),
    ("md",      ".md",      write_markdown),
    ("yaml",    ".yaml",    write_yaml),
    ("rtf",     ".rtf",     write_rtf),
    ("odt",     ".ods",     write_odt),
    ("parquet", ".parquet", write_parquet),
]


def load_rules():
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH) as f:
                return json.load(f).get("rules", DEFAULT_RULES)
        except Exception:
            pass
    return [dict(r) for r in DEFAULT_RULES]


def save_rules(rules):
    try:
        with open(CONFIG_PATH, "w") as f:
            json.dump({"rules": rules}, f, indent=2)
    except Exception:
        pass


def apply_rules(description, rules, bank="BCA"):
    """Match rules against description, respecting bank scope.
    A rule with bank='All Banks' matches any bank.
    A rule with a specific bank only matches that bank.
    """
    d = description.upper()
    for rule in rules:
        rule_bank = rule.get("bank", "All Banks")
        if rule_bank != "All Banks" and rule_bank != bank:
            continue
        kw = rule.get("keyword", "").upper()
        if kw and kw in d:
            return rule.get("type", "OTHER")
    return "OTHER"


# ════════════════════════════════════════════════════════════════════
#  GUI — THEMES
# ════════════════════════════════════════════════════════════════════

THEMES = {
    "light": {
        "bg":          "#F0F4F8",
        "bg2":         "#E2EAF4",
        "fg":          "#1A1A2E",
        "fg2":         "#555555",
        "topbar":      "#2E75B6",
        "topbar_fg":   "#FFFFFF",
        "topbar_sub":  "#BDD7EE",
        "accent":      "#2E75B6",
        "accent_fg":   "#FFFFFF",
        "success":     "#1A6B1A",
        "danger":      "#C00000",
        "warn_fg":     "#C87800",
        "entry_bg":    "#FFFFFF",
        "entry_fg":    "#1A1A2E",
        "list_bg":     "#FFFFFF",
        "list_sel":    "#BDD7EE",
        "log_bg":      "#1E1E1E",
        "log_fg":      "#D4D4D4",
        "status_bg":   "#D0DCE8",
        "status_fg":   "#333333",
        "tree_bg":     "#FFFFFF",
        "tree_alt":    "#EBF3FB",
        "tree_sel":    "#2E75B6",
        "tree_sel_fg": "#FFFFFF",
        "card_bg":     "#FFFFFF",
        "card_border": "#BDD7EE",
        "debit_fg":    "#C00000",
        "credit_fg":   "#1A6B1A",
        "opening_fg":  "#2E75B6",
        "sep":         "#BDD7EE",
    },
    "dark": {
        "bg":          "#1E1E2E",
        "bg2":         "#2A2A3E",
        "fg":          "#CDD6F4",
        "fg2":         "#888BA8",
        "topbar":      "#181825",
        "topbar_fg":   "#CDD6F4",
        "topbar_sub":  "#6C7086",
        "accent":      "#89B4FA",
        "accent_fg":   "#1E1E2E",
        "success":     "#A6E3A1",
        "danger":      "#F38BA8",
        "warn_fg":     "#FAB387",
        "entry_bg":    "#313244",
        "entry_fg":    "#CDD6F4",
        "list_bg":     "#313244",
        "list_sel":    "#45475A",
        "log_bg":      "#11111B",
        "log_fg":      "#CDD6F4",
        "status_bg":   "#181825",
        "status_fg":   "#888BA8",
        "tree_bg":     "#313244",
        "tree_alt":    "#2A2A3E",
        "tree_sel":    "#89B4FA",
        "tree_sel_fg": "#1E1E2E",
        "card_bg":     "#313244",
        "card_border": "#45475A",
        "debit_fg":    "#F38BA8",
        "credit_fg":   "#A6E3A1",
        "opening_fg":  "#89B4FA",
        "sep":         "#45475A",
    }
}


# ════════════════════════════════════════════════════════════════════
#  GUI
# ════════════════════════════════════════════════════════════════════

try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    _DND_AVAILABLE = True
except ImportError:
    _DND_AVAILABLE = False

_AppBase = TkinterDnD.Tk if _DND_AVAILABLE else tk.Tk

class App(_AppBase):
    def __repr__(self):
        return f"<App v{__version__}>"

    def __init__(self):
        super().__init__()
        self.title(f"BCA Statement Converter v{__version__}")
        self.resizable(True, True)
        self.minsize(1000, 660)

        cfg = {}
        if os.path.exists(CONFIG_PATH):
            try:
                cfg = json.load(open(CONFIG_PATH))
            except Exception:
                pass

        self.rules            = cfg.get("rules", [dict(r) for r in DEFAULT_RULES])
        self.merge_var        = tk.BooleanVar(value=True)
        self.fmt_var          = tk.StringVar(value="preview")
        self._dark            = cfg.get("dark_mode", False)
        self._offline         = cfg.get("offline_mode", False)
        self._fx_api_key      = EXCHANGE_RATE_API_KEY or cfg.get("fx_api_key", "") or ""
        self._last_out_folder = cfg.get("last_output_folder", "")
        self._budgets         = cfg.get("budgets", {})   # {type_label: monthly_limit}
        self._parsed_txns     = []        # last parsed transactions for preview
        self._preview_cache   = None      # pre-built search haystacks, invalidated on reload
        self._last_graph_data = None      # cached (ccy, dates, values, period) for re-plot
        self._txn_chart_ax    = None
        self._txn_chart_ylim  = None
        # Column ordering: list of _HEADERS keys in display order
        self._col_order   = cfg.get("col_order",   list(_HEADERS))
        # Ensure any new headers added since last save are included
        for h in _HEADERS:
            if h not in self._col_order:
                self._col_order.append(h)
        self._col_visible = cfg.get("col_visible", {h: True for h in _HEADERS})

        # Push to module-level so export writers use same order
        _this_module._active_col_order   = self._col_order
        _this_module._active_col_visible = self._col_visible

        self._apply_ttk_theme()
        self._build_ui()
        self._center()

    # ── theming ──────────────────────────────────────────────────────
    @property
    def T(self):
        return THEMES["dark" if self._dark else "light"]

    def _apply_ttk_theme(self):
        T  = self.T
        st = ttk.Style(self)
        st.theme_use("clam")
        st.configure(".",
            background=T["bg"], foreground=T["fg"],
            fieldbackground=T["entry_bg"], bordercolor=T["sep"],
            troughcolor=T["bg2"], selectbackground=T["tree_sel"],
            selectforeground=T["tree_sel_fg"])
        st.configure("TNotebook",         background=T["bg2"], borderwidth=0)
        st.configure("TNotebook.Tab",     background=T["bg2"], foreground=T["fg2"],
                     padding=[14,6], borderwidth=0)
        st.map("TNotebook.Tab",
               background=[("selected", T["bg"]), ("active", T["bg"])],
               foreground=[("selected", T["accent"]), ("active", T["fg"])])
        st.configure("Treeview",          background=T["tree_bg"],
                     foreground=T["fg"], fieldbackground=T["tree_bg"],
                     rowheight=22, borderwidth=0)
        st.configure("Treeview.Heading",  background=T["bg2"],
                     foreground=T["fg"], relief="flat", borderwidth=0)
        st.map("Treeview",
               background=[("selected", T["tree_sel"])],
               foreground=[("selected", T["tree_sel_fg"])])
        st.configure("TScrollbar",        background=T["bg2"],
                     troughcolor=T["bg"], borderwidth=0, arrowsize=12)
        st.configure("TProgressbar",      background=T["accent"],
                     troughcolor=T["bg2"])
        self.configure(bg=T["bg"])

    def _toggle_theme(self):
        self._dark = not self._dark
        try:
            cfg = json.load(open(CONFIG_PATH)) if os.path.exists(CONFIG_PATH) else {}
        except Exception:
            cfg = {}
        cfg["dark_mode"] = self._dark
        try:
            json.dump(cfg, open(CONFIG_PATH, "w"), indent=2)
        except Exception:
            pass
        # Rebuild entire UI
        for w in self.winfo_children():
            w.destroy()
        self._apply_ttk_theme()
        self._build_ui()

    def _toggle_offline(self):
        self._offline = not self._offline
        # Persist
        try:
            cfg = json.load(open(CONFIG_PATH)) if os.path.exists(CONFIG_PATH) else {}
        except Exception:
            cfg = {}
        cfg["offline_mode"] = self._offline
        try:
            json.dump(cfg, open(CONFIG_PATH, "w"), indent=2)
        except Exception:
            pass
        # Update button appearance without full rebuild
        T = self.T
        if self._offline:
            self._offline_btn.configure(
                text="🔴  Offline",
                bg=T["danger"], fg=T["accent_fg"],
                activebackground=T["danger"], activeforeground=T["accent_fg"])
        else:
            self._offline_btn.configure(
                text="🟢  Online",
                bg=T["topbar"], fg=T["topbar_sub"],
                activebackground=T["topbar"], activeforeground=T["topbar_sub"])
        # Update status bar
        self.status_var.set(
            "⚠ Offline mode — all network features disabled." if self._offline
            else "Online mode — network enabled.")
        # Refresh FX mode badge if already built
        if hasattr(self, "_update_fx_status"):
            self._update_fx_status()

    def _save_fx_key(self, *_):
        key = self.fx_key_var.get().strip()
        self._fx_api_key = key
        try:
            cfg = json.load(open(CONFIG_PATH)) if os.path.exists(CONFIG_PATH) else {}
        except Exception:
            cfg = {}
        cfg["fx_api_key"] = key
        try:
            json.dump(cfg, open(CONFIG_PATH, "w"), indent=2)
        except Exception:
            pass

    def _frame(self, parent, bg_key="bg", **kw):
        f = tk.Frame(parent, bg=self.T[bg_key], **kw)
        return f

    def _label(self, parent, bg_key="bg", fg_key="fg", **kw):
        l = tk.Label(parent, bg=self.T[bg_key], fg=self.T[fg_key], **kw)
        return l

    def _btn(self, parent, text, cmd, bg_key="accent", fg_key="accent_fg",
             bold=False, **kw):
        font = kw.pop("font", ("Arial", 9, "bold") if bold else ("Arial", 9))
        b = tk.Button(parent, text=text, command=cmd, font=font,
                      bg=self.T[bg_key], fg=self.T[fg_key],
                      activebackground=self.T[bg_key],
                      activeforeground=self.T[fg_key],
                      relief="flat", cursor="hand2", **kw)
        return b

    # ── layout ───────────────────────────────────────────────────────
    def _build_ui(self):
        T = self.T

        # Load logo
        self._logo_img = self._icon_img = None
        try:
            import base64 as _b64
            from io import BytesIO
            from PIL import Image, ImageTk
            raw     = _b64.b64decode(LOGO_B64)
            pil_img = Image.open(BytesIO(raw))
            self._icon_img = ImageTk.PhotoImage(pil_img.resize((32,32), Image.LANCZOS))
            self._logo_img = ImageTk.PhotoImage(pil_img.resize((34,34), Image.LANCZOS))
            self.wm_iconphoto(True, self._icon_img)
        except Exception:
            pass

        # ── top bar ──────────────────────────────────────────────────
        top = tk.Frame(self, bg=T["topbar"], padx=12, pady=8)
        top.pack(fill="x")

        if self._logo_img:
            tk.Label(top, image=self._logo_img,
                     bg=T["topbar"]).pack(side="left", padx=(0,8))

        tk.Label(top, text="BCA Statement Converter",
                 font=("Arial", 14, "bold"),
                 fg=T["topbar_fg"], bg=T["topbar"]).pack(side="left")
        tk.Label(top, text=f"v{__version__}",
                 font=("Arial", 9),
                 fg=T["topbar_sub"], bg=T["topbar"]).pack(side="left", padx=(6,0))

        # Dark mode toggle
        moon = "☀  Light" if self._dark else "🌙  Dark"
        self._btn(top, moon, self._toggle_theme,
                  bg_key="topbar", fg_key="topbar_sub",
                  font=("Arial", 9), padx=10, pady=2
                  ).pack(side="right", padx=4)

        # Offline mode toggle
        offline_lbl = "🔴  Offline" if self._offline else "🟢  Online"
        offline_cfg = dict(bg_key="danger", fg_key="accent_fg") if self._offline \
                      else dict(bg_key="topbar", fg_key="topbar_sub")
        self._offline_btn = self._btn(top, offline_lbl, self._toggle_offline,
                                      font=("Arial", 9), padx=10, pady=2,
                                      **offline_cfg)
        self._offline_btn.pack(side="right", padx=4)

        # ── notebook ─────────────────────────────────────────────────
        self._nb = ttk.Notebook(self)
        self._nb.pack(fill="both", expand=True, padx=10, pady=(8,0))

        self.tab_convert = tk.Frame(self._nb, bg=T["bg"])
        self.tab_preview = tk.Frame(self._nb, bg=T["bg"])
        self.tab_rules   = tk.Frame(self._nb, bg=T["bg"])
        self.tab_budget  = tk.Frame(self._nb, bg=T["bg"])
        self.tab_fx      = tk.Frame(self._nb, bg=T["bg"])
        self.tab_columns = tk.Frame(self._nb, bg=T["bg"])
        self._nb.add(self.tab_convert, text="  Convert  ")
        self._nb.add(self.tab_preview, text="  Preview  ")
        self._nb.add(self.tab_rules,   text="  Type Rules  ")
        self._nb.add(self.tab_budget,  text="  Budget  ")
        self._nb.add(self.tab_fx,      text="  Exchange Rates  ")
        self._nb.add(self.tab_columns, text="  Columns  ")

        self._build_convert_tab()
        self._build_preview_tab()
        self._build_rules_tab()
        self._build_budget_tab()
        self._build_fx_tab()
        self._build_columns_tab()

        # ── status bar ───────────────────────────────────────────────
        bar = tk.Frame(self, bg=T["status_bg"], height=24)
        bar.pack(fill="x", side="bottom")
        self.status_var = tk.StringVar(value="Ready.")
        tk.Label(bar, textvariable=self.status_var, font=("Arial", 9),
                 bg=T["status_bg"], fg=T["status_fg"],
                 anchor="w", padx=8).pack(fill="x")

    def _build_convert_tab(self):
        T   = self.T
        f   = self.tab_convert
        pad = dict(padx=14, pady=5)

        # ── left/right panes ─────────────────────────────────────────
        pane = tk.PanedWindow(f, orient="horizontal", bg=T["sep"],
                              sashwidth=4, sashrelief="flat")
        pane.pack(fill="both", expand=True, padx=8, pady=8)

        left_container = self._frame(pane, width=380)
        right = self._frame(pane)
        pane.add(left_container, minsize=340)
        pane.add(right,  minsize=300)

        # ── LEFT: fixed Convert bar at bottom (outside scroll) ───────
        convert_bar = tk.Frame(left_container, bg=T["bg"])
        convert_bar.pack(side="bottom", fill="x")
        tk.Frame(convert_bar, bg=T["sep"], height=1).pack(fill="x")
        btn_main = self._frame(convert_bar)
        btn_main.pack(fill="x", padx=14, pady=8)
        self.convert_btn = self._btn(btn_main, "▶  Convert",
                                     self._start_convert, bg_key="success",
                                     font=("Arial", 11, "bold"), padx=18, pady=7)
        self.convert_btn.pack(side="left")
        self.progress = ttk.Progressbar(btn_main, mode="determinate",
                                        maximum=100, length=150)
        self.progress.pack(side="left", padx=(12, 4))
        self.progress_pct = tk.StringVar(value="")
        self._label(btn_main, textvariable=self.progress_pct,
                    font=("Arial", 9, "bold"), fg_key="accent",
                    bg_key="bg").pack(side="left")

        # Time estimation row below progress
        time_row = self._frame(convert_bar)
        time_row.pack(fill="x", padx=14, pady=(0, 4))
        self.progress_time = tk.StringVar(value="")
        self._label(time_row, textvariable=self.progress_time,
                    font=("Arial", 8), fg_key="fg2",
                    bg_key="bg").pack(side="left")

        # ── LEFT: scrollable container ───────────────────────────────
        left_canvas = tk.Canvas(left_container, bg=T["bg"], highlightthickness=0, bd=0)
        left_sb     = ttk.Scrollbar(left_container, orient="vertical",
                                    command=left_canvas.yview)
        left_canvas.configure(yscrollcommand=left_sb.set)
        left_sb.pack(side="right", fill="y")
        left_canvas.pack(side="left", fill="both", expand=True)

        # Inner frame — all left-panel widgets go here
        inner = tk.Frame(left_canvas, bg=T["bg"])
        inner_win = left_canvas.create_window((0, 0), window=inner,
                                              anchor="nw", tags="inner")

        def _on_inner_config(e):
            left_canvas.configure(scrollregion=left_canvas.bbox("all"))
        def _on_canvas_resize(e):
            left_canvas.itemconfig("inner", width=e.width)
        inner.bind("<Configure>", _on_inner_config)
        left_canvas.bind("<Configure>", _on_canvas_resize)

        # Mouse-wheel scroll (Windows + Linux)
        def _on_mousewheel(e):
            left_canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        def _on_mousewheel_linux(e):
            left_canvas.yview_scroll(-1 if e.num == 4 else 1, "units")
        left_canvas.bind_all("<MouseWheel>",   _on_mousewheel)
        left_canvas.bind_all("<Button-4>",     _on_mousewheel_linux)
        left_canvas.bind_all("<Button-5>",     _on_mousewheel_linux)

        # Use `inner` as the parent for all left-panel widgets below
        left = inner

        # ── LEFT: inputs ─────────────────────────────────────────────
        # Bank selector
        bank_row = self._frame(left)
        bank_row.pack(fill="x", padx=14, pady=(10, 6))
        self._label(bank_row, text="Bank:", font=("Arial", 9, "bold"),
                    fg_key="fg2").pack(side="left", padx=(0, 8))

        self._bank_var = tk.StringVar(value=BANK_ADAPTERS[0].name)
        self._bank_map = {a.name: a.code for a in BANK_ADAPTERS}
        bank_cb = ttk.Combobox(bank_row, textvariable=self._bank_var,
                               values=[a.name for a in BANK_ADAPTERS],
                               state="readonly", width=30, font=("Arial", 9))
        bank_cb.pack(side="left")

        self._bank_badge = self._label(bank_row, text="● BCA",
                                       font=("Arial", 9, "bold"), fg_key="accent")
        self._bank_badge.pack(side="left", padx=(10, 0))

        def _on_bank_change(_=None):
            code  = self._bank_map.get(self._bank_var.get(), "BCA")
            adpt  = BANK_ADAPTER_MAP.get(code)
            badge = f"● {code}" + ("" if adpt and adpt.ready else "  ⚠ stub")
            self._bank_badge.configure(text=badge,
                fg=self.T["accent"] if adpt and adpt.ready else self.T["warn_fg"])
            if adpt and not adpt.ready and adpt.note:
                self.status_var.set(f"⚠ {adpt.name}: {adpt.note}")
        bank_cb.bind("<<ComboboxSelected>>", _on_bank_change)

        # Separator
        tk.Frame(left, bg=T["sep"], height=1).pack(fill="x", padx=14, pady=(0, 4))

        self._label(left, text="Input PDFs", font=("Arial", 9, "bold"),
                    bg_key="bg", fg_key="fg2").pack(anchor="w", padx=14, pady=(0,2))

        list_wrap = self._frame(left)
        list_wrap.pack(fill="both", expand=True, padx=14, pady=(0,6))

        self.pdf_listbox = tk.Listbox(
            list_wrap, font=("Consolas", 8), selectmode="extended", height=6,
            bg=T["list_bg"], fg=T["fg"], relief="flat", bd=0,
            selectbackground=T["list_sel"], selectforeground=T["fg"],
            highlightthickness=1, highlightbackground=T["sep"],
            highlightcolor=T["accent"])
        sb_h = ttk.Scrollbar(list_wrap, orient="horizontal",
                              command=self.pdf_listbox.xview)
        sb_v = ttk.Scrollbar(list_wrap, command=self.pdf_listbox.yview)
        self.pdf_listbox.configure(xscrollcommand=sb_h.set, yscrollcommand=sb_v.set)
        sb_v.pack(side="right", fill="y")
        self.pdf_listbox.pack(side="top", fill="both", expand=True)
        sb_h.pack(side="bottom", fill="x")

        # Drag-and-drop support
        if _DND_AVAILABLE:
            self.pdf_listbox.drop_target_register(DND_FILES)
            self.pdf_listbox.dnd_bind("<<Drop>>", self._on_drop_pdf)
            drop_hint = ""
        else:
            drop_hint = "  (install tkinterdnd2 for drag-and-drop)"
        drop_lbl_text = f"Drag & drop PDFs here{drop_hint}"
        self._label(list_wrap, text=drop_lbl_text, font=("Arial", 7),
                    fg_key="fg2").pack(anchor="w", padx=2)

        btn_row = self._frame(left)
        btn_row.pack(fill="x", padx=14, pady=(0,8))
        self._btn(btn_row, "＋ Add PDFs", self._browse_pdf,
                  padx=10, pady=4).pack(side="left", padx=(0,4))
        self._btn(btn_row, "✕ Remove", self._remove_pdf,
                  bg_key="danger", padx=8, pady=4).pack(side="left", padx=4)
        self._btn(btn_row, "Clear", self._clear_pdf,
                  bg_key="bg2", fg_key="fg2", padx=8, pady=4).pack(side="left", padx=4)

        # PDF password — no static field; asked via popup when a locked PDF is encountered
        self.pdf_pwd_var = tk.StringVar(value="")

        # Separator
        tk.Frame(left, bg=T["sep"], height=1).pack(fill="x", padx=14, pady=4)

        # Output folder
        self._label(left, text="Output folder", font=("Arial", 9, "bold"),
                    fg_key="fg2").pack(anchor="w", padx=14, pady=(2,2))
        out_row = self._frame(left)
        out_row.pack(fill="x", padx=14, pady=(0,8))
        self.out_var = tk.StringVar(value=self._last_out_folder)
        out_entry = tk.Entry(out_row, textvariable=self.out_var, font=("Arial", 9),
                             bg=T["entry_bg"], fg=T["entry_fg"], relief="flat",
                             insertbackground=T["fg"],
                             highlightthickness=1, highlightbackground=T["sep"],
                             highlightcolor=T["accent"])
        out_entry.pack(side="left", fill="x", expand=True, padx=(0,6))
        self._btn(out_row, "Browse…", self._browse_out,
                  padx=8, pady=3).pack(side="right")

        # Merge
        tk.Frame(left, bg=T["sep"], height=1).pack(fill="x", padx=14, pady=4)
        merge_row = self._frame(left)
        merge_row.pack(fill="x", padx=14)
        self.merge_var = tk.BooleanVar(value=True)
        ck = tk.Checkbutton(merge_row, text="Merge all PDFs into one file",
                            variable=self.merge_var, font=("Arial", 9),
                            bg=T["bg"], fg=T["fg"],
                            selectcolor=T["entry_bg"],
                            activebackground=T["bg"], activeforeground=T["fg"])
        ck.pack(side="left")

        # Date range filter
        tk.Frame(left, bg=T["sep"], height=1).pack(fill="x", padx=14, pady=(6,4))
        dr_hdr = self._frame(left)
        dr_hdr.pack(fill="x", padx=14)
        self._use_date_range = tk.BooleanVar(value=False)
        ck_dr = tk.Checkbutton(dr_hdr, text="Filter by date range",
                               variable=self._use_date_range,
                               font=("Arial", 9),
                               bg=T["bg"], fg=T["fg"],
                               selectcolor=T["entry_bg"],
                               activebackground=T["bg"], activeforeground=T["fg"],
                               command=self._on_date_range_toggle)
        ck_dr.pack(side="left")

        import datetime as _dt
        _today = _dt.date.today()
        _first = _today.replace(day=1)
        dr_row = self._frame(left)
        dr_row.pack(fill="x", padx=14, pady=(2,0))
        self._dr_from_var = tk.StringVar(value=_first.strftime("%d/%m/%Y"))
        self._dr_to_var   = tk.StringVar(value=_today.strftime("%d/%m/%Y"))
        self._label(dr_row, text="From:", font=("Arial", 9),
                    fg_key="fg2").pack(side="left", padx=(0,4))
        self._dr_from_e = tk.Entry(dr_row, textvariable=self._dr_from_var, width=11,
                                   font=("Arial", 9), bg=T["entry_bg"],
                                   fg=T["entry_fg"], relief="flat",
                                   insertbackground=T["fg"],
                                   highlightthickness=1, highlightbackground=T["sep"],
                                   highlightcolor=T["accent"])
        self._dr_from_e.pack(side="left", padx=(0,10))
        self._label(dr_row, text="To:", font=("Arial", 9),
                    fg_key="fg2").pack(side="left", padx=(0,4))
        self._dr_to_e = tk.Entry(dr_row, textvariable=self._dr_to_var, width=11,
                                 font=("Arial", 9), bg=T["entry_bg"],
                                 fg=T["entry_fg"], relief="flat",
                                 insertbackground=T["fg"],
                                 highlightthickness=1, highlightbackground=T["sep"],
                                 highlightcolor=T["accent"])
        self._dr_to_e.pack(side="left")
        self._on_date_range_toggle()   # set initial state
        tk.Frame(left, bg=T["sep"], height=1).pack(fill="x", padx=14, pady=6)
        self._label(left, text="Export as", font=("Arial", 9, "bold"),
                    fg_key="fg2").pack(anchor="w", padx=14, pady=(0,4))

        fmt_grid = self._frame(left)
        fmt_grid.pack(fill="x", padx=14, pady=(0,4))
        self.fmt_var = tk.StringVar(value="preview")

        # Preview Only — prominent standalone toggle at the top
        preview_row = self._frame(left)
        preview_row.pack(fill="x", padx=14, pady=(0,4))
        tk.Radiobutton(preview_row, text="👁  Preview Only  (no file saved)",
                       variable=self.fmt_var, value="preview",
                       font=("Arial", 9, "bold"),
                       bg=T["bg"], fg=T["accent"],
                       selectcolor=T["entry_bg"],
                       activebackground=T["bg"],
                       activeforeground=T["accent"]).pack(side="left")

        tk.Frame(left, bg=T["sep"], height=1).pack(fill="x", padx=14, pady=(0,4))

        formats = [(k, k.upper()) for k, _, _ in FORMAT_MAP] + [("all", "All")]
        for i, (val, label) in enumerate(formats):
            rb = tk.Radiobutton(fmt_grid, text=label, variable=self.fmt_var,
                                value=val, font=("Arial", 8),
                                bg=T["bg"], fg=T["fg"],
                                selectcolor=T["entry_bg"],
                                activebackground=T["bg"], activeforeground=T["fg"])
            rb.grid(row=i//4, column=i%4, sticky="w", padx=4, pady=1)

        # FX API key
        tk.Frame(left, bg=T["sep"], height=1).pack(fill="x", padx=14, pady=6)
        self._label(left, text="ExchangeRate-API key", font=("Arial", 9, "bold"),
                    fg_key="fg2").pack(anchor="w", padx=14, pady=(2,2))
        fx_row = self._frame(left)
        fx_row.pack(fill="x", padx=14, pady=(0,2))
        self.fx_key_var = tk.StringVar(value=self._fx_api_key)
        fx_entry = tk.Entry(fx_row, textvariable=self.fx_key_var, font=("Consolas", 8),
                            bg=T["entry_bg"], fg=T["entry_fg"], relief="flat",
                            show="*", insertbackground=T["fg"],
                            highlightthickness=1, highlightbackground=T["sep"],
                            highlightcolor=T["accent"])
        fx_entry.pack(side="left", fill="x", expand=True, padx=(0,6))
        self._btn(fx_row, "👁", lambda: fx_entry.configure(
                      show="" if fx_entry.cget("show") else "*"),
                  bg_key="bg2", fg_key="fg2", padx=6, pady=3).pack(side="right")
        fx_note_row = self._frame(left)
        fx_note_row.pack(fill="x", padx=14, pady=(0,4))
        self._label(fx_note_row,
                    text="Leave blank for free latest-rate fallback  •  Get key: exchangerate-api.com",
                    font=("Arial", 7), fg_key="fg2").pack(anchor="w")

        # FX rate mode status indicator
        fx_status_row = self._frame(left)
        fx_status_row.pack(fill="x", padx=14, pady=(0,6))
        self._label(fx_status_row, text="FX Rate Mode:",
                    font=("Arial", 8, "bold"), fg_key="fg2").pack(side="left", padx=(0,6))
        self._fx_mode_var = tk.StringVar()
        self._fx_mode_lbl = self._label(fx_status_row, textvariable=self._fx_mode_var,
                                        font=("Arial", 8, "bold"), fg_key="fg2")
        self._fx_mode_lbl.pack(side="left")

        def _update_fx_status(*_):
            if self._offline:
                self._fx_mode_var.set("🔴  Offline — no FX applied")
                self._fx_mode_lbl.configure(fg=self.T["danger"])
            elif self.fx_key_var.get().strip():
                self._fx_mode_var.set("🟢  Exact transaction date rate")
                self._fx_mode_lbl.configure(fg=self.T["success"])
            else:
                self._fx_mode_var.set("🟡  Today's rate (no API key)")
                self._fx_mode_lbl.configure(fg=self.T["warn_fg"])

        self._update_fx_status = _update_fx_status
        self.fx_key_var.trace_add("write", _update_fx_status)
        self.fx_key_var.trace_add("write", self._save_fx_key)
        _update_fx_status()   # set initial state

        # ── RIGHT: log ───────────────────────────────────────────────
        self._label(right, text="Log", font=("Arial", 9, "bold"),
                    fg_key="fg2").pack(anchor="w", padx=14, pady=(10,2))

        log_wrap = self._frame(right)
        log_wrap.pack(fill="both", expand=True, padx=14, pady=(0,12))

        self.log = tk.Text(log_wrap, font=("Consolas", 9),
                           bg=T["log_bg"], fg=T["log_fg"], relief="flat",
                           insertbackground=T["log_fg"],
                           state="disabled", wrap="word", bd=0,
                           highlightthickness=1, highlightbackground=T["sep"])
        # Color tags for log
        self.log.tag_configure("ok",    foreground="#A6E3A1")
        self.log.tag_configure("err",   foreground="#F38BA8")
        self.log.tag_configure("warn",  foreground="#FAB387")
        self.log.tag_configure("info",  foreground="#89B4FA")

        sb_log = ttk.Scrollbar(log_wrap, command=self.log.yview)
        self.log.configure(yscrollcommand=sb_log.set)
        self.log.pack(side="left", fill="both", expand=True)
        sb_log.pack(side="right", fill="y")

    def _build_preview_tab(self):
        T = self.T
        f = self.tab_preview

        # ── summary cards ────────────────────────────────────────────
        card_row = self._frame(f)
        card_row.pack(fill="x", padx=14, pady=(12,8))

        self._summary_cards = {}
        for key, label, icon in [
            ("total",   "Transactions", "📋"),
            ("debit",   "Total Debit",  "📤"),
            ("credit",  "Total Credit", "📥"),
            ("balance", "Net",          "💰"),
            ("accounts","Accounts",     "🏦"),
        ]:
            card = tk.Frame(card_row, bg=T["card_bg"], relief="flat",
                            highlightthickness=1, highlightbackground=T["card_border"])
            card.pack(side="left", fill="both", expand=True, padx=4)

            tk.Label(card, text=f"{icon}  {label}", font=("Arial", 8),
                     bg=T["card_bg"], fg=T["fg2"]).pack(anchor="w", padx=10, pady=(8,2))
            val_lbl = tk.Label(card, text="—", font=("Arial", 11, "bold"),
                               bg=T["card_bg"], fg=T["fg"])
            val_lbl.pack(anchor="w", padx=10, pady=(0,8))
            self._summary_cards[key] = val_lbl

        # ── filter bar ───────────────────────────────────────────────
        fbar = self._frame(f, bg_key="bg2")
        fbar.pack(fill="x", padx=14, pady=(0,6))

        self._label(fbar, text="🔍", font=("Arial", 11), bg_key="bg2",
                    fg_key="fg2").pack(side="left", padx=(10,4), pady=6)

        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self._filter_preview())
        search_e = tk.Entry(fbar, textvariable=self.search_var, font=("Arial", 9),
                            bg=T["entry_bg"], fg=T["entry_fg"], relief="flat",
                            width=28, insertbackground=T["fg"],
                            highlightthickness=1, highlightbackground=T["sep"])
        search_e.pack(side="left", padx=(0,12), pady=6)
        self._label(fbar, text="Type:", font=("Arial", 9),
                    bg_key="bg2", fg_key="fg2").pack(side="left", padx=(0,4))

        self.filter_type = tk.StringVar(value="All")
        self._type_menu = ttk.Combobox(fbar, textvariable=self.filter_type,
                                       state="readonly", width=16, font=("Arial", 9))
        self._type_menu["values"] = ["All"]
        self._type_menu.bind("<<ComboboxSelected>>", lambda _: self._filter_preview())
        self._type_menu.pack(side="left", padx=(0,12))

        self._label(fbar, text="DC:", font=("Arial", 9),
                    bg_key="bg2", fg_key="fg2").pack(side="left", padx=(0,4))
        self.filter_dc = tk.StringVar(value="All")
        dc_menu = ttk.Combobox(fbar, textvariable=self.filter_dc,
                               state="readonly", width=8, font=("Arial", 9))
        dc_menu["values"] = ["All", "DB", "CR"]
        dc_menu.bind("<<ComboboxSelected>>", lambda _: self._filter_preview())
        dc_menu.pack(side="left", padx=(0,12))

        self._btn(fbar, "✕ Clear", self._clear_filter,
                  bg_key="bg2", fg_key="fg2", padx=8, pady=3).pack(side="left")

        # Row count label
        self.preview_count = tk.StringVar(value="")
        self._label(fbar, textvariable=self.preview_count, font=("Arial", 8),
                    bg_key="bg2", fg_key="fg2").pack(side="right", padx=10)

        # ── transaction table ─────────────────────────────────────────
        tree_wrap = self._frame(f)
        tree_wrap.pack(fill="both", expand=True, padx=14, pady=(0,12))

        active = [h for h in self._col_order if self._col_visible.get(h, True)]
        label_map = dict(zip(_HEADERS, _HEADER_LABELS))
        width_map  = dict(zip(_HEADERS,
                               [60,110,80,160,88,70,100,90,110,100,260,110,50,110,110,110,110]))
        cols = active
        self.preview_tree = ttk.Treeview(tree_wrap, columns=cols,
                                          show="headings", selectmode="extended")
        for col in cols:
            self.preview_tree.heading(col, text=label_map.get(col, col),
                command=lambda c=col: self._sort_preview(c))
            self.preview_tree.column(col, width=width_map.get(col, 100),
                                     minwidth=50, stretch=(col=="keterangan"))

        self.preview_tree.tag_configure("db",      foreground=T["debit_fg"])
        self.preview_tree.tag_configure("cr",      foreground=T["credit_fg"])
        self.preview_tree.tag_configure("opening", foreground=T["opening_fg"])
        self.preview_tree.tag_configure("alt",     background=T["tree_alt"])

        # Right-click context menu
        self._preview_menu = tk.Menu(self, tearoff=0)
        self._preview_menu.add_command(label="Export selected rows…",
                                       command=self._export_selection)
        self._preview_menu.add_command(label="Copy cell value",
                                       command=self._copy_cell)
        self.preview_tree.bind("<Button-3>", self._show_preview_menu)
        self.preview_tree.bind("<Button-2>", self._show_preview_menu)  # macOS

        sb_px = ttk.Scrollbar(tree_wrap, orient="horizontal",
                              command=self.preview_tree.xview)
        sb_py = ttk.Scrollbar(tree_wrap, command=self.preview_tree.yview)
        self.preview_tree.configure(xscrollcommand=sb_px.set, yscrollcommand=sb_py.set)
        sb_py.pack(side="right", fill="y")
        self.preview_tree.pack(side="top", fill="both", expand=True)
        sb_px.pack(side="bottom", fill="x")

        self._sort_col = None
        self._sort_asc = True

    def _build_rules_tab(self):
        T = self.T
        f = self.tab_rules

        self._label(f, text="Keyword → Type mapping  (top-to-bottom, first match wins)",
                    font=("Arial", 9), fg_key="fg2").pack(anchor="w", padx=14, pady=(10,4))

        # ── filter bar ───────────────────────────────────────────────
        fbar = self._frame(f, bg_key="bg2")
        fbar.pack(fill="x", padx=14, pady=(0,4))

        self._label(fbar, text="Bank:", font=("Arial", 9),
                    bg_key="bg2", fg_key="fg2").pack(side="left", padx=(10,4), pady=6)
        self._rules_bank_var = tk.StringVar(value="All Banks")
        bank_cb = ttk.Combobox(fbar, textvariable=self._rules_bank_var,
                               values=RULE_BANKS, state="readonly",
                               width=14, font=("Arial", 9))
        bank_cb.bind("<<ComboboxSelected>>", lambda _: self._populate_tree())
        bank_cb.pack(side="left", padx=(0,14), pady=6)

        self._label(fbar, text="🔍", font=("Arial", 11),
                    bg_key="bg2", fg_key="fg2").pack(side="left", padx=(0,4), pady=6)
        self._rules_search_var = tk.StringVar()
        self._rules_search_var.trace_add("write", lambda *_: self._populate_tree())
        tk.Entry(fbar, textvariable=self._rules_search_var, font=("Arial", 9),
                 bg=T["entry_bg"], fg=T["entry_fg"], relief="flat",
                 width=22, insertbackground=T["fg"],
                 highlightthickness=1, highlightbackground=T["sep"],
                 highlightcolor=T["accent"]).pack(side="left", padx=(0,10), pady=6)

        self._rules_count_var = tk.StringVar(value="")
        self._label(fbar, textvariable=self._rules_count_var, font=("Arial", 8),
                    bg_key="bg2", fg_key="fg2").pack(side="right", padx=10)

        # ── treeview ─────────────────────────────────────────────────
        tree_frame = self._frame(f)
        tree_frame.pack(fill="both", expand=True, padx=14, pady=(0,4))

        self.tree = ttk.Treeview(tree_frame,
                                  columns=("bank", "keyword", "type"),
                                  show="headings", selectmode="browse", height=18)
        self.tree.heading("bank",    text="Bank")
        self.tree.heading("keyword", text="Keyword (contains)")
        self.tree.heading("type",    text="Type Label")
        self.tree.column("bank",    width=90,  minwidth=60, anchor="center")
        self.tree.column("keyword", width=260, minwidth=120)
        self.tree.column("type",    width=200, minwidth=80)

        sb = ttk.Scrollbar(tree_frame, command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        self._populate_tree()

        btn_row = self._frame(f)
        btn_row.pack(fill="x", padx=14, pady=(0,10))
        self._btn(btn_row, "＋ Add",    self._add_rule,    padx=10, pady=4).pack(side="left", padx=(0,4))
        self._btn(btn_row, "✎ Edit",   self._edit_rule,   padx=10, pady=4).pack(side="left", padx=4)
        self._btn(btn_row, "✕ Delete", self._delete_rule, bg_key="danger", padx=10, pady=4).pack(side="left", padx=4)
        self._btn(btn_row, "▲ Up",     self._move_up,     padx=10, pady=4).pack(side="left", padx=(14,4))
        self._btn(btn_row, "▼ Down",   self._move_down,   padx=10, pady=4).pack(side="left", padx=4)
        self._btn(btn_row, "↺ Reset",  self._reset_rules, bg_key="bg2", fg_key="fg2", padx=10, pady=4).pack(side="right")

    # ── FX dashboard ─────────────────────────────────────────────────
    def _build_columns_tab(self):
        T   = self.T
        f   = self.tab_columns

        # ── header ───────────────────────────────────────────────────
        hdr = self._frame(f)
        hdr.pack(fill="x", padx=14, pady=(14, 4))
        self._label(hdr, text="Column Order & Visibility",
                    font=("Arial", 11, "bold"), fg_key="fg").pack(side="left")
        self._label(hdr,
                    text="Drag with ▲▼ buttons · uncheck to hide from preview and export",
                    font=("Arial", 8), fg_key="fg2").pack(side="left", padx=(14, 0))

        tk.Frame(f, bg=T["sep"], height=1).pack(fill="x", padx=14, pady=(6, 0))

        # ── main layout: list on left, buttons on right ───────────────
        body = self._frame(f)
        body.pack(fill="both", expand=True, padx=14, pady=10)

        # Listbox with checkmark state
        list_frame = self._frame(body)
        list_frame.pack(side="left", fill="both", expand=True)

        self._col_lb = tk.Listbox(
            list_frame, font=("Arial", 10), selectmode="single",
            bg=T["list_bg"], fg=T["fg"], relief="flat", bd=0,
            selectbackground=T["list_sel"], selectforeground=T["fg"],
            activestyle="none",
            highlightthickness=1, highlightbackground=T["sep"],
            highlightcolor=T["accent"])
        sb = ttk.Scrollbar(list_frame, command=self._col_lb.yview)
        self._col_lb.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self._col_lb.pack(fill="both", expand=True)

        # Button panel
        btn_panel = self._frame(body)
        btn_panel.pack(side="left", fill="y", padx=(10, 0))

        self._btn(btn_panel, "▲  Move Up",   self._col_move_up,
                  padx=12, pady=5).pack(fill="x", pady=2)
        self._btn(btn_panel, "▼  Move Down", self._col_move_down,
                  padx=12, pady=5).pack(fill="x", pady=2)

        tk.Frame(btn_panel, bg=T["sep"], height=1).pack(fill="x", pady=8)

        self._btn(btn_panel, "☑  Show",  self._col_show,
                  bg_key="success", padx=12, pady=5).pack(fill="x", pady=2)
        self._btn(btn_panel, "☐  Hide",  self._col_hide,
                  bg_key="danger",  padx=12, pady=5).pack(fill="x", pady=2)

        tk.Frame(btn_panel, bg=T["sep"], height=1).pack(fill="x", pady=8)

        self._btn(btn_panel, "⟳  Reset",  self._col_reset,
                  bg_key="bg2", fg_key="fg2", padx=12, pady=5).pack(fill="x", pady=2)

        tk.Frame(btn_panel, bg=T["sep"], height=1).pack(fill="x", pady=8)

        self._btn(btn_panel, "✔  Apply",  self._col_apply,
                  bg_key="accent", fg_key="accent_fg",
                  font=("Arial", 9, "bold"), padx=12, pady=6).pack(fill="x", pady=2)

        self._label(btn_panel,
                    text="Apply refreshes\npreview & export",
                    font=("Arial", 7), fg_key="fg2", bg_key="bg").pack(pady=(4,0))

        # Populate
        self._col_lb_refresh()

    def _col_lb_refresh(self):
        """Redraw the columns listbox from _col_order and _col_visible."""
        lb  = self._col_lb
        sel = lb.curselection()
        lb.delete(0, "end")
        label_map = dict(zip(_HEADERS, _HEADER_LABELS))
        for h in self._col_order:
            tick = "☑" if self._col_visible.get(h, True) else "☐"
            lb.insert("end", f"  {tick}  {label_map.get(h, h)}")
        # restore selection
        if sel:
            idx = min(sel[0], lb.size()-1)
            lb.selection_set(idx)
            lb.see(idx)

    def _col_selected_idx(self):
        sel = self._col_lb.curselection()
        return sel[0] if sel else None

    def _col_move_up(self):
        idx = self._col_selected_idx()
        if idx is None or idx == 0: return
        o = self._col_order
        o[idx-1], o[idx] = o[idx], o[idx-1]
        self._col_lb_refresh()
        self._col_lb.selection_clear(0, "end")
        self._col_lb.selection_set(idx-1)
        self._col_lb.see(idx-1)

    def _col_move_down(self):
        idx = self._col_selected_idx()
        if idx is None or idx >= len(self._col_order)-1: return
        o = self._col_order
        o[idx], o[idx+1] = o[idx+1], o[idx]
        self._col_lb_refresh()
        self._col_lb.selection_clear(0, "end")
        self._col_lb.selection_set(idx+1)
        self._col_lb.see(idx+1)

    def _col_show(self):
        idx = self._col_selected_idx()
        if idx is None: return
        self._col_visible[self._col_order[idx]] = True
        self._col_lb_refresh()
        self._col_lb.selection_clear(0, "end")
        self._col_lb.selection_set(idx)

    def _col_hide(self):
        idx = self._col_selected_idx()
        if idx is None: return
        self._col_visible[self._col_order[idx]] = False
        self._col_lb_refresh()
        self._col_lb.selection_clear(0, "end")
        self._col_lb.selection_set(idx)

    def _col_reset(self):
        self._col_order   = list(_HEADERS)
        self._col_visible = {h: True for h in _HEADERS}
        self._col_lb_refresh()

    def _col_apply(self):
        """Save column config and refresh preview + export order."""
        self._save_col_config()
        # Push to module-level so writers respect order
        _this_module._active_col_order   = self._col_order
        _this_module._active_col_visible = self._col_visible
        # Rebuild preview treeview columns
        tree = self.preview_tree
        active = [h for h in self._col_order if self._col_visible.get(h, True)]
        label_map = dict(zip(_HEADERS, _HEADER_LABELS))
        width_map  = dict(zip(_HEADERS,
                               [60,110,80,160,88,70,100,90,110,100,260,110,50,110,110,110,110]))
        tree.configure(columns=active)
        for col in active:
            tree.heading(col, text=label_map.get(col, col),
                         command=lambda c=col: self._sort_preview(c))
            tree.column(col, width=width_map.get(col, 100),
                        minwidth=50, stretch=(col == "keterangan"))
        # Re-render existing data if any
        if self._parsed_txns:
            self._filter_preview()
        self.status_var.set("✔ Column layout applied.")

    def _save_col_config(self):
        try:
            cfg = json.load(open(CONFIG_PATH)) if os.path.exists(CONFIG_PATH) else {}
        except Exception:
            cfg = {}
        cfg["col_order"]   = self._col_order
        cfg["col_visible"] = self._col_visible
        try:
            json.dump(cfg, open(CONFIG_PATH, "w"), indent=2)
        except Exception:
            pass

    def _build_fx_tab(self):
        import datetime
        T = self.T
        f = self.tab_fx

        COMMON_CCY = ["USD","EUR","GBP","JPY","SGD","AUD","CNY","HKD","MYR","SAR","KRW","CHF","CAD"]

        # ── top control bar (rate table) ─────────────────────────────
        ctrl = self._frame(f, bg_key="bg2")
        ctrl.pack(fill="x", padx=14, pady=(12, 6))

        self._label(ctrl, text="Foreign → IDR", font=("Arial", 10, "bold"),
                    bg_key="bg2", fg_key="accent").pack(side="left", padx=(12, 12), pady=8)

        self._fx_mode_var = tk.StringVar(value="today")
        tk.Radiobutton(ctrl, text="Today", variable=self._fx_mode_var, value="today",
                       font=("Arial", 9), bg=T["bg2"], fg=T["fg"],
                       selectcolor=T["entry_bg"], activebackground=T["bg2"],
                       command=self._on_fx_mode_change).pack(side="left", padx=4, pady=8)
        tk.Radiobutton(ctrl, text="Historical date:", variable=self._fx_mode_var, value="hist",
                       font=("Arial", 9), bg=T["bg2"], fg=T["fg"],
                       selectcolor=T["entry_bg"], activebackground=T["bg2"],
                       command=self._on_fx_mode_change).pack(side="left", padx=4, pady=8)

        date_frame = self._frame(ctrl, bg_key="bg2")
        date_frame.pack(side="left", padx=(0, 12), pady=8)
        today = datetime.date.today()
        self._fx_day_var = tk.StringVar(value=f"{today.day:02d}")
        self._fx_mon_var = tk.StringVar(value=f"{today.month:02d}")
        self._fx_yr_var  = tk.StringVar(value=str(today.year))

        def _date_entry(parent, var, width, label):
            tk.Entry(parent, textvariable=var, width=width, font=("Arial", 9),
                     bg=T["entry_bg"], fg=T["entry_fg"], relief="flat", justify="center",
                     insertbackground=T["fg"],
                     highlightthickness=1, highlightbackground=T["sep"],
                     highlightcolor=T["accent"]).pack(side="left")
            self._label(parent, text=label, font=("Arial", 9),
                        bg_key="bg2", fg_key="fg2").pack(side="left", padx=(1, 4))

        _date_entry(date_frame, self._fx_day_var, 3, "/")
        _date_entry(date_frame, self._fx_mon_var, 3, "/")
        _date_entry(date_frame, self._fx_yr_var,  5, "")
        self._fx_date_frame = date_frame

        self._btn(ctrl, "🔄  Fetch Rates", self._fetch_fx_dashboard,
                  padx=12, pady=5).pack(side="left", padx=8, pady=8)

        self._fx_updated_var = tk.StringVar(value="")
        self._label(ctrl, textvariable=self._fx_updated_var, font=("Arial", 8),
                    bg_key="bg2", fg_key="fg2").pack(side="right", padx=12)

        self._on_fx_mode_change()

        # ── summary cards ────────────────────────────────────────────
        self._fx_rate_cards = {}
        for row_ccys in [COMMON_CCY[:7], COMMON_CCY[7:]]:
            strip = self._frame(f)
            strip.pack(fill="x", padx=14, pady=(0, 3))
            for ccy in row_ccys:
                card = tk.Frame(strip, bg=T["card_bg"],
                                highlightthickness=1, highlightbackground=T["card_border"],
                                cursor="hand2")
                card.pack(side="left", fill="both", expand=True, padx=2)
                self._label(card, text=ccy, font=("Arial", 8, "bold"),
                            bg_key="card_bg", fg_key="fg2").pack(anchor="w", padx=7, pady=(5,1))
                val = self._label(card, text="—", font=("Arial", 10, "bold"),
                                  bg_key="card_bg", fg_key="fg")
                val.pack(anchor="w", padx=7, pady=(0,5))
                self._fx_rate_cards[ccy] = val
                # clicking a card selects that currency in the graph
                for widget in (card, val):
                    widget.bind("<Button-1>", lambda e, c=ccy: self._select_graph_ccy(c))

        # ── main split: table (left) | graph (right) ─────────────────
        pane = tk.PanedWindow(f, orient="horizontal", bg=T["sep"],
                              sashwidth=4, sashrelief="flat")
        pane.pack(fill="both", expand=True, padx=14, pady=(4, 4))

        left_pane  = self._frame(pane)
        right_pane = self._frame(pane)
        pane.add(left_pane,  minsize=280)
        pane.add(right_pane, minsize=340)

        # Left: full rates table
        cols = ("currency", "name", "rate_idr", "buy_note", "sell_note")
        hdrs = ("CCY", "Description", "Mid Rate (IDR)", "Buy*", "Sell*")
        self._fx_tree = ttk.Treeview(left_pane, columns=cols,
                                      show="headings", selectmode="browse")
        for col, hdr, w in zip(cols, hdrs, [55, 155, 130, 100, 100]):
            self._fx_tree.heading(col, text=hdr)
            self._fx_tree.column(col, width=w, minwidth=40,
                                 anchor="center" if col != "name" else "w")
        self._fx_tree.tag_configure("alt", background=T["tree_alt"])
        self._fx_tree.bind("<<TreeviewSelect>>", self._on_fx_tree_select)

        sb_tbl = ttk.Scrollbar(left_pane, command=self._fx_tree.yview)
        self._fx_tree.configure(yscrollcommand=sb_tbl.set)
        self._fx_tree.pack(side="left", fill="both", expand=True)
        sb_tbl.pack(side="right", fill="y")

        # Right: graph panel
        graph_ctrl = self._frame(right_pane, bg_key="bg2")
        graph_ctrl.pack(fill="x", padx=6, pady=(0, 4))

        self._graph_ccy_var = tk.StringVar(value="USD")
        self._label(graph_ctrl, text="Currency:", font=("Arial", 9),
                    bg_key="bg2", fg_key="fg2").pack(side="left", padx=(8, 4), pady=6)
        graph_ccy_cb = ttk.Combobox(graph_ctrl, textvariable=self._graph_ccy_var,
                                     values=COMMON_CCY, state="readonly",
                                     width=7, font=("Arial", 9))
        graph_ccy_cb.pack(side="left", padx=(0, 10), pady=6)

        # Period quick-select buttons
        self._graph_period_var = tk.StringVar(value="1M")
        for period in ["7D","1M","3M","6M","1Y"]:
            btn = tk.Button(graph_ctrl, text=period,
                            font=("Arial", 8), relief="flat", cursor="hand2",
                            bg=T["accent"] if period=="1M" else T["bg2"],
                            fg=T["accent_fg"] if period=="1M" else T["fg2"],
                            padx=6, pady=3,
                            command=lambda p=period: self._set_graph_period(p))
            btn.pack(side="left", padx=2, pady=6)
            btn._period = period
        self._period_btns = [w for w in graph_ctrl.winfo_children()
                              if isinstance(w, tk.Button)]

        self._btn(graph_ctrl, "📈  Plot", self._fetch_graph_data,
                  padx=10, pady=3).pack(side="left", padx=(8,4), pady=6)

        # Graph mode toggle
        tk.Frame(graph_ctrl, bg=T["sep"], width=1).pack(side="left", fill="y", padx=6, pady=4)
        self._label(graph_ctrl, text="Show:", font=("Arial", 8),
                    bg_key="bg2", fg_key="fg2").pack(side="left", padx=(0, 4))
        self._graph_mode_var = tk.StringVar(value="rate")
        for val, label in [("rate", "Rate"), ("amount", "Txn Amounts")]:
            tk.Radiobutton(graph_ctrl, text=label, variable=self._graph_mode_var,
                           value=val, font=("Arial", 8),
                           bg=T["bg2"], fg=T["fg"],
                           selectcolor=T["entry_bg"],
                           activebackground=T["bg2"],
                           command=self._on_graph_mode_change
                           ).pack(side="left", padx=2)

        self._graph_status_var = tk.StringVar(value="Select a currency and click Plot")
        self._label(graph_ctrl, textvariable=self._graph_status_var, font=("Arial", 7),
                    bg_key="bg2", fg_key="fg2").pack(side="right", padx=8)

        # Y-axis scale slider (only shown in Txn Amounts mode)
        self._scale_row = self._frame(right_pane, bg_key="bg2")
        self._scale_row.pack(fill="x", padx=6, pady=(0, 2))
        self._label(self._scale_row, text="Y Scale:", font=("Arial", 8),
                    bg_key="bg2", fg_key="fg2").pack(side="left", padx=(8, 4))
        self._scale_var = tk.IntVar(value=100)
        self._scale_slider = ttk.Scale(self._scale_row, from_=5, to=100,
                                       orient="horizontal", length=180,
                                       variable=self._scale_var,
                                       command=self._on_scale_change)
        self._scale_slider.pack(side="left", padx=(0, 6))
        self._scale_lbl = self._label(self._scale_row, text="100%",
                                      font=("Arial", 8, "bold"), fg_key="accent",
                                      bg_key="bg2")
        self._scale_lbl.pack(side="left")
        self._btn(self._scale_row, "⟳", lambda: [self._scale_var.set(100),
                                                   self._on_scale_change("100")],
                  bg_key="bg2", fg_key="fg2", padx=5, pady=2).pack(side="left", padx=(6,0))
        self._scale_row.pack_forget()   # hidden until Txn Amounts mode

        # Matplotlib canvas placeholder
        self._graph_canvas_frame = self._frame(right_pane)
        self._graph_canvas_frame.pack(fill="both", expand=True, padx=6, pady=(0,4))
        self._graph_canvas  = None   # FigureCanvasTkAgg, created on first plot
        self._graph_toolbar = None

        note = self._label(f,
            text="* Buy/Sell ±0.5% est.  Graph data: Fawaz Ahmed Currency API (free, CDN-hosted, no key).  Snapshot: ExchangeRate-API paid key.",
            font=("Arial", 7), fg_key="fg2")
        note.pack(anchor="w", padx=16, pady=(0, 3))

    def _on_fx_mode_change(self):
        hist  = self._fx_mode_var.get() == "hist"
        state = "normal" if hist else "disabled"
        for child in self._fx_date_frame.winfo_children():
            try: child.configure(state=state)
            except Exception: pass

    def _get_fx_date_str(self):
        import datetime
        if self._fx_mode_var.get() == "today":
            d = datetime.date.today()
            return f"{d.day:02d}/{d.month:02d}/{d.year}"
        try:
            day = int(self._fx_day_var.get())
            mon = int(self._fx_mon_var.get())
            yr  = int(self._fx_yr_var.get())
            datetime.date(yr, mon, day)
            return f"{day:02d}/{mon:02d}/{yr}"
        except Exception:
            return None

    _CCY_NAMES = {
        "USD":"US Dollar","EUR":"Euro","GBP":"British Pound","JPY":"Japanese Yen",
        "SGD":"Singapore Dollar","AUD":"Australian Dollar","CNY":"Chinese Yuan",
        "HKD":"Hong Kong Dollar","MYR":"Malaysian Ringgit","SAR":"Saudi Riyal",
        "KRW":"South Korean Won","CHF":"Swiss Franc","CAD":"Canadian Dollar",
        "NZD":"New Zealand Dollar","THB":"Thai Baht","PHP":"Philippine Peso",
        "INR":"Indian Rupee","BRL":"Brazilian Real","TRY":"Turkish Lira",
        "ZAR":"South African Rand","AED":"UAE Dirham","QAR":"Qatari Riyal",
        "KWD":"Kuwaiti Dinar","BHD":"Bahraini Dinar","DKK":"Danish Krone",
        "SEK":"Swedish Krona","NOK":"Norwegian Krone","TWD":"Taiwan Dollar",
        "VND":"Vietnamese Dong","PKR":"Pakistani Rupee",
    }

    def _fetch_fx_dashboard(self):
        if self._offline:
            messagebox.showinfo("Offline", "Offline mode is enabled.\nDisable it via the 🔴 Offline button to fetch rates.")
            return
        import urllib.request
        date_str = self._get_fx_date_str()
        if date_str is None:
            messagebox.showwarning("Invalid date", "Please enter a valid date (DD/MM/YYYY).")
            return
        api_key = self.fx_key_var.get().strip()
        self._fx_updated_var.set("Fetching…")
        self._fx_tree.delete(*self._fx_tree.get_children())
        for card in self._fx_rate_cards.values():
            card.configure(text="…")

        def worker():
            try:
                data         = None
                source_label = ""
                if api_key:
                    parsed = _parse_date(date_str)
                    if parsed:
                        y, m, d = parsed
                        url = (f"https://v6.exchangerate-api.com/v6/{api_key}"
                               f"/history/IDR/{y}/{m}/{d}")
                        try:
                            with urllib.request.urlopen(url, timeout=10) as r:
                                data = json.loads(r.read())
                            if data.get('result') != 'success':
                                data = None
                            else:
                                source_label = f"Historical  {date_str}"
                        except Exception:
                            data = None
                if data is None:
                    url = "https://open.er-api.com/v6/latest/IDR"
                    with urllib.request.urlopen(url, timeout=10) as r:
                        data = json.loads(r.read())
                    if data.get('result') != 'success':
                        raise ValueError(f"API error: {data.get('error-type','unknown')}")
                    ts = data.get('time_last_update_utc', '')
                    source_label = f"Latest  ({ts[:16] if ts else 'now'})"

                idr_rates_inv = data.get('conversion_rates', data.get('rates', {}))
                rows = []
                for ccy, idr_per_unit_inv in sorted(idr_rates_inv.items()):
                    if ccy == 'IDR' or idr_per_unit_inv == 0:
                        continue
                    idr_per_ccy = round(1.0 / idr_per_unit_inv, 4)
                    name        = self._CCY_NAMES.get(ccy, "")
                    buy         = round(idr_per_ccy * 0.995, 2)
                    sell        = round(idr_per_ccy * 1.005, 2)
                    rows.append((ccy, name, idr_per_ccy, buy, sell))
                self.after(0, self._populate_fx_table, rows, source_label)
            except Exception as e:
                self.after(0, self._fx_updated_var.set, f"Error: {e}")
                self.after(0, messagebox.showerror, "FX Error", str(e))

        threading.Thread(target=worker, daemon=True).start()

    def _populate_fx_table(self, rows, source_label):
        import datetime
        tree = self._fx_tree
        tree.delete(*tree.get_children())
        for ccy, card_lbl in self._fx_rate_cards.items():
            match = next((r for r in rows if r[0] == ccy), None)
            if match:
                rate = match[2]
                txt = (f"{rate:,.0f}" if rate >= 1000 else
                       f"{rate:,.2f}" if rate >= 1 else f"{rate:.6f}")
                card_lbl.configure(text=txt)
            else:
                card_lbl.configure(text="—")
        for i, (ccy, name, rate, buy, sell) in enumerate(rows):
            fmt_r = lambda v: (f"{v:,.0f}" if v >= 1000 else
                               f"{v:,.2f}" if v >= 1 else f"{v:.6f}")
            tree.insert("", "end",
                        values=(ccy, name, fmt_r(rate), fmt_r(buy), fmt_r(sell)),
                        tags=("alt",) if i % 2 else ())
        now = datetime.datetime.now().strftime("%H:%M:%S")
        self._fx_updated_var.set(f"{source_label}  ·  fetched {now}")

    def _on_fx_tree_select(self, event):
        sel = self._fx_tree.selection()
        if sel:
            ccy = self._fx_tree.item(sel[0])["values"][0]
            self._select_graph_ccy(str(ccy))

    def _select_graph_ccy(self, ccy):
        self._graph_ccy_var.set(ccy)
        self._fetch_graph_data()

    def _set_graph_period(self, period):
        self._graph_period_var.set(period)
        # Re-colour period buttons
        T = self.T
        for btn in self._period_btns:
            active = (btn._period == period)
            btn.configure(
                bg=T["accent"] if active else T["bg2"],
                fg=T["accent_fg"] if active else T["fg2"])
        self._fetch_graph_data()

    def _on_graph_mode_change(self):
        mode = self._graph_mode_var.get()
        if mode == "amount":
            self._scale_row.pack(fill="x", padx=6, pady=(0, 2),
                                 before=self._graph_canvas_frame)
            self._plot_txn_amounts()
        else:
            self._scale_row.pack_forget()
            if hasattr(self, '_last_graph_data') and self._last_graph_data:
                ccy, dates, values, period = self._last_graph_data
                self._draw_graph(ccy, dates, values, period)

    def _on_scale_change(self, val=None):
        """Reapply Y-axis limits on the existing bar chart without re-fetching."""
        try:
            pct = int(float(val)) if val is not None else int(self._scale_var.get())
        except (TypeError, ValueError):
            return
        self._scale_lbl.configure(text=f"{pct}%")
        if self._txn_chart_ax is None or self._txn_chart_ylim is None:
            return
        ax = self._txn_chart_ax
        full_lo, full_hi = self._txn_chart_ylim
        ratio = pct / 100.0
        new_hi = full_hi * ratio if full_hi > 0 else full_hi
        new_lo = full_lo * ratio if full_lo < 0 else full_lo
        ax.set_ylim(new_lo if new_lo != 0 else -1, new_hi if new_hi != 0 else 1)
        if self._graph_canvas:
            self._graph_canvas.draw()   # force full redraw, not just idle

    def _plot_txn_amounts(self):
        """Plot transaction local amounts from parsed data for the selected currency."""
        ccy = self._graph_ccy_var.get()
        txns = self._parsed_txns
        if not txns:
            self._graph_status_var.set("No transaction data — convert a PDF first.")
            return
        import datetime
        try:
            import matplotlib
            matplotlib.use("TkAgg")
            import matplotlib.pyplot as plt
            import matplotlib.dates as mdates
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
            from matplotlib.figure import Figure
        except ImportError:
            self._graph_status_var.set("matplotlib not installed")
            return

        # Collect (date, local_amount) for matching currency transactions
        points = []
        for t in txns:
            if t.get('currency', 'IDR') != ccy and ccy != 'IDR':
                continue
            if ccy == 'IDR' and t.get('currency', 'IDR') != 'IDR':
                continue
            d = _transaction_to_dict(t)
            date_s = d.get('date', '')
            amt = d.get('local_amount') or d.get('currency_amount')
            if not date_s or amt is None or d.get('type') == 'OPENING':
                continue
            try:
                day, mon, yr = date_s.split('/')
                dt = datetime.date(int(yr), int(mon), int(day))
                points.append((dt, float(amt)))
            except Exception:
                continue

        if not points:
            self._graph_status_var.set(f"No {ccy} transactions found in loaded data.")
            return

        points.sort(key=lambda x: x[0])
        dates  = [p[0] for p in points]
        values = [p[1] for p in points]

        T = self.T
        bg     = T["bg"]
        fg     = T["fg"]
        card_bg= T["card_bg"]
        accent = T["accent"]
        grid_c = T["sep"]

        if self._graph_canvas:
            self._graph_canvas.get_tk_widget().destroy()
        if self._graph_toolbar:
            self._graph_toolbar.destroy()
        for w in self._graph_canvas_frame.winfo_children():
            w.destroy()

        fig = Figure(figsize=(5, 3.2), dpi=96)
        fig.patch.set_facecolor(bg)
        ax = fig.add_subplot(111)
        ax.set_facecolor(card_bg)

        # Bar chart: debit=red, credit=green
        colours = [T["credit_fg"] if v >= 0 else T["debit_fg"] for v in values]
        ax.bar(dates, values, color=colours, width=0.8, zorder=3)
        ax.axhline(0, color=grid_c, linewidth=0.8)

        ax.set_title(f"{ccy} Transaction Amounts (IDR)", color=fg,
                     fontsize=9, pad=6)
        ax.set_ylabel("IDR Amount", color=fg, fontsize=8)
        ax.tick_params(colors=fg, labelsize=7)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:,.0f}"))
        for spine in ax.spines.values():
            spine.set_edgecolor(grid_c)
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%d %b" if len(dates) <= 60 else "%b %y"))
        ax.xaxis.set_major_locator(mdates.AutoDateLocator())
        fig.autofmt_xdate(rotation=30, ha="right")
        ax.grid(True, color=grid_c, linewidth=0.5, linestyle="--", alpha=0.6, axis='y')
        fig.tight_layout(pad=1.4)

        canvas = FigureCanvasTkAgg(fig, master=self._graph_canvas_frame)
        canvas.draw()
        toolbar = NavigationToolbar2Tk(canvas, self._graph_canvas_frame)
        toolbar.update()
        toolbar.configure(bg=bg)
        canvas.get_tk_widget().pack(fill="both", expand=True)
        toolbar.pack(fill="x")
        self._graph_canvas  = canvas
        self._graph_toolbar = toolbar

        # Store for scale slider
        self._txn_chart_ax   = ax
        self._txn_chart_ylim = ax.get_ylim()
        # Reset slider to 100% on new plot
        self._scale_var.set(100)
        self._scale_lbl.configure(text="100%")

        total_cr = sum(v for v in values if v >= 0)
        total_db = sum(v for v in values if v < 0)
        self._graph_status_var.set(
            f"{len(values)} txns  ·  CR: {total_cr:,.0f}  ·  DB: {total_db:,.0f} IDR")

    def _fetch_graph_data(self):
        if self._graph_mode_var.get() == "amount":
            self._plot_txn_amounts()
            return
        if self._offline:
            self._graph_status_var.set("⚠ Offline mode — enable network to plot.")
            return
        import datetime, urllib.request
        from concurrent.futures import ThreadPoolExecutor, as_completed
        ccy    = self._graph_ccy_var.get()
        period = self._graph_period_var.get()

        days, step = {"7D":(7,1),"1M":(30,1),"3M":(90,3),
                      "6M":(180,7),"1Y":(365,7)}.get(period, (30,1))

        end_d   = datetime.date.today()
        start_d = end_d - datetime.timedelta(days=days)

        date_list = []
        d = start_d
        while d <= end_d:
            date_list.append(d)
            d += datetime.timedelta(days=step)
        if date_list[-1] != end_d:
            date_list.append(end_d)

        n = len(date_list)
        self._graph_status_var.set(f"Fetching {ccy}/IDR ({period}, {n} points)…")

        def worker():
            try:
                ccy_lower = ccy.lower()
                results   = {}   # date_str → idr_per_ccy
                completed = [0]

                def fetch_one(dt):
                    date_str = dt.strftime("%Y-%m-%d")
                    url = (f"https://cdn.jsdelivr.net/npm/@fawazahmed0/"
                           f"currency-api@{date_str}/v1/currencies/idr.json")
                    try:
                        with urllib.request.urlopen(url, timeout=10) as r:
                            data = json.loads(r.read())
                        inv = data.get("idr", {}).get(ccy_lower)
                        if inv and inv > 0:
                            return date_str, round(1.0 / inv, 4)
                    except Exception:
                        pass
                    return date_str, None

                with ThreadPoolExecutor(max_workers=12) as pool:
                    futures = {pool.submit(fetch_one, dt): dt for dt in date_list}
                    for fut in as_completed(futures):
                        date_str, rate = fut.result()
                        if rate is not None:
                            results[date_str] = rate
                        completed[0] += 1
                        pct = int(completed[0] / n * 100)
                        self.after(0, self._graph_status_var.set,
                                   f"Fetching {ccy}/IDR… {pct}%")

                dates  = sorted(results)
                values = [results[d] for d in dates]

                if len(dates) < 2:
                    raise ValueError(
                        f"Not enough data for {ccy}/IDR — currency may not be supported")

                self.after(0, self._draw_graph, ccy, dates, values, period)
                self._last_graph_data = (ccy, dates, values, period)

            except Exception as e:
                self.after(0, self._graph_status_var.set, f"Error: {e}")

        threading.Thread(target=worker, daemon=True).start()

    def _draw_graph(self, ccy, dates, values, period):
        import datetime
        T = self.T
        try:
            import matplotlib
            matplotlib.use("TkAgg")
            import matplotlib.pyplot as plt
            import matplotlib.dates as mdates
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
            from matplotlib.figure import Figure
        except ImportError:
            self._graph_status_var.set("matplotlib not installed — run: pip install matplotlib")
            return

        # Clear old canvas
        if self._graph_canvas:
            self._graph_canvas.get_tk_widget().destroy()
        if self._graph_toolbar:
            self._graph_toolbar.destroy()
        for w in self._graph_canvas_frame.winfo_children():
            w.destroy()

        # Dark/light colours
        bg     = T["bg"]
        fg     = T["fg"]
        accent = T["accent"]
        grid_c = T["sep"]
        card_bg= T["card_bg"]

        fig = Figure(figsize=(5, 3.2), dpi=96)
        fig.patch.set_facecolor(bg)
        ax = fig.add_subplot(111)
        ax.set_facecolor(card_bg)

        dt_dates = [datetime.datetime.strptime(d, "%Y-%m-%d") for d in dates]

        # Fill area under line
        ax.fill_between(dt_dates, values, alpha=0.15, color=accent)
        ax.plot(dt_dates, values, color=accent, linewidth=1.6, zorder=3)

        # Highlight min/max
        mn, mx = min(values), max(values)
        mn_i, mx_i = values.index(mn), values.index(mx)
        ax.scatter([dt_dates[mn_i]], [mn], color=T["debit_fg"],  s=40, zorder=5)
        ax.scatter([dt_dates[mx_i]], [mx], color=T["credit_fg"], s=40, zorder=5)
        ax.annotate(f"{mn:,.0f}", (dt_dates[mn_i], mn),
                    textcoords="offset points", xytext=(0,-14),
                    ha="center", fontsize=7, color=T["debit_fg"])
        ax.annotate(f"{mx:,.0f}", (dt_dates[mx_i], mx),
                    textcoords="offset points", xytext=(0,6),
                    ha="center", fontsize=7, color=T["credit_fg"])

        # Axes styling
        ax.set_title(f"{ccy} → IDR  ({period})", color=fg, fontsize=10, pad=8)
        ax.set_ylabel("IDR", color=fg, fontsize=8)
        ax.tick_params(colors=fg, labelsize=7)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v,_: f"{v:,.0f}"))
        for spine in ax.spines.values():
            spine.set_edgecolor(grid_c)
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%d %b" if len(dates)<=60 else "%b %y"))
        ax.xaxis.set_major_locator(mdates.AutoDateLocator())
        fig.autofmt_xdate(rotation=30, ha="right")
        ax.grid(True, color=grid_c, linewidth=0.5, linestyle="--", alpha=0.6)
        fig.tight_layout(pad=1.4)

        canvas = FigureCanvasTkAgg(fig, master=self._graph_canvas_frame)
        canvas.draw()
        toolbar = NavigationToolbar2Tk(canvas, self._graph_canvas_frame)
        toolbar.update()
        toolbar.configure(bg=bg)
        canvas.get_tk_widget().pack(fill="both", expand=True)
        toolbar.pack(fill="x")

        self._graph_canvas  = canvas
        self._graph_toolbar = toolbar

        last = values[-1]
        chg  = values[-1] - values[0]
        pct  = (chg / values[0] * 100) if values[0] else 0
        sign = "▲" if chg >= 0 else "▼"
        col  = T["credit_fg"] if chg >= 0 else T["debit_fg"]
        self._graph_status_var.set(
            f"Last: {last:,.2f} IDR  {sign} {abs(pct):.2f}% over {period}  "
            f"({len(dates)} data points)")

    # ── preview helpers ───────────────────────────────────────────────
    def _load_preview(self, transactions):
        """Populate the preview tab with parsed transactions."""
        self._parsed_txns = transactions          # no deepcopy — worker owns the data
        self._preview_cache = None                # invalidate search cache
        self._update_summary(transactions)
        self._refresh_type_menu(transactions)
        self._filter_preview()
        self._refresh_budget()
        self._nb.select(1)  # switch to Preview tab

    def _update_summary(self, txns):
        T = self.T
        total  = len([t for t in txns if t.get("type") != "OPENING"])
        debit  = sum(t.get("debit")  or 0 for t in txns)
        credit = sum(t.get("credit") or 0 for t in txns)
        net    = credit - debit
        accts  = len(set(t.get("account_no","") for t in txns))

        self._summary_cards["total"].configure(text=f"{total:,}")
        self._summary_cards["debit"].configure(
            text=f"{debit:,.0f}", fg=T["debit_fg"])
        self._summary_cards["credit"].configure(
            text=f"{credit:,.0f}", fg=T["credit_fg"])
        self._summary_cards["balance"].configure(
            text=f"{net:,.0f} IDR",
            fg=T["credit_fg"] if net >= 0 else T["debit_fg"])
        self._summary_cards["accounts"].configure(text=str(accts))

    def _refresh_type_menu(self, txns):
        types = sorted(set(t.get("type","") for t in txns if t.get("type")))
        self._type_menu["values"] = ["All"] + types
        self.filter_type.set("All")

    def _build_preview_cache(self):
        """Pre-compute search haystacks and DC flags once per load."""
        cache = []
        for t in self._parsed_txns:
            debit  = t.get("debit")
            credit = t.get("credit")
            dc = "DB" if debit is not None else ("CR" if credit is not None else "")
            # Build haystack from key visible fields only (fast, avoids internal _ keys)
            fields = (
                str(t.get("account_no",   "")),
                str(t.get("account_name", "")),
                str(t.get("date",         "")),
                str(t.get("keterangan",   "")),
                str(t.get("type",         "")),
                str(t.get("fasilitas",    "")),
                str(t.get("notes",        "")),
                str(t.get("currency",     "")),
            )
            cache.append((t, t.get("type",""), dc, " ".join(fields).upper()))
        self._preview_cache = cache

    def _filter_preview(self):
        search  = self.search_var.get().upper()
        f_type  = self.filter_type.get()
        f_dc    = self.filter_dc.get()

        # Build cache on first call or after reload
        if self._preview_cache is None:
            self._build_preview_cache()

        filtered = []
        for t, typ, dc, haystack in self._preview_cache:
            if f_type != "All" and typ != f_type:
                continue
            if f_dc != "All" and dc != f_dc:
                continue
            if search and search not in haystack:
                continue
            filtered.append(t)

        self._render_preview(filtered)
        self.preview_count.set(f"{len(filtered):,} rows")

    _PREVIEW_PAGE = 500   # rows rendered per chunk

    def _render_preview(self, txns):
        tree = self.preview_tree
        # Fast clear: delete all at once by IDs
        children = tree.get_children()
        if children:
            tree.delete(*children)

        active = [h for h in self._col_order if self._col_visible.get(h, True)]

        # Pre-build all row data first (no Tk calls in this loop)
        rows = []
        for i, t in enumerate(txns):
            d = _transaction_to_dict(t)
            c_amt  = f"{d['currency_amount']:,.2f}"   if d['currency_amount']   is not None else ""
            l_amt  = f"{d['local_amount']:,.2f}"      if d['local_amount']      is not None else ""
            c_bal  = f"{d['currency_balance']:,.2f}"  if d['currency_balance']  is not None else ""
            l_bal  = f"{d['local_balance']:,.2f}"     if d['local_balance']     is not None else ""
            fmt_map = {
                'seq_no': d['seq_no'], 'account_no': d['account_no'],
                'bank_name': d['bank_name'], 'account_name': d['account_name'],
                'date': d['date'], 'currency': d['currency'],
                'fx_rate_mode': d['fx_rate_mode'],
                'exchange_rate': f"{d['exchange_rate']:,.4f}" if d['exchange_rate'] is not None else '—',
                'fasilitas': d['fasilitas'], 'notes': d['notes'],
                'keterangan': d['keterangan'], 'type': d['type'], 'dc': d['dc'],
                'currency_amount': c_amt, 'local_amount': l_amt,
                'currency_balance': c_bal, 'local_balance': l_bal,
            }
            vals = tuple(fmt_map[h] for h in active)
            typ  = d["type"]
            dc   = d["dc"]
            if typ == "OPENING":
                tag = ("opening",)
            elif dc == "DB":
                tag = ("db", "alt") if i % 2 else ("db",)
            elif dc == "CR":
                tag = ("cr", "alt") if i % 2 else ("cr",)
            else:
                tag = ("alt",) if i % 2 else ()
            rows.append((vals, tag))

        # Insert in chunks so UI stays responsive between chunks
        def _insert_chunk(start):
            chunk = rows[start:start + self._PREVIEW_PAGE]
            for vals, tag in chunk:
                tree.insert("", "end", values=vals, tags=tag)
            nxt = start + self._PREVIEW_PAGE
            if nxt < len(rows):
                tree.after(1, _insert_chunk, nxt)   # yield to event loop then continue

        if rows:
            _insert_chunk(0)

    def _sort_preview(self, col):
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col
            self._sort_asc = True

        # Update heading arrows
        for c in _HEADERS:
            arrow = (" ▲" if self._sort_asc else " ▼") if c == col else ""
            lbl   = _HEADER_LABELS[_HEADERS.index(c)]
            self.preview_tree.heading(c, text=lbl + arrow)

        numeric = col in ("currency_amount", "local_amount", "currency_balance", "local_balance")
        def key(t):
            d = _transaction_to_dict(t)
            v = d.get(col)
            if numeric:
                try: return float(v or 0)
                except: return 0.0
            return str(v or "").lower()

        self._parsed_txns.sort(key=key, reverse=not self._sort_asc)
        self._filter_preview()

    def _clear_filter(self):
        self.search_var.set("")
        self.filter_type.set("All")
        self.filter_dc.set("All")
        self._filter_preview()

    # ── rules tree ───────────────────────────────────────────────────
    def _populate_tree(self):
        T = self.T
        self.tree.delete(*self.tree.get_children())
        bank_filter   = getattr(self, '_rules_bank_var',   None)
        search_filter = getattr(self, '_rules_search_var', None)
        sel_bank   = bank_filter.get()   if bank_filter   else "All Banks"
        search_kw  = search_filter.get().upper() if search_filter else ""

        shown = 0
        for i, r in enumerate(self.rules):
            rule_bank = r.get("bank", "All Banks")
            # Filter: "All Banks" filter shows everything;
            # specific bank filter shows that bank + All Banks rules
            if sel_bank != "All Banks" and rule_bank not in ("All Banks", sel_bank):
                continue
            if search_kw and search_kw not in r["keyword"].upper() \
                         and search_kw not in r["type"].upper():
                continue
            bank_label = "🌐 All Banks" if rule_bank == "All Banks" else rule_bank
            if r.get("locked"):
                tag = "locked"
            elif rule_bank == "All Banks":
                tag = "allbanks"
            else:
                tag = "even" if shown % 2 == 0 else "odd"
            self.tree.insert("", "end", iid=str(i),
                             values=(bank_label, r["keyword"], r["type"]), tags=(tag,))
            shown += 1

        self.tree.tag_configure("locked",   foreground=T["fg2"])
        self.tree.tag_configure("even",     background=T["tree_bg"])
        self.tree.tag_configure("odd",      background=T["tree_alt"])
        self.tree.tag_configure("allbanks", background=T["tree_bg"],
                                foreground=T["accent"])
        if hasattr(self, '_rules_count_var'):
            self._rules_count_var.set(f"{shown} / {len(self.rules)} rules")

    def _selected_index(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("No selection", "Please select a rule first.")
            return None
        return int(sel[0])

    # ── rule CRUD ────────────────────────────────────────────────────
    def _add_rule(self):
        self._rule_dialog("Add Rule", "", "", "BCA", callback=self._do_add)

    def _edit_rule(self):
        idx = self._selected_index()
        if idx is None: return
        r = self.rules[idx]
        if r.get("locked"):
            messagebox.showinfo("Locked", "This rule cannot be edited."); return
        self._rule_dialog("Edit Rule", r["keyword"], r["type"], r.get("bank","BCA"),
                          callback=lambda kw, tp, bk: self._do_edit(idx, kw, tp, bk))

    def _do_add(self, kw, tp, bk):
        self.rules.append({"keyword": kw, "type": tp, "bank": bk, "locked": False})
        save_rules(self.rules); self._populate_tree()

    def _do_edit(self, idx, kw, tp, bk):
        self.rules[idx].update({"keyword": kw, "type": tp, "bank": bk})
        save_rules(self.rules); self._populate_tree()

    def _delete_rule(self):
        idx = self._selected_index()
        if idx is None: return
        if self.rules[idx].get("locked"):
            messagebox.showinfo("Locked", "This rule cannot be deleted."); return
        if messagebox.askyesno("Delete", f"Delete rule '{self.rules[idx]['keyword']}'?"):
            self.rules.pop(idx); save_rules(self.rules); self._populate_tree()

    def _move_up(self):
        idx = self._selected_index()
        if idx is None or idx == 0: return
        self.rules[idx-1], self.rules[idx] = self.rules[idx], self.rules[idx-1]
        save_rules(self.rules); self._populate_tree()
        self.tree.selection_set(str(idx-1))

    def _move_down(self):
        idx = self._selected_index()
        if idx is None or idx >= len(self.rules)-1: return
        self.rules[idx+1], self.rules[idx] = self.rules[idx], self.rules[idx+1]
        save_rules(self.rules); self._populate_tree()
        self.tree.selection_set(str(idx+1))

    def _reset_rules(self):
        if messagebox.askyesno("Reset", "Reset all rules to default?"):
            self.rules = [dict(r) for r in DEFAULT_RULES]
            save_rules(self.rules); self._populate_tree()

    def _rule_dialog(self, title, keyword, type_label, bank, callback):
        T   = self.T
        dlg = tk.Toplevel(self)
        dlg.title(title); dlg.resizable(False, False)
        dlg.grab_set(); dlg.configure(bg=T["bg"])

        for row, lbl in enumerate(["Bank:", "Keyword (keterangan contains):", "Type label:"]):
            tk.Label(dlg, text=lbl, bg=T["bg"], fg=T["fg"],
                     font=("Arial", 10)).grid(row=row, column=0, sticky="w",
                                              padx=14, pady=(14,4) if row==0 else 4)

        bk_var = tk.StringVar(value=bank)
        kw_var = tk.StringVar(value=keyword)
        tp_var = tk.StringVar(value=type_label)

        # Bank dropdown — "All Banks" means rule applies to every bank
        bank_only = RULE_BANKS   # includes "All Banks" + specific banks
        ttk.Combobox(dlg, textvariable=bk_var, values=bank_only,
                     state="readonly", width=16,
                     font=("Arial", 10)).grid(row=0, column=1, sticky="w",
                                              padx=(0,14), pady=(14,4))

        for row, var in [(1, kw_var), (2, tp_var)]:
            tk.Entry(dlg, textvariable=var, width=30, font=("Arial", 10),
                     bg=T["entry_bg"], fg=T["entry_fg"], relief="flat",
                     insertbackground=T["fg"],
                     highlightthickness=1, highlightbackground=T["sep"],
                     highlightcolor=T["accent"]
                     ).grid(row=row, column=1, padx=(0,14), pady=4)

        def ok():
            bk = bk_var.get().strip()
            kw = kw_var.get().strip()
            tp = tp_var.get().strip()
            if not bk or not kw or not tp:
                messagebox.showwarning("Required", "All fields are required.", parent=dlg)
                return
            dlg.destroy(); callback(kw, tp, bk)

        br = tk.Frame(dlg, bg=T["bg"])
        br.grid(row=3, column=0, columnspan=2, pady=14)
        self._btn(br, "Save", ok, padx=16, pady=5).pack(side="left", padx=6)
        tk.Button(br, text="Cancel", command=dlg.destroy,
                  font=("Arial", 9), bg=T["bg2"], fg=T["fg"],
                  relief="flat", padx=12, pady=5).pack(side="left", padx=6)

        self.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  - 440) // 2
        y = self.winfo_y() + (self.winfo_height() - 210) // 2
        dlg.geometry(f"440x210+{x}+{y}")

    # ── date range toggle ─────────────────────────────────────────────
    def _on_date_range_toggle(self):
        state = "normal" if self._use_date_range.get() else "disabled"
        for w in (self._dr_from_e, self._dr_to_e):
            try: w.configure(state=state)
            except Exception: pass

    # ── drag and drop handler ─────────────────────────────────────────
    def _on_drop_pdf(self, event):
        """Handle files dropped onto the PDF listbox (requires tkinterdnd2)."""
        raw = event.data
        # tkinterdnd2 returns paths wrapped in braces if they contain spaces
        import re as _re
        paths = _re.findall(r'\{[^}]+\}|[^\s]+', raw)
        paths = [p.strip('{}') for p in paths]
        added = 0
        for p in paths:
            if p.lower().endswith('.pdf') and p not in self.pdf_listbox.get(0, "end"):
                self.pdf_listbox.insert("end", p)
                added += 1
        if added and not self.out_var.get():
            self.out_var.set(os.path.dirname(paths[0]))
        self.status_var.set(f"Added {added} PDF(s) via drag-and-drop.")

    # ── preview right-click ───────────────────────────────────────────
    def _show_preview_menu(self, event):
        try:
            iid = self.preview_tree.identify_row(event.y)
            if iid:
                if iid not in self.preview_tree.selection():
                    self.preview_tree.selection_set(iid)
                self._preview_menu_x = event.x
                self._preview_menu_col = self.preview_tree.identify_column(event.x)
                self._preview_menu.tk_popup(event.x_root, event.y_root)
        except Exception:
            pass

    def _copy_cell(self):
        sel = self.preview_tree.selection()
        if not sel: return
        col_id = getattr(self, '_preview_menu_col', '#1')
        col_idx = int(col_id.replace('#', '')) - 1
        vals = self.preview_tree.item(sel[0])['values']
        if 0 <= col_idx < len(vals):
            self.clipboard_clear()
            self.clipboard_append(str(vals[col_idx]))

    def _export_selection(self):
        sel = self.preview_tree.selection()
        if not sel:
            messagebox.showinfo("No selection", "Select rows first (Ctrl+click / Shift+click).")
            return
        out = self.out_var.get().strip()
        if not out:
            out = filedialog.askdirectory(title="Choose export folder")
            if not out: return
            self.out_var.set(out)

        # Collect transaction data for selected rows
        # Map iid → original transaction by matching values
        sel_iids = set(sel)
        sel_txns = []
        tree = self.preview_tree
        for iid in tree.get_children():
            if iid in sel_iids:
                vals = tree.item(iid)['values']
                # Reconstruct minimal dict from display values
                d = dict(zip(_HEADERS, vals))
                sel_txns.append(d)

        path = os.path.join(out, f"selection_{len(sel_txns)}_rows.xlsx")
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Selection"
            hdr_fill = PatternFill("solid", fgColor="2E75B6")
            hdr_font = Font(bold=True, color="FFFFFF")
            for ci, lbl in enumerate(_HEADER_LABELS, 1):
                c = ws.cell(1, ci, lbl)
                c.fill = hdr_fill; c.font = hdr_font
                c.alignment = Alignment(horizontal="center")
            for ri, row in enumerate(sel_txns, 2):
                for ci, key in enumerate(_HEADERS, 1):
                    ws.cell(ri, ci, row.get(key, ""))
            wb.save(path)
            messagebox.showinfo("Exported",
                f"Exported {len(sel_txns)} rows to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export error", str(e))

    # ── budget tab ────────────────────────────────────────────────────
    def _build_budget_tab(self):
        import datetime
        T = self.T
        f = self.tab_budget

        self._label(f, text="Monthly Budget per Type",
                    font=("Arial", 11, "bold"), fg_key="fg").pack(
                    anchor="w", padx=14, pady=(12, 2))
        self._label(f,
                    text="Set a monthly spending limit for each transaction type.  "
                         "Run a conversion first to populate types.",
                    font=("Arial", 8), fg_key="fg2").pack(anchor="w", padx=14, pady=(0,8))

        # Month selector
        ctrl = self._frame(f, bg_key="bg2")
        ctrl.pack(fill="x", padx=14, pady=(0,8))
        self._label(ctrl, text="Month:", font=("Arial", 9),
                    bg_key="bg2", fg_key="fg2").pack(side="left", padx=(10,4), pady=6)
        today = datetime.date.today()
        self._budget_month_var = tk.StringVar(
            value=today.strftime("%Y-%m"))
        month_e = tk.Entry(ctrl, textvariable=self._budget_month_var, width=10,
                           font=("Arial", 9), bg=T["entry_bg"], fg=T["entry_fg"],
                           relief="flat", insertbackground=T["fg"],
                           highlightthickness=1, highlightbackground=T["sep"])
        month_e.pack(side="left", padx=(0,8), pady=6)
        self._label(ctrl, text="(YYYY-MM)", font=("Arial", 8),
                    bg_key="bg2", fg_key="fg2").pack(side="left")
        self._btn(ctrl, "📊  Refresh", self._refresh_budget,
                  padx=10, pady=3).pack(side="left", padx=(12,0), pady=6)

        # Budget table
        bud_wrap = self._frame(f)
        bud_wrap.pack(fill="both", expand=True, padx=14, pady=(0,4))

        bcols = ("type", "budget", "actual", "remaining", "status")
        bhdrs = ("Type", "Monthly Budget (IDR)", "Actual Spent", "Remaining", "Status")
        self._budget_tree = ttk.Treeview(bud_wrap, columns=bcols,
                                          show="headings", selectmode="browse")
        for col, hdr, w in zip(bcols, bhdrs, [130,160,150,150,120]):
            self._budget_tree.heading(col, text=hdr)
            self._budget_tree.column(col, width=w, minwidth=60,
                                     anchor="w" if col=="type" else "e")
        self._budget_tree.tag_configure("over",    foreground=T["debit_fg"])
        self._budget_tree.tag_configure("ok",      foreground=T["credit_fg"])
        self._budget_tree.tag_configure("nobudget",foreground=T["fg2"])
        self._budget_tree.tag_configure("alt",     background=T["tree_alt"])

        sb_b = ttk.Scrollbar(bud_wrap, command=self._budget_tree.yview)
        self._budget_tree.configure(yscrollcommand=sb_b.set)
        self._budget_tree.pack(side="left", fill="both", expand=True)
        sb_b.pack(side="right", fill="y")

        # Double-click to set budget
        self._budget_tree.bind("<Double-1>", self._edit_budget_row)

        # Edit row
        edit_row = self._frame(f)
        edit_row.pack(fill="x", padx=14, pady=(4,10))
        self._label(edit_row, text="Double-click a row to set its budget limit  ·  "
                    "Enter 0 to clear", font=("Arial", 7), fg_key="fg2").pack(side="left")
        self._btn(edit_row, "💾  Save Budgets", self._save_budgets,
                  padx=10, pady=3).pack(side="right")

    def _refresh_budget(self):
        import datetime
        tree = self._budget_tree
        tree.delete(*tree.get_children())
        if not self._parsed_txns:
            self._label(self.tab_budget,
                        text="No data — convert a PDF first.",
                        font=("Arial", 9), fg_key="fg2")
            return

        month_str = self._budget_month_var.get().strip()   # "YYYY-MM"
        try:
            y, m = int(month_str[:4]), int(month_str[5:7])
        except Exception:
            messagebox.showwarning("Invalid month", "Use YYYY-MM format.")
            return

        # Aggregate actual spending by type for the chosen month
        actual = {}
        for t in self._parsed_txns:
            date = t.get("date", "")
            try:
                d, mo, yr = date.split("/")
                if int(yr) != y or int(mo) != m: continue
            except Exception:
                continue
            tp    = t.get("type", "OTHER")
            debit = t.get("debit") or 0
            actual[tp] = actual.get(tp, 0) + debit

        # Build rows: every known type (with spending) + any budgeted type
        all_types = sorted(set(list(actual.keys()) + list(self._budgets.keys())))
        for i, tp in enumerate(all_types):
            budget  = self._budgets.get(tp, 0)
            spent   = actual.get(tp, 0)
            remain  = budget - spent if budget else None
            if budget == 0:
                status = "—  no limit"
                tag    = "nobudget"
            elif spent > budget:
                over   = spent - budget
                status = f"▲ Over by {over:,.0f}"
                tag    = "over"
            else:
                pct    = int(spent / budget * 100) if budget else 0
                status = f"✔ {pct}% used"
                tag    = "ok"
            if i % 2: tag = (tag, "alt")

            tree.insert("", "end", iid=tp, tags=(tag,),
                        values=(tp,
                                f"{budget:,.0f}" if budget else "—",
                                f"{spent:,.0f}",
                                f"{remain:,.0f}" if remain is not None else "—",
                                status))

    def _edit_budget_row(self, event):
        T   = self.T
        sel = self._budget_tree.selection()
        if not sel: return
        tp      = sel[0]   # iid == type label
        current = self._budgets.get(tp, 0)

        dlg = tk.Toplevel(self)
        dlg.title(f"Budget: {tp}")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.configure(bg=T["bg"])

        tk.Label(dlg, text=f"Monthly limit for  {tp}  (IDR):",
                 bg=T["bg"], fg=T["fg"],
                 font=("Arial", 10)).grid(row=0, column=0, padx=14, pady=(16,4), sticky="w")
        amt_var = tk.StringVar(value=str(current))
        tk.Entry(dlg, textvariable=amt_var, width=20, font=("Arial", 10),
                 bg=T["entry_bg"], fg=T["entry_fg"], relief="flat",
                 insertbackground=T["fg"],
                 highlightthickness=1, highlightbackground=T["sep"],
                 highlightcolor=T["accent"]).grid(row=0, column=1, padx=(0,14), pady=(16,4))

        def ok():
            try:
                v = float(amt_var.get().replace(",", ""))
                if v == 0:
                    self._budgets.pop(tp, None)
                else:
                    self._budgets[tp] = v
                dlg.destroy()
                self._refresh_budget()
            except ValueError:
                messagebox.showwarning("Invalid", "Enter a number.", parent=dlg)

        br = tk.Frame(dlg, bg=T["bg"])
        br.grid(row=1, column=0, columnspan=2, pady=12)
        self._btn(br, "Save", ok, padx=14, pady=4).pack(side="left", padx=6)
        tk.Button(br, text="Cancel", command=dlg.destroy,
                  font=("Arial", 9), bg=T["bg2"], fg=T["fg"],
                  relief="flat", padx=10, pady=4).pack(side="left", padx=6)

        self.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  - 420)//2
        y = self.winfo_y() + (self.winfo_height() - 140)//2
        dlg.geometry(f"420x140+{x}+{y}")

    def _save_budgets(self):
        try:
            cfg = json.load(open(CONFIG_PATH)) if os.path.exists(CONFIG_PATH) else {}
        except Exception:
            cfg = {}
        cfg["budgets"] = self._budgets
        try:
            json.dump(cfg, open(CONFIG_PATH, "w"), indent=2)
            self.status_var.set("Budgets saved.")
        except Exception as e:
            messagebox.showerror("Save error", str(e))

    # ── file pickers ──────────────────────────────────────────────────
    def _ask_password(self, filename):
        """Show a modal password dialog. Returns password string or None if cancelled."""
        T    = self.T
        dlg  = tk.Toplevel(self)
        dlg.title("PDF Password Required")
        dlg.resizable(False, False)
        dlg.configure(bg=T["bg"])
        dlg.grab_set()

        # Centre over main window
        self.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  - 360) // 2
        y = self.winfo_y() + (self.winfo_height() - 180) // 2
        dlg.geometry(f"360x200+{x}+{y}")

        tk.Label(dlg, text="🔒  Password Required",
                 font=("Arial", 11, "bold"),
                 bg=T["bg"], fg=T["fg"]).pack(padx=20, pady=(18, 4))
        tk.Label(dlg, text=f'"{filename}"',
                 font=("Arial", 8), bg=T["bg"],
                 fg=T["fg2"], wraplength=320).pack(padx=20)
        tk.Label(dlg, text="💡 BCA: Date of Birth (DDMMYYYY)",
                 font=("Arial", 8), bg=T["bg"],
                 fg=T["fg2"]).pack(pady=(2, 8))

        pwd_var = tk.StringVar()
        row = tk.Frame(dlg, bg=T["bg"]); row.pack(fill="x", padx=20)
        entry = tk.Entry(row, textvariable=pwd_var, show="*",
                         font=("Consolas", 10),
                         bg=T["entry_bg"], fg=T["entry_fg"], relief="flat",
                         insertbackground=T["fg"],
                         highlightthickness=1, highlightbackground=T["sep"],
                         highlightcolor=T["accent"])
        entry.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self._btn(row, "👁", lambda: entry.configure(
                      show="" if entry.cget("show") else "*"),
                  bg_key="bg2", fg_key="fg2", padx=6, pady=3).pack(side="right")

        result = [None]

        def _ok(_=None):
            result[0] = pwd_var.get()
            dlg.destroy()

        def _cancel():
            result[0] = None
            dlg.destroy()

        btn_row = tk.Frame(dlg, bg=T["bg"]); btn_row.pack(pady=12)
        self._btn(btn_row, "✔  Unlock & Continue", _ok,
                  bg_key="success", font=("Arial", 9, "bold"),
                  padx=12, pady=5).pack(side="left", padx=4)
        self._btn(btn_row, "✕  Skip this file", _cancel,
                  bg_key="danger", padx=10, pady=5).pack(side="left", padx=4)

        entry.focus_set()
        entry.bind("<Return>", _ok)
        entry.bind("<Escape>", lambda _: _cancel())
        dlg.wait_window()
        return result[0]

    def _browse_pdf(self):
        bank = self._bank_map.get(self._bank_var.get(), "Bank")
        paths = filedialog.askopenfilenames(
            title=f"Select {bank} Statement PDFs",
            filetypes=[("PDF files", "*.pdf")])
        for p in paths:
            if p not in self.pdf_listbox.get(0, "end"):
                self.pdf_listbox.insert("end", p)
        if paths and not self.out_var.get():
            folder = os.path.dirname(paths[0])
            self.out_var.set(folder)
            self._persist_last_folder(folder)

    def _remove_pdf(self):
        for idx in reversed(self.pdf_listbox.curselection()):
            self.pdf_listbox.delete(idx)

    def _clear_pdf(self):
        self.pdf_listbox.delete(0,"end")

    def _browse_out(self):
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            self.out_var.set(path)
            self._persist_last_folder(path)

    def _persist_last_folder(self, folder):
        self._last_out_folder = folder
        try:
            cfg = json.load(open(CONFIG_PATH)) if os.path.exists(CONFIG_PATH) else {}
        except Exception:
            cfg = {}
        cfg["last_output_folder"] = folder
        try:
            json.dump(cfg, open(CONFIG_PATH, "w"), indent=2)
        except Exception:
            pass

    # ── log ───────────────────────────────────────────────────────────
    def _log(self, msg, tag=""):
        self.log.configure(state="normal")
        self.log.insert("end", msg+"\n", tag)
        self.log.see("end")
        self.log.configure(state="disabled")

    # ── convert ───────────────────────────────────────────────────────
    def _start_convert(self):
        import datetime as _dt
        pdfs  = list(self.pdf_listbox.get(0,"end"))
        out   = self.out_var.get().strip()
        merge = self.merge_var.get()
        fmt   = self.fmt_var.get()

        if not pdfs:
            messagebox.showwarning("No files","Please add at least one PDF."); return
        for p in pdfs:
            if not os.path.isfile(p):
                messagebox.showwarning("Missing",f"File not found:\n{p}"); return
        if len(pdfs) > 12 and not merge:
            if messagebox.askyesno(
                "Merge Required",
                f"You have {len(pdfs)} PDFs loaded.\n\n"
                f"Merge mode is required for more than 12 files to avoid\n"
                f"creating too many output files at once.\n\n"
                f"Enable Merge and continue?",
                icon="warning"
            ):
                self.merge_var.set(True)
                merge = True
            else:
                return
        if not out:
            out = os.path.dirname(pdfs[0])
            self.out_var.set(out)
            self._persist_last_folder(out)

        # Date range
        use_dr   = self._use_date_range.get()
        dr_from  = dr_to = None
        if use_dr:
            try:
                def _parse_dmy(s):
                    d, m, y = s.strip().split("/")
                    return _dt.date(int(y), int(m), int(d))
                dr_from = _parse_dmy(self._dr_from_var.get())
                dr_to   = _parse_dmy(self._dr_to_var.get())
                if dr_from > dr_to:
                    messagebox.showwarning("Date range", "'From' must be before 'To'.")
                    return
            except Exception:
                messagebox.showwarning("Date range",
                    "Invalid date format. Use DD/MM/YYYY.")
                return

        self.convert_btn.configure(state="disabled")
        self.progress["value"] = 0
        self.progress_pct.set("0%")
        self.progress_time.set("Estimating…")
        self.status_var.set("Converting…")

        import time as _time
        _start_time = [_time.monotonic()]   # mutable so worker closure can read it

        rules      = [dict(r) for r in self.rules]
        fx_key     = self.fx_key_var.get().strip()
        pdf_pwd    = [self.pdf_pwd_var.get()]   # list so inner loop can mutate it
        bank       = self._bank_map.get(self._bank_var.get(), "BCA")
        adapter    = BANK_ADAPTER_MAP.get(bank, BANK_ADAPTER_MAP["BCA"])
        n_files = len(pdfs)

        def _fmt_time(secs):
            secs = int(secs)
            if secs < 60:
                return f"{secs}s"
            return f"{secs//60}m {secs%60:02d}s"

        def _set_progress(pct):
            pct = max(0, min(100, int(pct)))
            elapsed = _time.monotonic() - _start_time[0]
            if pct > 2:
                total_est = elapsed / (pct / 100.0)
                remaining = max(0, total_est - elapsed)
                time_str  = f"Elapsed: {_fmt_time(elapsed)}  ·  ETA: {_fmt_time(remaining)}"
            else:
                time_str  = f"Elapsed: {_fmt_time(elapsed)}"
            self.after(0, lambda p=pct: self.progress.configure(value=p))
            self.after(0, self.progress_pct.set, f"{pct}%")
            self.after(0, self.progress_time.set, time_str)

        def _apply_date_filter(txns):
            if not use_dr:
                return txns
            out_txns = []
            for t in txns:
                date_s = t.get("date", "")
                try:
                    d, m, y = date_s.split("/")
                    td = _dt.date(int(y), int(m), int(d))
                    if dr_from <= td <= dr_to:
                        out_txns.append(t)
                except Exception:
                    out_txns.append(t)
            return out_txns

        def worker():
            try:
                self.after(0, self._log, f"Bank: {bank}  |  Adapter: {adapter.name}", "info")
                if not adapter.ready:
                    raise NotImplementedError(
                        f"{adapter.name} parser is not yet implemented.")
                orig = _this_module.infer_type
                _this_module.infer_type = lambda d: apply_rules(d, rules, bank)

                # Progress weights per file (must sum to 100):
                # parse=50, balance+seq=15, fx=20, save=15
                W_PARSE = 50; W_BAL = 15; W_FX = 20; W_SAVE = 15

                def file_progress(file_idx, stage_pct):
                    """Overall % = completed files + current file's stage fraction."""
                    per_file = 100.0 / n_files
                    done     = file_idx * per_file
                    current  = stage_pct / 100.0 * per_file
                    _set_progress(done + current)

                def save(txns, stem):
                    if fmt == "preview":
                        return []
                    saved = []
                    for key, ext, fn in FORMAT_MAP:
                        if fmt in (key, "all"):
                            p = os.path.join(out, stem + ext)
                            try:
                                fn(txns, p); saved.append(p)
                            except ImportError as ie:
                                self.after(0, self._log,
                                           f"  ⚠ Skipped {ext}: {ie}", "warn")
                    return saved

                def do_fx(txns, file_idx):
                    if self._offline:
                        for t in txns:
                            if t.get('currency', 'IDR') != 'IDR':
                                t['_fx_rate'] = None
                                t['_fx_rate_mode'] = 'Offline'
                                t['_local_amount'] = None
                                t['_local_balance'] = None
                        return

                    file_progress(file_idx, W_PARSE + W_BAL + (W_FX / 2))
                    def log_cb(msg, tag):
                        self.after(0, self._log, msg, tag)
                    apply_fx_rates(txns, fx_key, log_cb)

                # Process files
                all_txns = []
                for i, pdf_path in enumerate(pdfs):
                    stem = os.path.splitext(os.path.basename(pdf_path))[0]
                    self.after(0, self._log, f"▶ {stem}.pdf", "info")

                    success = False
                    attempts = 0
                    while not success and attempts < 3:
                        try:
                            txns = adapter.parse(pdf_path, pdf_password=pdf_pwd[0])
                            success = True
                            
                            txns = _apply_date_filter(txns)

                            if not txns:
                                self.after(0, self._log, "  ⚠ No transactions found/matched", "warn")
                            else:
                                self.after(0, self._log, f"  ✔ {len(txns)} transactions parsed", "ok")

                            file_progress(i, W_PARSE)

                            _calc_running_balance(txns)
                            _assign_seq_numbers(txns)
                            file_progress(i, W_PARSE + W_BAL)

                            do_fx(txns, i)
                            file_progress(i, W_PARSE + W_BAL + W_FX)

                            if merge:
                                all_txns.extend(txns)
                            else:
                                saved = save(txns, stem)
                                for p in saved:
                                    self.after(0, self._log, f"  💾 {os.path.basename(p)}")

                            file_progress(i, 100)

                        except Exception as e:
                            err_type = type(e).__name__
                            err_msg = str(e).lower()
                            
                            if "pdfminerexception" in err_type or "password" in err_msg or "incorrect" in err_msg or not str(e):
                                new_pwd = self._ask_password(os.path.basename(pdf_path))
                                if new_pwd == "SKIP" or new_pwd is None:
                                    self.after(0, self._log, f"  ⚠ Skipped password-protected file", "warn")
                                    break
                                else:
                                    pdf_pwd[0] = new_pwd
                                    attempts += 1
                            else:
                                self.after(0, self._log, f"  ✘ Error: {err_type}", "err")
                                for line in str(e).split('\n'):
                                    self.after(0, self._log, f"    {line}", "err")
                                break

                # End loop
                if merge and all_txns:
                    self.after(0, self.status_var.set, "Saving merged file…")
                    saved = save(all_txns, "BCA_Merged_Statements")
                    for p in saved:
                        self.after(0, self._log, f"💾 Merged: {os.path.basename(p)}", "ok")

                target_txns = all_txns if merge else (txns if 'txns' in locals() else [])
                self.after(0, self._load_preview, target_txns)
                
                self.after(0, self.progress.configure, {"value": 100})
                self.after(0, self.progress_pct.set, "100%")
                self.after(0, self.status_var.set, "✔ Conversion complete.")

            except Exception as e:
                self.after(0, self._log, f"Fatal error: {e}", "err")
                self.after(0, messagebox.showerror, "Error", str(e))
            finally:
                self.after(0, self.convert_btn.configure, {"state": "normal"})
                _this_module.infer_type = orig

        threading.Thread(target=worker, daemon=True).start()

    def _center(self):
        self.update_idletasks()
        w, h = 1100, 700
        x = (self.winfo_screenwidth()  - w)//2
        y = (self.winfo_screenheight() - h)//2
        self.geometry(f"{w}x{h}+{x}+{y}")


# ════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import multiprocessing
    multiprocessing.freeze_support()   # required for PyInstaller --onefile

    try:
        app = App()
        app.mainloop()
    except Exception as _exc:
        import traceback
        _tb = traceback.format_exc()
        try:
            import tkinter as _tk
            _root = _tk.Tk()
            _root.withdraw()
            from tkinter import messagebox as _mb
            _mb.showerror(
                "Startup Error",
                f"BCA Converter failed to start:\n\n{_exc}\n\n"
                f"See error_log.txt next to the .exe for details."
            )
            _root.destroy()
        except Exception:
            pass
        _log_path = os.path.join(
            os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else __file__),
            "error_log.txt"
        )
        with open(_log_path, "a") as _f:
            _f.write(_tb + "\n")
        sys.exit(1)